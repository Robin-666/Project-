import unittest,re
from test_page.ZT_information import ZT_ZM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log
from test_page.People_Assist_JM import People_Assist


class TestC2_01(unittest.TestCase):
    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.driver.maximize_window()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        # cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\辅助核算.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["辅助核算"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\辅助核算.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["辅助核算"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ZT_infor
        ZT_infor = ZT_ZM(self.driver, self.base_url, '')

    def People_FZHS(self):
        global PEOPLE
        PEOPLE = People_Assist(self.driver, self.base_url, '')

    def Log_In_People_Assist(self):
        '''从登陆界面进入人员辅助核算界面'''
        global page, Data ,PEOPLE
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        PEOPLE = People_Assist(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        self.driver.maximize_window()
        sleep(1)
        PEOPLE.Assist_Button()
        page.switch_frame(Data.myframe)
        PEOPLE.People_FZHS_button()
        PEOPLE.switch_frame(Data.people_frame)

    def Select_Button_Clear_All(self):
        self.People_FZHS()
        PEOPLE.Bing_state_OK()
        PEOPLE.All_Select_Button()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        page.OK_Msg()
        self.driver.implicitly_wait(10)
        page.OK_Msg()



    def test_A_People_Assist(self):
        self.Log_In_People_Assist()
        PEOPLE.Year_2019_people()
        sleep(1)
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.User_Name_Bind()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        Msg = PEOPLE.Msg_Frame()
        if "是否继续？？"in Msg:
            page.switch_frame_default()
            PEOPLE.Save_frame1()
            Msg2 = page.Dialog()
            if Msg2 == "成功自动绑定1条数据！":
                page.OK_Msg()
                PEOPLE.switch_frame(Data.myframe)
                PEOPLE.switch_frame(Data.people_frame)
                PEOPLE.Bing_state_OK()
                if PEOPLE.List_Display_two()==PEOPLE.List_Display_four():
                    self.WriteXlsx(2, 11, "PASS")
                    self.WriteXlsx(3, 11, "PASS")
                else:
                    self.WriteXlsx(2, 11, "ERROR")
                    self.WriteXlsx(3, 11, "ERROR")
                    do_log.error("报错：实际与预期不符")
            else:
                self.WriteXlsx(2, 11, "ERROR")
                self.WriteXlsx(3, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(2, 11, "ERROR")
            self.WriteXlsx(3, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(2, 10, "Y")
        self.WriteXlsx(3, 10, "Y")