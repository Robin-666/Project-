import unittest,re
from test_page.ZT_information import ZT_ZM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log


class TestC2_01(unittest.TestCase):

    def Log_In_DFKM(self):
        '''从登陆界面进入税金科目配置'''
        global page, Data
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        page.open()
        page.input_username(Data.name)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        page.Sbu_Config()
        page.switch_frame(Data.myframe)

    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)
    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ZT_infor
        ZT_infor = ZT_ZM(self.driver, self.base_url, '')

    def addIn_SJ_config(self):
        page.SJ_Button()
        page.SJ_ZT_All()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.SJ_Click_Row1()
        page.switch_frame_default()
        page.switch_frame(Data.frame2)

    #科目配置窗口账套查询_0530
    def test_A_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.addIn_SJ_config()
        self.Replace(531,8)
        ZT_infor.ZT_name()
        ZT_infor.Send_KEY_name(C["按账套名称查询"])
        ZT_infor.YS_Check_button()
        Lab1 = ZT_infor.Lab1_1()
        Lab2 = ZT_infor.Lab1()
        Lab3 = ZT_infor.Lab3()#平行记帐测试账套002
        if Lab1=="2019" and Lab2=="测试帐套001" and Lab3=="平行记帐测试账套002":
            self.WriteXlsx(531, 11, "PASS")
        else:
            self.WriteXlsx(531, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(531, 10, "Y")
        ZT_infor.ZT_name_clear()
        self.Replace(532,8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])
        ZT_infor.YS_Check_button()
        Lab4 = ZT_infor.Lab1()
        Lab5 = ZT_infor.Lab2()
        if Lab4 == Lab5 =="测试帐套001":
            self.WriteXlsx(532, 11, "PASS")
        else:
            self.WriteXlsx(532, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(532, 10, "Y")
        ZT_infor.ZT_name_clear()
        self.Replace(533, 8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])
        ZT_infor.YS_Check_button()
        Lab6 = ZT_infor.Lab4()
        if Lab6 =="":
            self.WriteXlsx(533, 11, "PASS")
        else:
            self.WriteXlsx(533, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(533, 10, "Y")
        ZT_infor.ZT_name_clear()
        ZT_infor.YS_Check_button()
        Lab7 = ZT_infor.Lab3()
        if Lab7 =="平行记帐测试账套002":
            self.WriteXlsx(534, 11, "PASS")
        else:
            self.WriteXlsx(534, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(534, 10, "Y")
        ZT_infor.ZT_name_clear()
        self.Replace(535, 8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])
        ZT_infor.YS_Check_button()
        Lab8 = ZT_infor.Lab4()
        if Lab8 == "":
            self.WriteXlsx(535, 11, "PASS")
        else:
            self.WriteXlsx(535, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(535, 10, "Y")
    #科目配置窗口账套查询_0535
    def test_B_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.addIn_SJ_config()
        self.Replace(536, 8)
        ZT_infor.ZT_Num()
        ZT_infor.Send_KEY(C["按账套号查询"])
        ZT_infor.YS_Check_button()
        Lab1 = ZT_infor.Lab1_1()
        Lab2 = ZT_infor.Lab1()
        Lab3 = ZT_infor.Lab3()  # 平行记帐测试账套002
        if Lab1 == "2019" and Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002":
            self.WriteXlsx(536, 11, "PASS")
        else:
            self.WriteXlsx(536, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(536, 10, "Y")
        ZT_infor.ZT_Num_clear()
        self.Replace(537, 8)
        ZT_infor.Send_KEY(C["按账套号查询"])
        ZT_infor.YS_Check_button()
        Lab4 = ZT_infor.Lab1()
        if Lab4 =="平行记帐测试账套002":
            self.WriteXlsx(537, 11, "PASS")
        else:
            self.WriteXlsx(537, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(537, 10, "Y")
        ZT_infor.ZT_Num_clear()
        self.Replace(538, 8)
        ZT_infor.Send_KEY(C["按账套号查询"])
        ZT_infor.YS_Check_button()
        Lab5 = ZT_infor.Lab4()
        if Lab5 == "":
            self.WriteXlsx(538, 11, "PASS")
        else:
            self.WriteXlsx(538, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(538, 10, "Y")
        ZT_infor.ZT_Num_clear()
        ZT_infor.YS_Check_button()
        Lab6 = ZT_infor.Lab1_1()
        Lab7 = ZT_infor.Lab1()
        Lab8 = ZT_infor.Lab3()  # 平行记帐测试账套002
        if Lab6 == "2019" and Lab7 == "测试帐套001" and Lab8 == "平行记帐测试账套002":
            self.WriteXlsx(539, 11, "PASS")
        else:
            self.WriteXlsx(539, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(539, 10, "Y")
        ZT_infor.ZT_Num_clear()
        self.Replace(540,8)
        ZT_infor.Send_KEY(C["按账套号查询"])
        ZT_infor.YS_Check_button()
        Lab9 = ZT_infor.Lab4()
        if Lab9 == "":
            self.WriteXlsx(540, 11, "PASS")
        else:
            self.WriteXlsx(540, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(540, 10, "Y")

    def test_C_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.addIn_SJ_config()
        self.Replace(541, 8)
        ZT_infor.ZT_year()
        ZT_infor.Send_KEY_year(C["按年度查询"])
        ZT_infor.YS_Check_button()
        Lab1 = ZT_infor.CW_XSYS()
        # self.assertEqual(Lab1,"条/共3条记录")   断言只能判断一次。
        if Lab1 == "条/共3条记录":
            self.WriteXlsx(541, 11, "PASS")
        else:
            self.WriteXlsx(541, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(541, 10, "Y")
        ZT_infor.ZT_year_clear()
        self.Replace(542, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])
        ZT_infor.YS_Check_button()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共1条记录":
            self.WriteXlsx(542, 11, "PASS")
        else:
            self.WriteXlsx(542, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(542, 10, "Y")
        ZT_infor.ZT_year_clear()
        self.Replace(543, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])
        ZT_infor.YS_Check_button()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共0条记录":
            self.WriteXlsx(543, 11, "PASS")
        else:
            self.WriteXlsx(543, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(543, 10, "Y")
        ZT_infor.ZT_year_clear()
        ZT_infor.YS_Check_button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共3条记录":
            self.WriteXlsx(544, 11, "PASS")
        else:
            self.WriteXlsx(544, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(544, 10, "Y")
        ZT_infor.ZT_year_clear()
        self.Replace(545, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])
        ZT_infor.YS_Check_button()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5 == "条/共0条记录":
            self.WriteXlsx(545, 11, "PASS")
        else:
            self.WriteXlsx(545, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(545, 10, "Y")

    def test_D_Subject_config_SJ(self):
        self.Log_In_DFKM()
        page.SJ_Button()
        page.SJ_ZT_All()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.JXS_Deve_button()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        # ZT_infor.switch_frame_default()
        page.OK_Msg()
        if Msg1 =="请选择账套信息!":
            self.WriteXlsx(550, 11, "PASS")
        else:
            self.WriteXlsx(550, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(550, 10, "Y")
        page.switch_frame(Data.frame1)
        ZT_infor.XXS_Deve_Input()
        ZT_infor.switch_frame_default()
        Msg2 = ZT_infor.Dialog()
        page.OK_Msg()
        if Msg2 == "请选择账套信息!":
            self.WriteXlsx(551, 11, "PASS")
        else:
            self.WriteXlsx(551, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(551, 10, "Y")
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg3 = ZT_infor.Dialog()
        page.OK_Msg()
        if Msg3 == "请选择账套信息!":
            self.WriteXlsx(552, 11, "PASS")
        else:
            self.WriteXlsx(552, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(552, 10, "Y")

    def AddIn_SJ_config_JXS(self):
        '''点击税金科目配置，账套选择001，新增进入科目配置窗口，点击进项税科目名称栏，弹出科目信息窗口'''
        page.SJ_Button()
        page.SJ_ZT_001()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.JXS_Deve_button()
        page.switch_frame_default()
        page.switch_frame(Data.frame2)


    def test_E_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_JXS()
        self.Replace(554,8)
        ZT_infor.Input_send_key(C["定位"])
        ZT_infor.search_Button()
        Lab1 = ZT_infor.CW_mytree_20_span()
        if Lab1 =="1101 交易性金融资产":
            self.WriteXlsx(554, 11, "PASS")
        else:
            self.WriteXlsx(554, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(554, 10, "Y")
        ZT_infor.search_Button()
        ZT_infor.search_Button()
        Lab2 = ZT_infor.CW_mytree_107_span()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab2 =="2221 应交税费" and Lab3=="条/共4条记录":
            self.WriteXlsx(555, 11, "PASS")
        else:
            self.WriteXlsx(555, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(555, 10, "Y")

    def test_F_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_JXS()
        ZT_infor.CW_mytree_1_span()
        Lab1 =ZT_infor.CW_XSYS()
        if Lab1 == "条/共188条记录":
            self.WriteXlsx(557, 11, "PASS")
        else:
            self.WriteXlsx(557, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(557, 10, "Y")
        ZT_infor.CW_mytree_2_span()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共90条记录":
            self.WriteXlsx(558, 11, "PASS")
        else:
            self.WriteXlsx(558, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(558, 10, "Y")
        ZT_infor.CW_mytree_93_span()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共44条记录":
            self.WriteXlsx(559, 11, "PASS")
        else:
            self.WriteXlsx(559, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(559, 10, "Y")
        ZT_infor.CW_mytree_138_span()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 =="条/共7条记录":
            self.WriteXlsx(560, 11, "PASS")
        else:
            self.WriteXlsx(560, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(560, 10, "Y")
        ZT_infor.CW_mytree_146_span()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5 =="条/共7条记录":
            self.WriteXlsx(561, 11, "PASS")
        else:
            self.WriteXlsx(561, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(561, 10, "Y")
        ZT_infor.CW_mytree_162_span()
        Lab6 = ZT_infor.CW_XSYS()
        if Lab6 == "条/共33条记录":
            self.WriteXlsx(562, 11, "PASS")
        else:
            self.WriteXlsx(562, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(562, 10, "Y")

    def test_G_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_JXS()
        self.Replace(565,8)
        ZT_infor.Check_Input_KMBM()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 =="条/共3条记录":
            self.WriteXlsx(565, 11, "PASS")
        else:
            self.WriteXlsx(565, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(565, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(566, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共12条记录":
            self.WriteXlsx(566, 11, "PASS")
        else:
            self.WriteXlsx(566, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(566, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(567, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共0条记录":
            self.WriteXlsx(567, 11, "PASS")
        else:
            self.WriteXlsx(567, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(567, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        ZT_infor.Check_Button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共188条记录":
            self.WriteXlsx(568, 11, "PASS")
        else:
            self.WriteXlsx(568, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(568, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(569, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5 == "条/共0条记录":
            self.WriteXlsx(569, 11, "PASS")
        else:
            self.WriteXlsx(569, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(569, 10, "Y")

    def test_H_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_JXS()
        self.Replace(570, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共18条记录":
            self.WriteXlsx(570, 11, "PASS")
        else:
            self.WriteXlsx(570, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(570, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(571, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共4条记录":
            self.WriteXlsx(571, 11, "PASS")
        else:
            self.WriteXlsx(571, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(571, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(572, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共0条记录":
            self.WriteXlsx(572, 11, "PASS")
        else:
            self.WriteXlsx(572, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(572, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        ZT_infor.Check_Button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共188条记录":
            self.WriteXlsx(573, 11, "PASS")
        else:
            self.WriteXlsx(573, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(573, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(574, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5 == "条/共0条记录":
            self.WriteXlsx(574, 11, "PASS")
        else:
            self.WriteXlsx(574, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(574, 10, "Y")

    def test_I_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_JXS()
        ZT_infor.CW_mytree_93_span()
        ZT_infor.Check_Input_KMBM()
        self.Replace(575,8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共7条记录":
            self.WriteXlsx(575, 11, "PASS")
        else:
            self.WriteXlsx(575, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(575, 10, "Y")
        ZT_infor.CW_mytree_138_span()
        ZT_infor.clear_KMBM_Input()
        self.Replace(576, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共2条记录":
            self.WriteXlsx(576, 11, "PASS")
        else:
            self.WriteXlsx(576, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(576, 10, "Y")
        ZT_infor.CW_mytree_162_span()
        ZT_infor.Check_Input_KMMC()
        self.Replace(577, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共11条记录":
            self.WriteXlsx(577, 11, "PASS")
        else:
            self.WriteXlsx(577, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(577, 10, "Y")
        ZT_infor.CW_mytree_2_span()
        ZT_infor.Clear_KMMC_Input()
        self.Replace(578, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共10条记录":
            self.WriteXlsx(578, 11, "PASS")
        else:
            self.WriteXlsx(578, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(578, 10, "Y")
    def test_J_Subject_config_SJ(self):
        self.Log_In_DFKM()
        page.SJ_Button()
        page.SJ_ZT_001()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        page.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg3 = ZT_infor.Dialog()
        page.OK_Msg()
        if Msg3 == "进项税科目信息和销项税科目信息不能同时为空!请选择...":
            self.WriteXlsx(582, 11, "PASS")
        else:
            self.WriteXlsx(582, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(582, 10, "Y")

    '''点击税金科目配置，账套选择001，新增进入科目配置窗口，点击销项税科目名称栏，弹出科目信息窗口'''
    def AddIn_SJ_config_XXS(self):
        page.SJ_Button()
        page.SJ_ZT_001()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.XXS_Deve_Input()
        page.switch_frame_default()
        page.switch_frame(Data.frame2)

    def test_K_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_XXS()
        self.Replace(584,8)
        ZT_infor.Input_send_key(C["定位"])
        ZT_infor.search_Button()
        Lab1 = ZT_infor.CW_mytree_20_span()
        if Lab1 == "1101 交易性金融资产":
            self.WriteXlsx(584, 11, "PASS")
        else:
            self.WriteXlsx(584, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(584, 10, "Y")
        ZT_infor.search_Button()
        ZT_infor.search_Button()
        Lab2 = ZT_infor.CW_mytree_107_span()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab2 == "2221 应交税费" and Lab3 == "条/共4条记录":
            self.WriteXlsx(585, 11, "PASS")
        else:
            self.WriteXlsx(585, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(585, 10, "Y")

    def test_L_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_XXS()
        ZT_infor.CW_mytree_1_span()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共188条记录":
            self.WriteXlsx(587, 11, "PASS")
        else:
            self.WriteXlsx(587, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(587, 10, "Y")
        ZT_infor.CW_mytree_2_span()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共90条记录":
            self.WriteXlsx(588, 11, "PASS")
        else:
            self.WriteXlsx(588, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(588, 10, "Y")
        ZT_infor.CW_mytree_93_span()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共44条记录":
            self.WriteXlsx(589, 11, "PASS")
        else:
            self.WriteXlsx(589, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(589, 10, "Y")
        ZT_infor.CW_mytree_138_span()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共7条记录":
            self.WriteXlsx(590, 11, "PASS")
        else:
            self.WriteXlsx(590, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(590, 10, "Y")
        ZT_infor.CW_mytree_146_span()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5 == "条/共7条记录":
            self.WriteXlsx(591, 11, "PASS")
        else:
            self.WriteXlsx(591, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(591, 10, "Y")
        ZT_infor.CW_mytree_162_span()
        Lab6 = ZT_infor.CW_XSYS()
        if Lab6 == "条/共33条记录":
            self.WriteXlsx(592, 11, "PASS")
        else:
            self.WriteXlsx(592, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(592, 10, "Y")

    def test_M_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_XXS()
        self.Replace(595, 8)
        ZT_infor.Check_Input_KMBM()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共3条记录":
            self.WriteXlsx(595, 11, "PASS")
        else:
            self.WriteXlsx(595, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(595, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(596, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共12条记录":
            self.WriteXlsx(596, 11, "PASS")
        else:
            self.WriteXlsx(596, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(596, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(597, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共0条记录":
            self.WriteXlsx(597, 11, "PASS")
        else:
            self.WriteXlsx(597, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(597, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        ZT_infor.Check_Button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共188条记录":
            self.WriteXlsx(598, 11, "PASS")
        else:
            self.WriteXlsx(598, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(598, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(599, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5 == "条/共0条记录":
            self.WriteXlsx(599, 11, "PASS")
        else:
            self.WriteXlsx(599, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(599, 10, "Y")

    def test_N_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_XXS()
        self.Replace(600, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共18条记录":
            self.WriteXlsx(600, 11, "PASS")
        else:
            self.WriteXlsx(600, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(600, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(601, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共4条记录":
            self.WriteXlsx(601, 11, "PASS")
        else:
            self.WriteXlsx(601, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(601, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(602, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共0条记录":
            self.WriteXlsx(602, 11, "PASS")
        else:
            self.WriteXlsx(602, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(602, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        ZT_infor.Check_Button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共188条记录":
            self.WriteXlsx(603, 11, "PASS")
        else:
            self.WriteXlsx(603, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(603, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(604, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5 == "条/共0条记录":
            self.WriteXlsx(604, 11, "PASS")
        else:
            self.WriteXlsx(604, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(604, 10, "Y")

    def test_O_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.AddIn_SJ_config_XXS()
        ZT_infor.CW_mytree_93_span()
        ZT_infor.Check_Input_KMBM()
        self.Replace(605, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共7条记录":
            self.WriteXlsx(605, 11, "PASS")
        else:
            self.WriteXlsx(605, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(605, 10, "Y")
        ZT_infor.CW_mytree_138_span()
        ZT_infor.clear_KMBM_Input()
        self.Replace(606, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共2条记录":
            self.WriteXlsx(606, 11, "PASS")
        else:
            self.WriteXlsx(606, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(606, 10, "Y")
        ZT_infor.CW_mytree_162_span()
        ZT_infor.Check_Input_KMMC()
        self.Replace(607, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共11条记录":
            self.WriteXlsx(607, 11, "PASS")
        else:
            self.WriteXlsx(607, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(607, 10, "Y")
        ZT_infor.CW_mytree_2_span()
        ZT_infor.Clear_KMMC_Input()
        self.Replace(608, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共10条记录":
            self.WriteXlsx(608, 11, "PASS")
        else:
            self.WriteXlsx(608, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(608, 10, "Y")

    def test_P_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2019_SJ()
        page.SJ_ZT_001()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        ZT_infor.JXS_Deve_button()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_mytree_93_span()
        ZT_infor.JXSE_SJ()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.XXS_Deve_Input()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()
        ZT_infor.XXSE_SJ()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg=="保存成功!":
            page.OK_Msg()
            ZT_infor.switch_frame(Data.myframe)
            Lab1 = ZT_infor.ZJM_Data_Display_1()
            Lab2 = ZT_infor.ZJM_Data_Display_2()
            if Lab1 =="22210101[应交税费-应交增值税-进项税额]" and Lab2=="22210102[应交税费-应交增值税-销项税额]":
                self.WriteXlsx(613, 11, "PASS")
            else:
                self.WriteXlsx(613, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
            self.WriteXlsx(613, 10, "Y")
        else:
            self.WriteXlsx(613, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(613, 10, "Y")

    def test_Q_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2019_SJ()
        page.SJ_ZT_001()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        ZT_infor.JXS_Deve_button()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_mytree_93_span()
        ZT_infor.JXSE_SJ()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.XXS_Deve_Input()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()
        ZT_infor.XXSE_SJ()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg =="此账套已绑定税金科目信息!":
            self.WriteXlsx(614, 11, "PASS")
        else:
            self.WriteXlsx(614, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(614, 10, "Y")

    def test_R_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2020_SJ()
        page.SJ_ZT_001()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        ZT_infor.JXS_Deve_button()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_mytree_93_span()
        ZT_infor.JXSE_SJ()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            ZT_infor.switch_frame_default()
            page.OK_Msg()
            ZT_infor.switch_frame(Data.myframe)
            Lab1 = ZT_infor.ZJM_Data_Display_1()
            if Lab1 == "22210101[应交税费-应交增值税-进项税额]" :
                self.WriteXlsx(615, 11, "PASS")
            else:
                self.WriteXlsx(615, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
            self.WriteXlsx(615, 10, "Y")

    def test_S_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2019_SJ()
        page.SJ_ZT_002()
        page.SJ_Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        ZT_infor.XXS_Deve_Input()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_JZC()
        ZT_infor.JZC_Dispaly()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            page.OK_Msg()
            ZT_infor.switch_frame(Data.myframe)
            Lab1 = ZT_infor.ZJM_Data_Display_2()
            if Lab1 == "320101[本年盈余-行政事业盈余]":
                self.WriteXlsx(616, 11, "PASS")
            else:
                self.WriteXlsx(616, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
            self.WriteXlsx(616, 10, "Y")
        else:
            self.WriteXlsx(616, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(616, 10, "Y")

    def test_T_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2019_SJ()
        page.SJ_ZT_001()
        ZT_infor.Lab_td_6()
        page.Change_Year_button()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row1()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.Select_Year_2020_ZT()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        if Msg=="此账套已绑定税金科目信息!":
            self.WriteXlsx(617, 11, "PASS")
        else:
            self.WriteXlsx(617, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(617, 10, "Y")

    def test_U_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2019_SJ()
        page.SJ_ZT_001()
        ZT_infor.Lab_td_6()
        page.Change_Year_button()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row1()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.ZT_002_TEST()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        if Msg == "此账套已绑定税金科目信息!":
            self.WriteXlsx(618, 11, "PASS")
        else:
            self.WriteXlsx(618, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(618, 10, "Y")

    def test_V_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2020_SJ()
        page.SJ_ZT_001()
        ZT_infor.Lab_td_6()
        ZT_infor.SJ_Delete_Button()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        if Msg=="数据删除后将无法恢复,请确认是否进行删除!":
            ZT_infor.switch_frame_default()
            page.OK_Msg()
            page.OK_Msg()
            ZT_infor.switch_frame(Data.myframe)
            self.WriteXlsx(619, 11, "PASS")
        else:
            self.WriteXlsx(619, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(619, 10, "Y")

    def test_W_Subject_config_SJ(self):
        self.Log_In_DFKM()
        self.Glo()
        page.SJ_Button()
        ZT_infor.Year_2019_SJ()
        page.SJ_ZT_002()
        ZT_infor.Lab_td_8()
        ZT_infor.SJ_Delete_Button()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            ZT_infor.switch_frame_default()
            page.OK_Msg()
            page.OK_Msg()
            ZT_infor.switch_frame(Data.myframe)
            self.WriteXlsx(620, 11, "PASS")
        else:
            self.WriteXlsx(620, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(620, 10, "Y")

    '''科目配置用例执行完成后，进行新增枚举带科目的配置（初始化后续的操作）'''
    def Log_In_DF_KMPZ(self):
        '''从登陆界面进入借方科目配置,2020,001,新增按钮'''
        global page, Data
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        page.open()
        page.input_username(Data.name)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        page.Sbu_Config()
        page.switch_frame(Data.myframe)
        page.JF_select_2020()
        page.JF_select_YWDJMJ_A()
        page.JF_ZT_and_001()
        page.Add_Button_JF()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)

    def test_X_Subject_config_SJ(self):
        self.Log_In_DF_KMPZ()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.JF_YJMJ_A()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.span_text_A()
        ZT_infor.Next_JM()
        ZT_infor.Select_1131()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        if page.Dialog()=="保存成功!":
            self.WriteXlsx(672, 11, "PASS")
        else:
            self.WriteXlsx(672, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(672, 10, "Y")
        page.OK_Msg()

    def test_Y_Subject_config_SJ(self):
        self.Log_In_DF_KMPZ()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.JF_YJMJ_DDDDD()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.span_text_A()
        ZT_infor.Next_JM()
        ZT_infor.Select_1146()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        if page.Dialog() == "保存成功!":
            self.WriteXlsx(673, 11, "PASS")
        else:
            self.WriteXlsx(673, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(673, 10, "Y")
        page.OK_Msg()

    def test_Z_Subject_config_SJ(self):
        self.Log_In_DF_KMPZ()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.JF_YJMJ_C()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.span_text_A()
        ZT_infor.Next_JM()
        ZT_infor.Select_1147()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        if page.Dialog() == "保存成功!":
            self.WriteXlsx(674, 11, "PASS")
        else:
            self.WriteXlsx(674, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(674, 10, "Y")
        page.OK_Msg()

    def test_Za_Subject_config_SJ(self):
        self.Log_In_DF_KMPZ()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.JF_YJMJ_B()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.span_text_A()
        ZT_infor.Next_JM()
        ZT_infor.Select_1148()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        if page.Dialog() == "保存成功!":
            self.WriteXlsx(675, 11, "PASS")
        else:
            self.WriteXlsx(675, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(675, 10, "Y")
        page.OK_Msg()

    def Log_In_DF_KMPZ_Button(self):
        '''从登陆界面进入贷方科目配置,2020,001,新增按钮'''
        global page, Data
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        page.open()
        page.input_username(Data.name)
        page.input_password(Data.pwd)
        page.click_submit()
        # self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        page.Sbu_Config()
        page.switch_frame(Data.myframe)
        page.DF_Button()
        page.select_2020()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)


    def test_Zb_Subject_config_SJ(self):
        self.Log_In_DF_KMPZ_Button()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.JF_YJMJ_DDDDD()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.span_text_A()
        ZT_infor.Select_100202()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        if page.Dialog() == "保存成功!":
            self.WriteXlsx(676, 11, "PASS")
        else:
            self.WriteXlsx(676, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(676, 10, "Y")
        page.OK_Msg()

    def test_Zc_Subject_config_SJ(self):
        self.Log_In_DF_KMPZ_Button()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.JF_YJMJ_B()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.span_text_A()
        js = "var q=document.documentElement.scrllTop=1000"
        self.driver.execute_script(js)
        ZT_infor.Select_1012()
        # ZT_infor.Next_JM()
        # ZT_infor.Select_1148()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        if page.Dialog() == "保存成功!":
            self.WriteXlsx(677, 11, "PASS")
        else:
            self.WriteXlsx(677, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(677, 10, "Y")
        page.OK_Msg()


    def test_Zd_Subject_config_SJ(self):
        self.Log_In_DF_KMPZ_Button()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.JF_YJMJ_C()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.span_text_A()
        ZT_infor.Next_JM()
        ZT_infor.Next_JM()
        ZT_infor.Select_1304()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        page.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        if page.Dialog() == "保存成功!":
            self.WriteXlsx(678, 11, "PASS")
        else:
            self.WriteXlsx(678, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(678, 10, "Y")
        page.OK_Msg()


if __name__ == '__main__':
    unittest.main(verbosity=2)