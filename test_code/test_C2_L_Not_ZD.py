import unittest,re
from test_page.ZT_information import ZT_ZM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log
from test_page.People_Assist_JM import People_Assist
from test_page.Wait_ZD_JM import Wait_ZD



class TestC2_01(unittest.TestCase):
    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\凭证制单_不制单.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["凭证制单_不制单"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\凭证制单_不制单.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["凭证制单_不制单"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ZT_infor
        ZT_infor = ZT_ZM(self.driver, self.base_url, '')

    def People_FZHS(self):
        global PEOPLE
        PEOPLE = People_Assist(self.driver, self.base_url, '')



    def Log_In_Wait_ZD(self):
        '''从登陆界面进入不制单界面'''
        global page, Data ,Wait_ZD_ZJM
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        Wait_ZD_ZJM = Wait_ZD(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        page.CWJC()
        page.Win_PZZD()
        page.Win_Split()
        Wait_ZD_ZJM.Not_ZD_JM()
        Wait_ZD_ZJM.switch_frame(Data.myframe)


    def test_A_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_New_LB_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        sleep(1)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(2, 11, "PASS")
        else:
            self.WriteXlsx(2, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(2, 10, "Y")
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_2020_ZT()
        self.Glo()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        sleep(1)
        if ZT_infor.CW_XSYS() == "条/共8条记录":
            self.WriteXlsx(3, 11, "PASS")
        else:
            self.WriteXlsx(3, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(3, 10, "Y")
        Wait_ZD_ZJM.Select_BD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Select_C2_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        sleep(1)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(4, 11, "PASS")
        else:
            self.WriteXlsx(4, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(4, 10, "Y")
        Wait_ZD_ZJM.Select_BD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Select_Robin_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        sleep(1)
        if ZT_infor.CW_XSYS() == "条/共7条记录":
            self.WriteXlsx(5, 11, "PASS")
        else:
            self.WriteXlsx(5, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(5, 10, "Y")
        Wait_ZD_ZJM.Select_Year_2019()
        sleep(1)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(6, 11, "PASS")
        else:
            self.WriteXlsx(6, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(6, 10, "Y")
        Wait_ZD_ZJM.Select_Year_2020()
        sleep(1)
        if ZT_infor.CW_XSYS() == "条/共7条记录":
            self.WriteXlsx(7, 11, "PASS")
        else:
            self.WriteXlsx(7, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(7, 10, "Y")

    def test_B_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_2020_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Select_BD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_Robin_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Not_ZD_MX()
        Wait_ZD_ZJM.switch_frame_default()
        if page.Dialog()=="请选择需要查看明细的表单!":
            self.WriteXlsx(9, 11, "PASS")
        else:
            self.WriteXlsx(9, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(9, 10, "Y")
        Wait_ZD_ZJM.switch_frame_default()
        page.OK_Msg()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.List_1()
        Wait_ZD_ZJM.Not_ZD_MX()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        if Wait_ZD_ZJM.List_Display_4()=="不制单" and Wait_ZD_ZJM.List_Display_5() !="":
            self.WriteXlsx(8, 11, "PASS")
        else:
            self.WriteXlsx(8, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(8, 10, "Y")

    def test_C_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_111111_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Select_BD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_Robin_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Clear_Select_people_input()
        Wait_ZD_ZJM.Select_people_button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Select_CXB_people()
        Wait_ZD_ZJM.Select_Right_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共4条记录":
            self.WriteXlsx(10, 11, "PASS")
        else:
            self.WriteXlsx(10, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(10, 10, "Y")

    def test_D_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_111111_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Select_BD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_Robin_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Clear_Select_people_input()
        Wait_ZD_ZJM.Select_people_button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Select_CXA_1()
        Wait_ZD_ZJM.Select_CXA_people()
        Wait_ZD_ZJM.Select_Right_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共3条记录":
            self.WriteXlsx(11, 11, "PASS")
        else:
            self.WriteXlsx(11, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(11, 10, "Y")

    def test_E_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_111111_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Select_BD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_Robin_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Clear_Select_people_input()
        self.Replace(12, 8)
        Wait_ZD_ZJM.Advanced_Query_Input(C["标题"])
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(12, 11, "PASS")
        else:
            self.WriteXlsx(12, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(12, 10, "Y")
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Clear_Advanced_Query_Input()
        self.Replace(13, 8)
        Wait_ZD_ZJM.Advanced_Query_Input(C["标题"])
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(13, 11, "PASS")
        else:
            self.WriteXlsx(13, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(13, 10, "Y")

    def test_F_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_U8_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Select_BD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_R2_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Select_Year_2020()
        Wait_ZD_ZJM.Month_3()
        Wait_ZD_ZJM.List_1()
        Wait_ZD_ZJM.Re_ZD_Button()
        Wait_ZD_ZJM.switch_frame_default()
        if page.Dialog()=="恢复制单操作成功!":
            self.WriteXlsx(14, 11, "PASS")
        else:
            self.WriteXlsx(14, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(14, 10, "Y")





















if __name__ == '__main__':
    unittest.main(verbosity=2)




















