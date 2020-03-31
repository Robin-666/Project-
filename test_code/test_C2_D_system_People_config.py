import unittest,re
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.support.select import Select

class TestC2_02(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.33:8081"
        print("Test Start")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\人员对照.xlsx"
        # Excel_path = '../test_case\人员对照.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["Sheet1"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\人员对照.xlsx"
        # Excel_path = '../test_case\人员对照.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["Sheet1"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def System_config(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxa")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        # self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                # print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
        self.driver.find_element_by_xpath("//div[@title = '人员对照']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        # self.driver.switch_to.frame("templateSysMgr")  # 切换到人员对照页面中
    #按用户名自动绑定-正常
    def test_a_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@id='north']/a[1]").click()#按用户名自动绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        self.driver.switch_to.default_content()
        if msg1 == "当前选择【单位A】将会按照【单位A】下OA用户名与ERP用户名一致，进行自动绑定；请确认已绑定人员是否更新?" and msg2=="成功自动绑定6条数据！":
            self.WriteXlsx(2, 11, "Pass")
        else:
            self.WriteXlsx(2, 11, "Error")
        self.WriteXlsx(2, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
    #按用户名自动绑定-重复绑定
    def test_b_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@id='north']/a[1]").click()  # 按用户名自动绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
        sleep(3)
        self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//div[@id='north']/a[1]").click()  # 按用户名自动绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg3 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg3)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg4 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg4)
        self.driver.switch_to.default_content()
        if msg1 == msg3 =="当前选择【单位A】将会按照【单位A】下OA用户名与ERP用户名一致，进行自动绑定；请确认已绑定人员是否更新?":
            if msg2 == msg4 =="成功自动绑定6条数据！":
                self.WriteXlsx(4, 11, "Pass")
            else:
                self.WriteXlsx(4, 11, "Error")
        self.WriteXlsx(4, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
    #一对一手动绑定-预带最近账套-是
    def test_c_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td/div/input").click()#选择副书记
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_center']/div[1]/div[5]/table/tbody/tr[11]/td/div/input").click()#选择四川演示
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()#保存绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1) #用户绑定成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
        sleep(1)
        self.driver.switch_to.frame('myiframe')
        msg2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[6]/div").text
        print(msg2)
        if msg1 =='用户绑定成功!' and msg2 =='是':
            self.WriteXlsx(5, 11, "Pass")
        else:
            self.WriteXlsx(5, 11, "Error")
        self.WriteXlsx(5, 10, "Y")
    #一对一手动绑定-预带最近账套-否
    def test_d_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td/div/input").click()#选择OA-办公室人员B
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@id='east_center']/div[1]/div[5]/table/tbody/tr[10]/td/div/input").click()  # 选择ERP-都满
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/div/input").click()  # 预带最近账套-否
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()  # 保存绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 用户绑定成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
        sleep(1)
        self.driver.switch_to.frame('myiframe')
        msg2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[6]/div").text
        print(msg2)
        if msg1 =='用户绑定成功!' and msg2 =='否':
            self.WriteXlsx(6,11, "Pass")
        else:
            self.WriteXlsx(6, 11, "Error")
        self.WriteXlsx(6, 10, "Y")
    def Del(self):
        self.driver.find_element_by_xpath("//div[@id='north']/a[3]").click()  # 全部清除
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='autobtn']").click()#自动绑定
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()


    #多对一手动绑定-预带最近账套-是
    def test_e_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td/div/input").click()#财务人员A
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[9]/td/div/input").click()  # 综合科人员A
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[13]/td/div/input").click()  # 监察室人员A
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@id='east_center']/div[1]/div[5]/table/tbody/tr[11]/td/div/input").click()  # 选择四川演示
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()  # 保存绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 当前选择了多个协同用户绑定一个ERP用户，是否继续？
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(2)
        # self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)#	用户绑定成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
        sleep(1)
        self.driver.switch_to.frame('myiframe')
        sleep(1)
        msg3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[4]/div").text  # 查看财务人员A对应的ERP账套
        msg4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[9]/td[4]/div").text  #
        msg5 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[14]/td[4]/div").text
        msg6 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[6]/div").text  #预带最近账套-是
        msg7 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[9]/td[6]/div").text
        msg8 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[14]/td[6]/div").text
        print("msg3~msg4:",msg3,msg4)
        print("msg5:",msg5)
        print("msg6~msg7:", msg6, msg7)
        print('msg8',msg8)
        if msg1 == '当前选择了多个协同用户绑定一个ERP用户，是否继续？':
            if msg2 == '用户绑定成功!':
                if msg3  == msg5 =="测试用户" and msg6  == msg8 == "是":
                    self.WriteXlsx(7, 11, "Pass")
                else:
                    self.WriteXlsx(7, 11, "Error")
            else:
                self.WriteXlsx(7, 11, "Error")
        else:
            self.WriteXlsx(7, 11, "Error")
        self.WriteXlsx(7, 10, "Y")
        # self.Del()
    # 多对一手动绑定-预带最近账套-否
    def test_f_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td/div/input").click()  # 副书记
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[11]/td/div/input").click() #综合人员B
        sleep(1)
        # js = "var q=document.documentElement.scrollTop=300"
        # self.driver.execute_script(js)
        # print("操作滚动条")
        # sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[15]/td/div/input").click() #监察室人员B
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@id='east_center']/div[1]/div[5]/table/tbody/tr[11]/td/div/input").click()  # 选择四川演示
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/div/input").click()#预带最近账套-否
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()#保存绑定
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 当前选择了多个协同用户绑定一个ERP用户，是否继续？
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(2)
        # self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)  # 用户绑定成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
        sleep(1)
        self.driver.switch_to.frame('myiframe')
        sleep(1)
        msg3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[6]/div").text  # 查看财务人员A对应的ERP账套
        msg4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[11]/td[6]/div").text
        msg5 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[15]/td[6]/div").text
        msg6 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 查看ERP名称
        msg7 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[11]/td[4]/div").text
        msg8 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[15]/td[4]/div").text
        print("msg3~msg4:", msg3, msg4)
        print("msg5:",msg5)
        print("msg6~msg7:", msg6, msg7)
        print("msg8:", msg8)
        if msg1 == '当前选择了多个协同用户绑定一个ERP用户，是否继续？':
            if msg2 == '用户绑定成功!':
                if msg3 == msg4 == "否" and msg6 == msg7 == "测试用户":
                    self.WriteXlsx(8, 11, "Pass")
                else:
                    self.WriteXlsx(8, 11, "Error")
            else:
                self.WriteXlsx(8, 11, "Error")
        else:
            self.WriteXlsx(8, 11, "Error")
        self.WriteXlsx(8, 10, "Y")
        self.Del()
    #未选择OA/ERP人员进行保存绑定
    def test_g_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()  # 保存绑定
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 请选择至少一个协同用户进行绑定!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if msg1 == '请选择至少一个协同用户进行绑定!':
            self.WriteXlsx(10, 11, "Pass")
        else:
            self.WriteXlsx(10, 11, "Error")
        self.WriteXlsx(10, 10, "Y")
    #只选择OA人员进行保存绑定
    def test_h_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td/div/input").click()  # 副书记
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()  # 保存绑定
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 	是否要更新‘预带最近账套’？
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        self.driver.switch_to.default_content()
        sleep(1)
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)#请选择已绑定的协同用户进行更新!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if msg1 == "是否要更新‘预带最近账套’？" and msg2 =='请选择已绑定的协同用户进行更新!':
            self.WriteXlsx(11, 11, "Pass")
        else:
            self.WriteXlsx(11, 11, "Error")
        self.WriteXlsx(11, 10, "Y")
        self.driver.switch_to.frame("myiframe")
        self.Del()
    #只选择ERP人员进行保存绑定
    def test_i_People_config(self):
        self.System_config()
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@id='east_center']/div[1]/div[5]/table/tbody/tr[11]/td/div/input").click()  # 选择四川演示
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()  # 保存绑定
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 请选择至少一个协同用户进行绑定!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if msg1 =='请选择至少一个协同用户进行绑定!':
            self.WriteXlsx(12, 11, "Pass")
        else:
            self.WriteXlsx(12, 11, "Error")
        self.WriteXlsx(12, 10, "Y")
    #直接点击清除绑定
    def test_j_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@id='north']/a[2]").click()#清除绑定按钮
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 请至少选择一条数据进行清除绑定!
        if msg1 =='请至少选择一条数据进行清除绑定!':
            self.WriteXlsx(13, 11, "Pass")
        else:
            self.WriteXlsx(13, 11, "Error")
        self.WriteXlsx(13, 10, "Y")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
    #选择OA用户名进行清除绑定
    def test_k_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td/div/input").click()  # 副书记
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='north']/a[2]").click()  # 清除绑定按钮
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 是否要清除当前所选的人员？
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        self.driver.switch_to.default_content()
        sleep(1)
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)  # 清除数据成功！请重新对照人员!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if msg1 == '是否要清除当前所选的人员？' and  msg2 =="清除数据成功！请重新对照人员!":
            self.WriteXlsx(14, 11, "Pass")
        else:
            self.WriteXlsx(14, 11, "Error")
        self.WriteXlsx(14, 10, "Y")
    #OA用户名与ERP用户名绑定后进行清除绑定
    def test_l_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td/div/input").click()  # 办公室人员B
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@id='east_center']/div[1]/div[5]/table/tbody/tr[7]/td/div/input").click()  # 选择李响
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()  # 保存绑定
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 	用户绑定成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame('myiframe')
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td/div/input").click()  # 办公室人员B
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='north']/a[2]").click()  # 清除绑定按钮
        sleep(1)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)  # 是否要清除当前所选的人员？
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        self.driver.switch_to.default_content()
        sleep(1)
        msg3 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg3)  # 清除数据成功！请重新对照人员!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame('myiframe')
        sleep(1)
        msg4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[4]/div").text
        print('msg4:',msg4)
        if msg1 =='用户绑定成功!' and msg2=='是否要清除当前所选的人员？' and msg3 =='清除数据成功！请重新对照人员!':
            if msg4 == ' ':
                self.WriteXlsx(15, 11, "Pass")
            else:
                self.WriteXlsx(15, 11, "Error")
        self.WriteXlsx(15, 10, "Y")
    #直接点击全部清除
    def test_m_People_config(self):
        self.System_config()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='north']/a[3]").click()  # 清除绑定按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)#将清空所有绑定的人员信息，是否继续
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)  # 清空成功
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if msg1 == '将清空所有绑定的人员信息，是否继续' and msg2 == '清空成功':
            self.WriteXlsx(16,11, "Pass")
        else:
            self.WriteXlsx(16, 11, "Error")
        self.WriteXlsx(16, 10, "Y")
    # OA用户名与ERP用户名绑定后进行全部清除
    def test_n_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td/div/input").click()  # 办公室人员B
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@id='east_center']/div[1]/div[5]/table/tbody/tr[7]/td/div/input").click()  # 选择李响
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()  # 保存绑定
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 用户绑定成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame('myiframe')
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td/div/input").click()  # 办公室人员B
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='north']/a[3]").click()  # 全部清除按钮
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)  # 将清空所有绑定的人员信息，是否继续
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        self.driver.switch_to.default_content()
        sleep(1)
        msg3 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg3)#清空成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if msg1 == '用户绑定成功!' and msg2 == '将清空所有绑定的人员信息，是否继续' and msg3 == '清空成功':
            self.WriteXlsx(17, 11, "Pass")
        else:
            self.WriteXlsx(17, 11, "Error")
        self.WriteXlsx(17, 10, "Y")
    #按照OA用户名进行查询
    def test_o_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//select[@id='bindingtype']/option[1]").click()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='searchOaInfo']").click()
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[2]")  # 鼠标悬停,OA用户名
        ActionChains(self.driver).move_to_element(move).perform()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[2]").click()
        sleep(1)
        self.Replace(18, 8)
        self.driver.find_element_by_xpath("//input[@id='searchOaInput']").send_keys(C["输入"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearchOa']").click()
        sleep(1)
        msg1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[3]").text
        print(msg1)
        if "a" in msg1:
            self.WriteXlsx(18, 11, "Pass")
        else:
            self.WriteXlsx(18, 11, "Error")
        self.WriteXlsx(18, 10, "Y")
    #按照OA登录名进行查询
    def test_p_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//select[@id='bindingtype']/option[1]").click()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='searchOaInfo']").click()
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[3]")  # 鼠标悬停,OA登录名名
        ActionChains(self.driver).move_to_element(move).perform()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[3]").click()#OA登录名名
        sleep(1)
        self.Replace(19, 8)
        self.driver.find_element_by_xpath("//input[@id='searchOaInput']").send_keys(C["输入"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearchOa']").click()
        sleep(1)
        msg1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[3]").text
        print(msg1)
        if "a" in msg1:
            self.WriteXlsx(19, 11, "Pass")
        else:
            self.WriteXlsx(19, 11, "Error")
        self.WriteXlsx(19, 10, "Y")
    #按照ERP用户编码进行查询
    def test_q_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@id='north']/a[1]").click()  # 按用户名自动绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//select[@id='bindingtype']/option[1]").click()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='searchOaInfo']").click()
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[5]")  # 鼠标悬停,ERP用户编码
        ActionChains(self.driver).move_to_element(move).perform()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[5]").click()  # ERP用户编码
        sleep(1)
        self.Replace(20, 8)
        self.driver.find_element_by_xpath("//input[@id='searchOaInput']").send_keys(C["输入"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearchOa']").click()
        sleep(1)
        msg3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]/div").text
        print(msg3)
        if "S" not in msg3:
            self.WriteXlsx(20, 11, "Pass")
        else:
            self.WriteXlsx(20, 11, "Error")
        self.WriteXlsx(20, 10, "Y")
    #按照ERP用户名进行查询
    def test_r_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//select[@id='bindingtype']/option[1]").click()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='searchOaInfo']").click()
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[4]")  # 鼠标悬停,ERP用户名
        ActionChains(self.driver).move_to_element(move).perform()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='chooseoption']/p[4]").click()  # ERP用户名
        sleep(1)
        self.Replace(21, 8)
        self.driver.find_element_by_xpath("//input[@id='searchOaInput']").send_keys(C["输入"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearchOa']").click()
        sleep(1)
        msg3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text
        print(msg3)
        if "A" in msg3:
            self.WriteXlsx(21, 11, "Pass")
        else:
            self.WriteXlsx(21, 11, "Error")
        self.WriteXlsx(21, 10, "Y")
    #ERP-按照ERP用户编码进行查询
    def test_s_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@id='searchErpInfo']").click()#查找条件
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@id='rightbox']/div[3]/div/div/div[2]/p[2]")  # 鼠标悬停,ERP用户名
        ActionChains(self.driver).move_to_element(move).perform()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='rightbox']/div[3]/div/div/div[2]/p[2]").click()
        sleep(2)
        self.Replace(22, 8)
        self.driver.find_element_by_xpath("//input[@id='searchErpInput']").send_keys(C["输入"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearchErp']").click()
        sleep(1)
        msg1 = self.driver.find_element_by_xpath("//table[@id='mytable2']/tbody/tr[1]/td[2]/div").text
        print(msg1)
        if "17" in msg1:
            self.WriteXlsx(22, 11, "Pass")
        else:
            self.WriteXlsx(22, 11, "Error")
        self.WriteXlsx(22, 10, "Y")
    #按照ERP用户名进行查询
    def test_t_People_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@id='searchErpInfo']").click()#查找条件
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@id='rightbox']/div[3]/div/div/div[2]/p[3]")  # 鼠标悬停,ERP用户名
        ActionChains(self.driver).move_to_element(move).perform()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='rightbox']/div[3]/div/div/div[2]/p[3]").click()
        sleep(1)
        self.Replace(23,8)
        self.driver.find_element_by_xpath("//input[@id='searchErpInput']").send_keys(C["输入"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearchErp']").click()
        sleep(1)
        msg1 = self.driver.find_element_by_xpath("//table[@id='mytable2']/tbody/tr[1]/td[3]/div").text
        print(msg1)
        if "崔雄A" in msg1:
            self.WriteXlsx(23, 11, "Pass")
        else:
            self.WriteXlsx(23, 11, "Error")
        self.WriteXlsx(23, 10, "Y")
    def test_u_People_config(self):
        self.System_config()
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='bindingtype']/option[2]").click()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[1]/div/input").click()#cxa
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_center']/div/div[6]/div/a[3]").click()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable2']/tbody/tr[1]/td[1]/div/input").click()#cxd
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='east_south']/a").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()



if __name__ == '__main__':
    unittest.main(verbosity=2)
