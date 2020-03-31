import unittest,re
from test_page.ZT_information import ZT_ZM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log


class TestC2_01(unittest.TestCase):
    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Log_In_DFKM(self):
        '''从登陆界面进入贷方科目配置'''
        global page, Data
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        page.open()
        page.input_username(Data.name)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        page.Sbu_Config()
        page.switch_frame(Data.myframe)
    '''贷方科目配置
    1、点击贷方科目配置，年度选择2019，选择业务枚举多级-业务枚举(单位A)，账套选择全部2、点击新增按钮'''
    def addIn_KM_config(self):
        page.DF_Button()#贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()#业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_All()#新增进入科目配置界面
        page.Add_Button()
        page.switch_frame_default()#切出来
        page.switch_frame(Data.frame1)
        page.addInZT()#账套信息界面
        page.switch_frame_default()  # 切出来
        page.switch_frame(Data.frame2)
    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ZT_infor
        ZT_infor = ZT_ZM(self.driver, self.base_url, '')
    #科目映射窗口账套查询_0261
    def test_A_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_name()
        self.Replace(262, 8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])#
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab1()
        Lab2 = ZT_infor.Lab3()
        if Lab1 == "测试帐套001" and Lab2 =="平行记帐测试账套002":
            self.WriteXlsx(262, 11, "PASS")
        else:
            self.WriteXlsx(262, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(262, 10, "Y")
    #科目映射窗口账套查询_0262
    def test_B_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_name()
        self.Replace(263, 8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])  #
        ZT_infor.Check_Button()
        if ZT_infor.Lab1()== ZT_infor.Lab2()=="测试帐套001":
            self.WriteXlsx(263, 11, "PASS")
        else:
            self.WriteXlsx(263, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(263, 10, "Y")
    #科目映射窗口账套查询_0263
    def test_C_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_name()#按账套名称查询
        self.Replace(264, 8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])  #
        ZT_infor.Check_Button()
        if ZT_infor.Lab4()== "":
            self.WriteXlsx(264, 11, "PASS")
        else:
            self.WriteXlsx(264, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(264, 10, "Y")
    #科目映射窗口账套查询_0264
    def test_D_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_name()  # 按账套名称查询
        self.Replace(265, 8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab4()
        ZT_infor.ZT_name_clear()
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.Lab1()
        Lab3 = ZT_infor.Lab3()
        if Lab1=="" and Lab2=="测试帐套001" and Lab3=="平行记帐测试账套002":
            self.WriteXlsx(265, 11, "PASS")
        else:
            self.WriteXlsx(265, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(265, 10, "Y")
    #科目映射窗口账套查询，特殊符号查询_0265
    def test_E_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_name()  # 按账套名称查询
        self.Replace(266, 8)
        ZT_infor.Send_KEY_name(C["按账套名称查询"])  #
        ZT_infor.Check_Button()
        if ZT_infor.Lab4()=="":
            self.WriteXlsx(266, 11, "PASS")
        else:
            self.WriteXlsx(266, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(266, 10, "Y")
    #科目映射窗口账套查询_0266
    def test_F_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_Num()#按账套号查询
        self.Replace(267, 8)
        ZT_infor.Send_KEY(C["按账套号查询"])  #
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.Lab1()
        Lab3 = ZT_infor.Lab3()
        if Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002":
            self.WriteXlsx(267, 11, "PASS")
        else:
            self.WriteXlsx(267, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(267, 10, "Y")
    #科目映射窗口账套查询_0267
    def test_G_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_Num()
        self.Replace(268, 8)
        ZT_infor.Send_KEY(C["按账套号查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab1()
        if Lab1 =="平行记帐测试账套002":
            self.WriteXlsx(268, 11, "PASS")
        else:
            self.WriteXlsx(268, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(268, 10, "Y")
    #科目映射窗口账套查询_0268
    def test_H_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_Num()
        self.Replace(269, 8)
        ZT_infor.Send_KEY(C["按账套号查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab4()
        if Lab1 == "":
            self.WriteXlsx(269, 11, "PASS")
        else:
            self.WriteXlsx(269, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(269, 10, "Y")
    #科目映射窗口账套查询_0269
    def test_I_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_Num()
        self.Replace(270, 8)
        ZT_infor.Send_KEY(C["按账套号查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab4()
        ZT_infor.ZT_Num_clear()
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.Lab1()
        Lab3 = ZT_infor.Lab3()
        if Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002" and Lab1=="":
            self.WriteXlsx(270, 11, "PASS")
        else:
            self.WriteXlsx(270, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(270, 10, "Y")
    #科目映射窗口账套查询，特殊符号查询_0270
    def test_J_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_Num()
        self.Replace(271, 8)
        ZT_infor.Send_KEY(C["按账套号查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab4()
        if Lab1 == "":
            self.WriteXlsx(271, 11, "PASS")
        else:
            self.WriteXlsx(271, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(271, 10, "Y")
    #科目映射窗口账套查询_0271
    def test_K_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_year()
        self.Replace(272, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])  #
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.Lab1()
        Lab3 = ZT_infor.Lab3()
        if Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002" :#and Lab1 == ""
            self.WriteXlsx(272, 11, "PASS")
        else:
            self.WriteXlsx(272, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(272, 10, "Y")
    #科目映射窗口账套查询_0272
    def test_L_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_year()
        self.Replace(273, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])  #
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.Lab1_1()
        Lab3 = ZT_infor.Lab1()
        if Lab2 == "2020"and Lab3 == "测试帐套001":
            self.WriteXlsx(273, 11, "PASS")
        else:
            self.WriteXlsx(273, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(273, 10, "Y")
    #科目映射窗口账套查询_0273
    def test_M_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_year()
        self.Replace(274, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab4()
        if Lab1 == "":
            self.WriteXlsx(274, 11, "PASS")
        else:
            self.WriteXlsx(274, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(274, 10, "Y")
    #科目映射窗口账套查询_0274
    def test_N_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_year()
        self.Replace(275, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab4()
        ZT_infor.ZT_year_clear()
        ZT_infor.Check_Button()
        Lab2 = ZT_infor.Lab1()
        Lab3 = ZT_infor.Lab3()
        if Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002" and Lab1 == "":
            self.WriteXlsx(275, 11, "PASS")
        else:
            self.WriteXlsx(275, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(275, 10, "Y")
    #科目映射窗口账套查询，特殊符号查询_0275
    def test_O_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.ZT_year()
        self.Replace(276, 8)
        ZT_infor.Send_KEY_year(C["按年度查询"])  #
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.Lab4()
        if Lab1 == "":
            self.WriteXlsx(276, 11, "PASS")
        else:
            self.WriteXlsx(276, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(276, 10, "Y")
    #科目映射窗口账套选择_0276.
    def test_P_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.First_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab1 = ZT_infor.ZT_infor_input()
        # print("Lab1:", Lab1)
        if Lab1 =="测试帐套001[001, 2019]":
            self.WriteXlsx(277, 11, "PASS")
        else:
            self.WriteXlsx(277, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(277, 10, "Y")

    #科目映射窗口账套选择_0278
    def test_Q_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.First_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        sleep(1)
        Lab1 = ZT_infor.ZT_infor_input()
        print("Lab1",Lab1)
        if Lab1 == "平行记帐测试账套002[002, 2019]":
            self.WriteXlsx(279, 11, "PASS")
        else:
            self.WriteXlsx(279, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(279, 10, "Y")
    #科目映射窗口账套清空_0279
    def test_R_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.addIn_KM_config()
        self.Glo()
        ZT_infor.First_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        # ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame1)
        Lab1 = ZT_infor.ZT_infor_input()
        page.addInZT()  # 账套信息界面
        page.switch_frame_default()  # 切出来
        page.switch_frame(Data.frame3)
        page.switch_frame_default()
        ZT_infor.Clear_frame2()#清空
        # sleep(1)
        ZT_infor.switch_frame(Data.frame1)
        Lab2 = ZT_infor.ZT_infor_input()
        if Lab1=="测试帐套001[001, 2019]" and Lab2=="":
            self.WriteXlsx(280, 11, "PASS")
        else:
            self.WriteXlsx(280, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(280, 10, "Y")
    def DF_YWMJDJ_ZTall_And_New(self):
        '''点击贷方科目配置，映射类型选择业务枚举多级-业务枚举(单位A)，账套选择全部，新增进入科目配置窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_All()
        page.Add_Button()  # 新增进入科目配置界面
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
    #科目映射窗口账套必选验证_0280
    def test_S_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.DF_YWMJDJ_ZTall_And_New()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        Lab = ZT_infor.Dialog()
        if Lab =="请选择账套信息!":
            self.WriteXlsx(281, 11, "PASS")
        else:
            self.WriteXlsx(281, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(281, 10, "Y")
    #科目映射窗口账套必选验证_0281
    def test_T_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.DF_YWMJDJ_ZTall_And_New()
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        Lab = ZT_infor.Dialog()
        if Lab == "请选择账套信息!":
            self.WriteXlsx(282, 11, "PASS")
        else:
            self.WriteXlsx(282, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(282, 10, "Y")
    #科目映射窗口账套必选验证_0282
    def test_U_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.DF_YWMJDJ_ZTall_And_New()
        self.Glo()
        ZT_infor.Click_Row5()
        ZT_infor.switch_frame_default()
        Lab = ZT_infor.Dialog()
        if Lab == "请选择账套信息!":
            self.WriteXlsx(283, 11, "PASS")
        else:
            self.WriteXlsx(283, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(283, 10, "Y")
    #科目映射窗口账套必选验证_0283
    def test_V_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.DF_YWMJDJ_ZTall_And_New()
        self.Glo()
        ZT_infor.Click_Row6()
        ZT_infor.switch_frame_default()
        Lab = ZT_infor.Dialog()
        if Lab == "请选择账套信息!":
            self.WriteXlsx(284, 11, "PASS")
        else:
            self.WriteXlsx(284, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(284, 10, "Y")
    #科目映射窗口账套必选验证_0284
    def test_W_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.DF_YWMJDJ_ZTall_And_New()
        self.Glo()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Lab = ZT_infor.Dialog()
        if Lab == "请选择账套信息!":
            self.WriteXlsx(285, 11, "PASS")
        else:
            self.WriteXlsx(285, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(285, 10, "Y")
    def Select_001And_Join_KM(self):
        '''点击贷方科目配置，映射类型选择业务枚举多级-业务枚举(单位A)，
        账套选择001，新增进入科目配置窗口，点击业务枚举多级-业务枚举(单位A)栏位，进入映射信息窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()
        page.Add_Button()  # 新增进入科目配置界面
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
    #科目配置枚举窗口定位_0287
    def test_X_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.Input_text()
        self.Replace(288,8)
        ZT_infor.Input_send_key(C["定位"])
        ZT_infor.search_Button()
        Lab1 = ZT_infor.span_text_AA()
        if Lab1=="业务枚举二级AA":
            self.WriteXlsx(288, 11, "PASS")
        else:
            self.WriteXlsx(288, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(288, 10, "Y")
        Lab2 = ZT_infor.Tbody_list_01()
        Lab3 = ZT_infor.Tbody_list_02()
        Lab4 = ZT_infor.Tbody_list_03()
        if Lab2=="业务枚举二级AA" and Lab3=="业务枚举三级AAA" and Lab4=="业务枚举三级AAB":
            self.WriteXlsx(289, 11, "PASS")
        else:
            self.WriteXlsx(289, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(289, 10, "Y")
    #科目配置枚举窗口定位_0290
    def test_Y_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.Input_text()
        self.Replace(291, 8)
        ZT_infor.Input_send_key(C["定位"])
        ZT_infor.search_Button()
        ZT_infor.search_Button()
        Lab = ZT_infor.span_text_AB()
        if Lab =="业务枚举二级AB":
            self.WriteXlsx(291, 11, "PASS")
        else:
            self.WriteXlsx(291, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(291, 10, "Y")
        Lab2 = ZT_infor.Tbody_list_01()
        if Lab2=="业务枚举二级AB":
            self.WriteXlsx(292, 11, "PASS")
        else:
            self.WriteXlsx(292, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(292, 10, "Y")
    #科目配置枚举窗口定位_0292
    def test_Z_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.span_text_B()
        lab = ZT_infor.Tbody_list_01()
        if lab =="业务枚举一级B":
            self.WriteXlsx(293, 11, "PASS")
        else:
            self.WriteXlsx(293, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(293, 10, "Y")
    #科目配置枚举窗口定位_0293
    def test_Za_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.span_text_A()
        Lab2 = ZT_infor.Tbody_list_01()
        Lab3 = ZT_infor.Tbody_list_02()
        Lab4 = ZT_infor.Tbody_list_03()
        if Lab2=="业务枚举一级A" and Lab3=="业务枚举二级AA" and Lab4=="业务枚举二级AB":
            self.WriteXlsx(294, 11, "PASS")
        else:
            self.WriteXlsx(294, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(294, 10, "Y")
    #科目配置枚举窗口定位_0294
    def test_Zb_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.mytree_2_switch()
        ZT_infor.mytree_3_span()
        Lab2 = ZT_infor.Tbody_list_01()
        Lab3 = ZT_infor.Tbody_list_02()
        Lab4 = ZT_infor.Tbody_list_03()
        if Lab2 == "业务枚举二级AA" and Lab3 == "业务枚举三级AAA" and Lab4 == "业务枚举三级AAB":
            self.WriteXlsx(295, 11, "PASS")
        else:
            self.WriteXlsx(295, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(295, 10, "Y")
    #科目配置枚举窗口查询_0297
    def test_Zc_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.YS_search_button()
        ZT_infor.YS_search_name()
        self.Replace(298,8)
        ZT_infor.YS_send_key(C["名称"])
        ZT_infor.YS_Check_button()
        Lab2 = ZT_infor.Tbody_list_01()
        Lab3 = ZT_infor.Tbody_list_02()
        if Lab2=="业务枚举二级AA" and Lab3=="业务枚举二级AB":
            self.WriteXlsx(298, 11, "PASS")
        else:
            self.WriteXlsx(298, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(298, 10, "Y")
    #科目配置枚举窗口查询_0298
    def test_Zd_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.YS_search_button()
        ZT_infor.YS_search_name()
        self.Replace(299, 8)
        ZT_infor.YS_send_key(C["名称"])
        ZT_infor.YS_Check_button()
        Lab2 = ZT_infor.Tbody_list_01()
        Lab3 = ZT_infor.Tbody_list_02()
        Lab4 = ZT_infor.Tbody_list_03()
        if Lab2 == "业务枚举二级AA" and Lab3 == "业务枚举三级AAA" and Lab4 == "业务枚举三级AAB":
            self.WriteXlsx(299, 11, "PASS")
        else:
            self.WriteXlsx(299, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(299, 10, "Y")
    #科目配置枚举窗口查询_0299
    def test_Ze_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.YS_search_button()
        ZT_infor.YS_search_name()
        self.Replace(300, 8)
        ZT_infor.YS_send_key(C["名称"])
        ZT_infor.YS_Check_button()
        lab = ZT_infor.YS_All_list()
        if lab =="":
            self.WriteXlsx(300, 11, "PASS")
        else:
            self.WriteXlsx(300, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(300, 10, "Y")
        ZT_infor.clear_YS_send_key()
        ZT_infor.YS_Check_button()
        Lab2 = ZT_infor.Tbody_list_01()
        Lab3 = ZT_infor.Tbody_list_02()
        Lab4 = ZT_infor.Tbody_list_03()
        if Lab2 == "业务枚举一级A" and Lab3 == "业务枚举一级B" and Lab4 == "业务枚举二级AA":
            self.WriteXlsx(301, 11, "PASS")
        else:
            self.WriteXlsx(301, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(301, 10, "Y")
        ZT_infor.clear_YS_send_key()
        self.Replace(302, 8)
        ZT_infor.YS_send_key(C["名称"])
        ZT_infor.YS_Check_button()
        lab5 = ZT_infor.YS_All_list()
        if lab5 == "":
            self.WriteXlsx(302, 11, "PASS")
        else:
            self.WriteXlsx(302, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(302, 10, "Y")
    #科目配置枚举窗口选择_0304
    def test_Zf_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM()
        ZT_infor.First_td_01()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab = ZT_infor.ZT_infor_input_YWMJ()
        if Lab =="业务枚举一级A":
            self.WriteXlsx(305, 11, "PASS")
        else:
            self.WriteXlsx(305, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(305, 10, "Y")
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_frame2()
        ZT_infor.switch_frame(Data.frame1)
        # ZT_infor.switch_frame_default()
        Lab2 = ZT_infor.ZT_infor_input_YWMJ()
        if Lab2 == "":
            self.WriteXlsx(306, 11, "PASS")
        else:
            self.WriteXlsx(306, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(306, 10, "Y")
    #科目配置枚举必填验证_0306
    def test_Zg_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()  # 新增进入科目配置界面
        page.Add_Button()
        page.switch_frame_default()  # 切出来
        page.switch_frame(Data.frame1)
        page.switch_frame_default()
        ZT_infor.Save_frame1()
        page.switch_frame_default()
        Msg = ZT_infor.Dialog()
        if Msg=="请选填业务枚举多级-业务枚举(单位A)!":
            self.WriteXlsx(307, 11, "PASS")
        else:
            self.WriteXlsx(307, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(307, 10, "Y")
    def Select_001And_Join_KM_CW(self):
        '''点击贷方科目配置，映射类型选择业务枚举多级-业务枚举(单位A)，账套选择001，
        新增进入科目配置窗口，点击财务科目栏，弹出科目信息窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()
        page.Add_Button()  # 新增进入科目配置界面
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)

    #科目配置财务科目定位_0323
    def test_Zh_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(324,8)
        ZT_infor.CW_INPUT_Key(C["定位"])
        ZT_infor.CW_search()
        Lab = ZT_infor.CW_mytree_20_span()
        if Lab == "1101 交易性金融资产":
            self.WriteXlsx(324, 11, "PASS")
        else:
            self.WriteXlsx(324, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(324, 10, "Y")
    #科目配置财务科目定位_0324
    def test_Zi_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(325, 8)
        ZT_infor.CW_INPUT_Key(C["定位"])
        ZT_infor.CW_search()
        ZT_infor.CW_search()
        ZT_infor.CW_search()
        Lab1 = ZT_infor.CW_mytree_107_span()
        Lab2 = ZT_infor.CW_List_one1()
        Lab3 = ZT_infor.CW_List_two1()
        Lab4 = ZT_infor.CW_List_three1()
        Lab5 = ZT_infor.CW_List_four1()
        if Lab1 =="2221 应交税费":
            if Lab2=="应交税费" and Lab3=="应交增值税" and Lab4=="进项税额" and Lab5=="销项税额":
                self.WriteXlsx(325, 11, "PASS")
            else:
                self.WriteXlsx(325, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(325, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(325, 10, "Y")
    #科目配置财务科目定位，手工定位_0326
    def test_Zj_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        ZT_infor.CW_mytree_1_span()#点击科目
        Lab1 = ZT_infor.CW_XSYS()
        Lab2 = ZT_infor.CW_YS()
        if Lab1=="条/共188条记录" and Lab2=="共10页":
            self.WriteXlsx(327, 11, "PASS")
        else:
            self.WriteXlsx(327, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(327, 10, "Y")
        ZT_infor.CW_mytree_2_span()#资产
        Lab3 = ZT_infor.CW_XSYS()
        Lab4 = ZT_infor.CW_YS()
        if Lab3 == "条/共90条记录" and Lab4 == "共5页":
            self.WriteXlsx(328, 11, "PASS")
        else:
            self.WriteXlsx(328, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(328, 10, "Y")
        ZT_infor.CW_mytree_93_span()#负债
        Lab5 = ZT_infor.CW_XSYS()
        Lab6 = ZT_infor.CW_YS()
        if Lab5 == "条/共44条记录" and Lab6 == "共3页":
            self.WriteXlsx(329, 11, "PASS")
        else:
            self.WriteXlsx(329, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(329, 10, "Y")
    #科目配置财务科目定位，手工定位_0329
    def test_Zk_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        ZT_infor.CW_mytree_138_span()  # 共同
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        # Lab3 = ZT_infor.CW_List_three1()
        Lab4 = ZT_infor.CW_List_four1()
        Lab5 = ZT_infor.CW_List_five1()
        Lab6 = ZT_infor.CW_List_six1()
        if Lab1=="清算资金往来" and Lab2=="货币兑换" :
            if Lab4 =="科目无权限" and Lab5=="衍生工具" and Lab6=="套期工具":
                self.WriteXlsx(330, 11, "PASS")
            else:
                self.WriteXlsx(330, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(330, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(330, 10, "Y")
        ZT_infor.CW_mytree_146_span()
        Lab7 = ZT_infor.CW_XSYS()
        if Lab7 == "条/共7条记录" :
            self.WriteXlsx(331, 11, "PASS")
        else:
            self.WriteXlsx(331, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(331, 10, "Y")
    #科目配置财务科目定位，手工定位_0331
    def test_Zl_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        ZT_infor.CW_mytree_162_span()
        # Lab1 = ZT_infor.CW_List_one1()
        # Lab2 = ZT_infor.CW_List_two1()
        # Lab3 = ZT_infor.CW_List_three1()
        # if Lab1=="主营业务收入" and Lab2=="分保费用" and Lab3=="以前年度损益调整":
        Lab7 = ZT_infor.CW_XSYS()
        if Lab7 == "条/共33条记录":
            self.WriteXlsx(332, 11, "PASS")
        else:
            self.WriteXlsx(332, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(332, 10, "Y")
    #科目配置财务科目定位，手工定位_0332
    def test_Zm_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        ZT_infor.CW_mytree_2_switch()
        ZT_infor.CW_mytree_4_span()
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        Lab3 = ZT_infor.CW_List_three1()
        Lab4 = ZT_infor.CW_List_four1()
        if Lab1=="银行存款" and Lab2=="现金流量科目一" and Lab3=="现金流量科目2" and Lab4=="银行科目A":
            self.WriteXlsx(333, 11, "PASS")
        else:
            self.WriteXlsx(333, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(333, 10, "Y")
        ZT_infor.CW_mytree_15_span()
        Lab5 = ZT_infor.CW_XSYS()
        if Lab5=="条/共1条记录":
            self.WriteXlsx(334, 11, "PASS")
        else:
            self.WriteXlsx(334, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(334, 10, "Y")
    #    科目配置财务科目查询_0334
    def test_Zn_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(335,8)
        ZT_infor.Check_Input_KMBM()#查询
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.YS_Check_button()#查询按钮
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        Lab3 = ZT_infor.CW_List_three1()
        if Lab1=="现金流量科目一" and Lab2=="现金流量科目2" and Lab3=="银行科目A":
            self.WriteXlsx(335, 11, "PASS")
        else:
            self.WriteXlsx(335, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(335, 10, "Y")
    #科目配置财务科目查询_0335
    def test_Zo_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(336, 8)
        ZT_infor.Check_Input_KMBM()  # 查询
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.YS_Check_button()  # 查询按钮
        LAB = ZT_infor.CW_XSYS()
        if LAB=="条/共12条记录":
            self.WriteXlsx(336, 11, "PASS")
        else:
            self.WriteXlsx(336, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(336, 10, "Y")
    #科目配置财务科目查询_0336
    def test_Zp_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(337, 8)
        ZT_infor.Check_Input_KMBM()  # 查询
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.YS_Check_button()  # 查询按钮
        LAB = ZT_infor.CW_XSYS()
        if LAB=="条/共0条记录":
            self.WriteXlsx(337, 11, "PASS")
        else:
            self.WriteXlsx(337, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(337, 10, "Y")
    #科目配置财务科目查询_0337
    def test_Zq_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(338, 8)
        ZT_infor.Check_Input_KMBM()  # 查询
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.YS_Check_button()
        LAB1 = ZT_infor.CW_XSYS()
        ZT_infor.clear_KMBM_Input()
        ZT_infor.YS_Check_button()
        LAB2 = ZT_infor.CW_XSYS()
        if LAB1 == "条/共0条记录" and LAB2 =="条/共188条记录":
            self.WriteXlsx(338, 11, "PASS")
        else:
            self.WriteXlsx(338, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(338, 10, "Y")
    #科目配置财务科目查询，特殊符号查询_0338
    def test_Zr_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(339, 8)
        ZT_infor.Check_Input_KMBM()  # 查询
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.YS_Check_button()
        Lab = ZT_infor.CW_All_text()
        if Lab=="":
            self.WriteXlsx(339, 11, "PASS")
        else:
            self.WriteXlsx(339, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(339, 10, "Y")
    #科目配置财务科目查询_0339
    def test_Zs_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(340, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.YS_Check_button()
        # Lab1 = ZT_infor.CW_List_one1()
        # Lab2 = ZT_infor.CW_List_two1()
        # Lab3 = ZT_infor.CW_List_three1()
        # Lab4 = ZT_infor.CW_List_four1()
        # Lab5 = ZT_infor.CW_List_five1()
        # if Lab1=="应交税费" and Lab2=="差旅费科目" and Lab3=="维修费科目" and Lab4=="招待费科目" and Lab5=="分保费用":
        LAB1 = ZT_infor.CW_XSYS()
        if LAB1 == "条/共18条记录":
            self.WriteXlsx(340, 11, "PASS")
        else:
            self.WriteXlsx(340, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(340, 10, "Y")
    #科目配置财务科目查询_0340
    def test_Zt_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(341, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.YS_Check_button()
        # Lab1 = ZT_infor.CW_List_one1()
        # Lab2 = ZT_infor.CW_List_two1()
        # if Lab1 == "应交税费" and Lab2 == "应交增值税":
        LAB1 = ZT_infor.CW_XSYS()
        if LAB1 == "条/共4条记录":
            self.WriteXlsx(341, 11, "PASS")
        else:
            self.WriteXlsx(341, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(341, 10, "Y")
    #科目配置财务科目查询_0341
    def test_Zu_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(342, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.YS_Check_button()
        Lab1 = ZT_infor.CW_All_text()
        if Lab1=="":
            self.WriteXlsx(342, 11, "PASS")
        else:
            self.WriteXlsx(342, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(342, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        ZT_infor.YS_Check_button()
        Lab2 = ZT_infor.CW_XSYS()
        Lab3 = ZT_infor.CW_YS()
        if Lab2 == "条/共188条记录" and Lab3 == "共10页":
            self.WriteXlsx(343, 11, "PASS")
        else:
            self.WriteXlsx(343, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(343, 10, "Y")
    #科目配置财务科目查询，特殊符号查询_0343
    def test_Zv_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(344, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.YS_Check_button()
        Lab1 = ZT_infor.CW_All_text()
        if Lab1 == "":
            self.WriteXlsx(344, 11, "PASS")
        else:
            self.WriteXlsx(344, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(344, 10, "Y")
    #科目配置财务科目查询_0344
    def test_Zw_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(345, 8)
        ZT_infor.CW_mytree_93_span()
        ZT_infor.Check_Input_KMBM()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.YS_Check_button()
        # Lab1 = ZT_infor.CW_List_one1()
        # Lab2 = ZT_infor.CW_List_two1()
        # if Lab1 == "应付账款" and Lab2 == "销项税额":
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共7条记录":
            self.WriteXlsx(345, 11, "PASS")
        else:
            self.WriteXlsx(345, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(345, 10, "Y")
    #科目配置财务科目查询_0345
    def test_Zx_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(346, 8)
        ZT_infor.CW_mytree_138_span()
        ZT_infor.Check_Input_KMBM()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.YS_Check_button()
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        if Lab1 == "货币兑换" and Lab2 == "被套期项目":
            self.WriteXlsx(346, 11, "PASS")
        else:
            self.WriteXlsx(346, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(346, 10, "Y")
    #科目配置财务科目查询_0346
    def test_Zy_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(347, 8)
        ZT_infor.CW_mytree_162_span()
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.YS_Check_button()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共11条记录":
            self.WriteXlsx(347, 11, "PASS")
        else:
            self.WriteXlsx(347, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(347, 10, "Y")
    #科目配置财务科目查询_0347
    def test_Zz_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        self.Replace(348, 8)
        ZT_infor.CW_mytree_2_span()
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.YS_Check_button()
        # Lab1 = ZT_infor.CW_List_one1()
        # Lab2 = ZT_infor.CW_List_two1()
        # Lab3 = ZT_infor.CW_List_three1()
        # if Lab1 =="银行存款" and Lab2=="应收账款" and Lab3=="其他应收款":
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 == "条/共10条记录":
            self.WriteXlsx(348, 11, "PASS")
        else:
            self.WriteXlsx(348, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(348, 10, "Y")
    #科目配置财务科目选择_0348
    def test_Zza_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        ZT_infor.CW_mytree_138_span()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab = ZT_infor.ZT_infor_input_CWKM()
        if Lab=="3001[清算资金往来]":
            self.WriteXlsx(349, 11, "PASS")
        else:
            self.WriteXlsx(349, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(349, 10, "Y")
    #科目配置财务科目选择_0349
    def test_Zzb_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        ZT_infor.CW_mytree_146_span()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab = ZT_infor.ZT_infor_input_CWKM()
        if Lab == "4001[实收资本]":
            self.WriteXlsx(350, 11, "PASS")
        else:
            self.WriteXlsx(350, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(350, 10, "Y")
    #科目配置财务科目清空_0350
    def test_Zzc_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_001And_Join_KM_CW()
        self.Glo()
        ZT_infor.CW_mytree_162_span()
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab1 = ZT_infor.ZT_infor_input_CWKM()#6542[分保费用]
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab2 = ZT_infor.ZT_infor_input_CWKM()  # 空
        if Lab1=="6011[利息收入]":
            self.WriteXlsx(351, 11, "PASS")
        else:
            self.WriteXlsx(351, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(351, 10, "Y")
    #科目配置财务科目必填验证_0351
    def test_Zzd_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()
        page.Add_Button()  # 新增进入科目配置界面
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg=="财务科目信息和预算科目信息不能同时为空!请选择...":
            self.WriteXlsx(352, 11, "PASS")
        else:
            self.WriteXlsx(352, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(352, 10, "Y")

    def Select_SJ_andJoin_SJ(self):
        '''点击贷方科目配置，映射类型选择业务枚举多级-业务枚举(单位A)，账套选择001，
        新增进入科目配置窗口，点击税金科目栏，弹出科目信息窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()
        page.Add_Button()  # 新增进入科目配置界面
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row5()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)

    # 科目配置税金科目定位_0353
    def test_Zze_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        self.Replace(354,8)
        ZT_infor.Input_text()
        ZT_infor.Input_send_key(C["定位"])
        ZT_infor.search_Button()
        Lab = ZT_infor.CW_mytree_20_span()
        if Lab == "1101 交易性金融资产":
            self.WriteXlsx(354, 11, "PASS")
        else:
            self.WriteXlsx(354, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(354, 10, "Y")
    #科目配置税金科目定位_0354
    def test_Zzf_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        self.Replace(355, 8)
        ZT_infor.Input_text()
        ZT_infor.Input_send_key(C["定位"])
        ZT_infor.CW_search()
        ZT_infor.CW_search()
        ZT_infor.CW_search()
        Lab1 = ZT_infor.CW_mytree_107_span()
        Lab2 = ZT_infor.CW_List_one1()
        Lab3 = ZT_infor.CW_List_two1()
        Lab4 = ZT_infor.CW_List_three1()
        Lab5 = ZT_infor.CW_List_four1()
        if Lab1 == "2221 应交税费":
            if Lab2 == "应交税费" and Lab3 == "应交增值税" and Lab4 == "进项税额" and Lab5 == "销项税额":
                self.WriteXlsx(355, 11, "PASS")
            else:
                self.WriteXlsx(355, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(355, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(355, 10, "Y")
    #科目配置税金科目定位，特殊符号定位_0355
    @unittest.skip("直接跳过此用例")
    def test_Zzg_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        self.Replace(356, 8)
        ZT_infor.Input_text()
        ZT_infor.Input_send_key(C["定位"])
        ZT_infor.CW_search()
        Lab = ZT_infor.CW_All_text()
        if Lab=="":
            self.WriteXlsx(356, 11, "PASS")
        else:
            self.WriteXlsx(356, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(356, 10, "Y")
    #科目配置税金科目定位，手工定位_0356
    def test_Zzh_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.CW_mytree_1_span()  # 点击科目
        Lab1 = ZT_infor.CW_XSYS()
        Lab2 = ZT_infor.CW_YS()
        if Lab1 == "条/共188条记录" and Lab2 == "共10页":
            self.WriteXlsx(357, 11, "PASS")
        else:
            self.WriteXlsx(357, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(357, 10, "Y")
        ZT_infor.CW_mytree_2_span()  # 资产
        Lab3 = ZT_infor.CW_XSYS()
        Lab4 = ZT_infor.CW_YS()
        if Lab3 == "条/共90条记录" and Lab4 == "共5页":
            self.WriteXlsx(358, 11, "PASS")
        else:
            self.WriteXlsx(358, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(358, 10, "Y")
        ZT_infor.CW_mytree_93_span()  # 负债
        Lab5 = ZT_infor.CW_XSYS()
        Lab6 = ZT_infor.CW_YS()
        if Lab5 == "条/共44条记录" and Lab6 == "共3页":
            self.WriteXlsx(359, 11, "PASS")
        else:
            self.WriteXlsx(359, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(359, 10, "Y")
    #科目配置税金科目定位，手工定位_0359
    def test_Zzi_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.CW_mytree_138_span()  # 共同
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        Lab3 = ZT_infor.CW_List_three1()
        Lab4 = ZT_infor.CW_List_four1()
        Lab5 = ZT_infor.CW_List_five1()
        Lab6 = ZT_infor.CW_List_six1()
        if Lab1 == "清算资金往来" and Lab2 == "货币兑换" and Lab3 == "科目有权限":
            if Lab4 == "科目有无限" and Lab5 == "衍生工具" and Lab6 == "套期工具":
                self.WriteXlsx(360, 11, "PASS")
            else:
                self.WriteXlsx(360, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(360, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(360, 10, "Y")
        ZT_infor.CW_mytree_146_span()
        # Lab7 = ZT_infor.CW_List_one1()
        # Lab8 = ZT_infor.CW_List_two1()
        # if Lab7 == "实收资本" and Lab8 == "库存股":
        Lab7 = ZT_infor.CW_XSYS()
        if Lab7 == "条/共7条记录":
            self.WriteXlsx(361, 11, "PASS")
        else:
            self.WriteXlsx(361, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(361, 10, "Y")
        ZT_infor.CW_mytree_162_span()
        # Lab9 = ZT_infor.CW_List_one1()
        # Lab10 = ZT_infor.CW_List_two1()
        # Lab11 = ZT_infor.CW_List_three1()
        # if Lab9 == "主营业务收入" and Lab10 == "分保费用" and Lab11 == "以前年度损益调整":
        Lab9 = ZT_infor.CW_XSYS()
        if Lab9 == "条/共33条记录":
            self.WriteXlsx(362, 11, "PASS")
        else:
            self.WriteXlsx(362, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(362, 10, "Y")
    #科目配置税金科目定位，手工定位_0362
    def test_Zzj_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.CW_mytree_2_switch()
        ZT_infor.CW_mytree_4_span()
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        Lab3 = ZT_infor.CW_List_three1()
        Lab4 = ZT_infor.CW_List_four1()
        if Lab1 == "银行存款" and Lab2 == "现金流量科目一" and Lab3 == "现金流量科目2" and Lab4 == "银行科目A":
            self.WriteXlsx(363, 11, "PASS")
        else:
            self.WriteXlsx(363, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(363, 10, "Y")
        ZT_infor.CW_mytree_22_span()
        # Lab5 = ZT_infor.CW_All_text()
        Lab9 = ZT_infor.CW_XSYS()
        if Lab9 == "条/共1条记录":
            self.WriteXlsx(364, 11, "PASS")
        else:
            self.WriteXlsx(364, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(364, 10, "Y")
    #科目配置税金科目查询_0364
    def test_Zzk_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.YS_search_button()
        self.Replace(365,8)
        ZT_infor.Check_Input_KMBM()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        Lab3 = ZT_infor.CW_List_three1()
        if Lab1=="现金流量科目一" and Lab2=="现金流量科目2" and Lab3=="银行科目A":
            self.WriteXlsx(365, 11, "PASS")
        else:
            self.WriteXlsx(365, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(365, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(366, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4=="条/共12条记录":
            self.WriteXlsx(366, 11, "PASS")
        else:
            self.WriteXlsx(366, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(366, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        self.Replace(367, 8)
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab5 = ZT_infor.CW_All_text()
        if Lab5 =="":
            self.WriteXlsx(367, 11, "PASS")
        else:
            self.WriteXlsx(367, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(367, 10, "Y")
        ZT_infor.clear_KMBM_Input()
        ZT_infor.Check_Button()
        Lab6 = ZT_infor.CW_XSYS()
        if Lab6 == "条/共188条记录":
            self.WriteXlsx(368, 11, "PASS")
        else:
            self.WriteXlsx(368, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(368, 10, "Y")
        self.Replace(369, 8)
        ZT_infor.clear_KMBM_Input()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab7 = ZT_infor.CW_All_text()
        if Lab7 == "":
            self.WriteXlsx(369, 11, "PASS")
        else:
            self.WriteXlsx(369, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(369, 10, "Y")
    #科目配置税金科目查询_0369
    def test_Zzl_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.YS_search_button()
        self.Replace(370, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        # Lab1 = ZT_infor.CW_List_one1()
        # Lab2 = ZT_infor.CW_List_two1()
        # Lab3 = ZT_infor.CW_List_three1()
        # Lab4 = ZT_infor.CW_List_four1()
        # Lab5 = ZT_infor.CW_List_five1()
        # if Lab1=="应交税费" and Lab2=="差旅费科目" and Lab3=="维修费科目" and Lab4=="招待费科目" and Lab5=="分保费用":
        Lab6 = ZT_infor.CW_XSYS()
        if Lab6 == "条/共18条记录":
            self.WriteXlsx(370, 11, "PASS")
        else:
            self.WriteXlsx(370, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(370, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(371, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        # Lab6 = ZT_infor.CW_List_one1()
        # Lab7 = ZT_infor.CW_List_two1()
        # if Lab6=="应交税费" and Lab7=="应交增值税":
        Lab7 = ZT_infor.CW_XSYS()
        if Lab7 == "条/共4条记录":
            self.WriteXlsx(371, 11, "PASS")
        else:
            self.WriteXlsx(371, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(371, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        self.Replace(372, 8)
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        Lab8 = ZT_infor.CW_All_text()
        if Lab8 == "":
            self.WriteXlsx(372, 11, "PASS")
        else:
            self.WriteXlsx(372, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(372, 10, "Y")
        ZT_infor.Clear_KMMC_Input()
        ZT_infor.Check_Button()
        Lab9 = ZT_infor.CW_XSYS()
        if Lab9 == "条/共188条记录":
            self.WriteXlsx(373, 11, "PASS")
        else:
            self.WriteXlsx(373, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(373, 10, "Y")

    #科目配置税金科目查询_0374
    def test_Zzm_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.CW_mytree_93_span()
        ZT_infor.YS_search_button()
        self.Replace(375, 8)
        ZT_infor.Check_Input_KMBM()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        # Lab1 = ZT_infor.CW_List_one1()
        # Lab2 = ZT_infor.CW_List_two1()
        # if Lab1=="应付账款" and Lab2=="销项税额":
        Lab9 = ZT_infor.CW_XSYS()
        if Lab9 == "条/共7条记录":
            self.WriteXlsx(375, 11, "PASS")
        else:
            self.WriteXlsx(375, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(375, 10, "Y")
    #科目配置税金科目查询_0375
    def test_Zzn_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.CW_mytree_138_span()
        self.Replace(376,8)
        ZT_infor.YS_search_button()#
        ZT_infor.Check_Input_KMBM()
        ZT_infor.KMBM_Input(C["科目编码"])
        ZT_infor.Check_Button()
        Lab1 = ZT_infor.CW_List_one1()
        Lab2 = ZT_infor.CW_List_two1()
        if Lab1=="货币兑换" and Lab2=="被套期项目":
            self.WriteXlsx(376, 11, "PASS")
        else:
            self.WriteXlsx(376, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(376, 10, "Y")
        ZT_infor.CW_mytree_162_span()
        self.Replace(377, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        # Lab3 = ZT_infor.CW_List_one1()
        # if Lab3=="分保费用":
        Lab9 = ZT_infor.CW_XSYS()
        if Lab9 == "条/共11条记录":
            self.WriteXlsx(377, 11, "PASS")
        else:
            self.WriteXlsx(377, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(377, 10, "Y")
        ZT_infor.CW_mytree_2_span()
        self.Replace(378, 8)
        ZT_infor.Check_Input_KMMC()
        ZT_infor.KMMC_Input(C["科目名称"])
        ZT_infor.Check_Button()
        # Lab4 = ZT_infor.CW_List_one1()
        # Lab5 = ZT_infor.CW_List_two1()
        # LAB6 = ZT_infor.CW_List_three1()
        # if Lab4 == "银行存款" and Lab5 == "应收账款" and LAB6=="其他应收款":
        Lab19 = ZT_infor.CW_XSYS()
        if Lab19 == "条/共10条记录":
            self.WriteXlsx(378, 11, "PASS")
        else:
            self.WriteXlsx(378, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(378, 10, "Y")
    #科目配置税金科目选择_0378
    def test_Zzo_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andJoin_SJ()
        ZT_infor.CW_mytree_138_span()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab = ZT_infor.ZT_infor_input_CWSJ()
        if Lab =="3001[清算资金往来]":
            self.WriteXlsx(379, 11, "PASS")
        else:
            self.WriteXlsx(379, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(379, 10, "Y")
        ZT_infor.Click_Row5()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_146_span()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        Lab1 = ZT_infor.ZT_infor_input_CWSJ()
        if Lab1 == "4001[实收资本]":
            self.WriteXlsx(380, 11, "PASS")
        else:
            self.WriteXlsx(380, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(380, 10, "Y")
    def Select_SJ_andAdd_new(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择业务枚举多级-业务枚举(单位A)，账套选择001，新增进入科目配置窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()  # 新增进入科目配置界面
        page.Add_Button()
        page.switch_frame_default()  # 切出来
        page.switch_frame(Data.frame1)

    def YWMJ_A(self):
        '''业务枚举多级-业务枚举(单位A)选择业务枚举一级A'''
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
    def CW_3001(self):
        '''财务科目选择3001清算资金往来科目'''
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_138_span()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
    def CW_4001(self):
        '''财务科目选择4001实收资本科目'''
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_146_span()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)

    #账套+年度+对象+部门重复校验_0381
    def test_Zzp_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.CW_3001()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg=="保存成功!":
            self.WriteXlsx(382, 11, "PASS")
        else:
            self.WriteXlsx(382, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(382, 10, "Y")
    #账套+年度+对象+部门重复校验_0382
    def test_Zzq_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.CW_4001()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(383, 11, "PASS")
        else:
            self.WriteXlsx(383, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(383, 10, "Y")
    #账套+年度+对象+部门重复校验_0383
    def test_Zzr_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.CW_4001()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame4)
        ZT_infor.ZK_Button()
        ZT_infor.Dep_ZH()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame4()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            self.WriteXlsx(384, 11, "PASS")
        else:
            self.WriteXlsx(384, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(384, 10, "Y")
    #账套+年度+对象+部门重复校验_0384
    def test_Zzs_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2020()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()  # 新增进入科目配置界面
        page.Add_Button()
        page.switch_frame_default()  # 切出来
        page.switch_frame(Data.frame1)
        self.YWMJ_A()
        self.CW_3001()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            self.WriteXlsx(385, 11, "PASS")
        else:
            self.WriteXlsx(385, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(385, 10, "Y")
    #账套+年度+对象+部门重复校验_0385
    def test_Zzt_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_002()  # 新增进入科目配置界面
        page.Add_Button()
        page.switch_frame_default()  # 切出来
        page.switch_frame(Data.frame1)
        self.YWMJ_A()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_JZC()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            self.WriteXlsx(386, 11, "PASS")
        else:
            self.WriteXlsx(386, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(386, 10, "Y")
    #账套+年度+对象+部门重复校验_0386
    def test_Zzu_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_001()
        self.Glo()
        ZT_infor.CW_List_one()
        page.Dele()
        page.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.switch_frame_default()
        page.OK_Msg()
        Msg2 = ZT_infor.Dialog()
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(387, 11, "PASS")
        else:
            self.WriteXlsx(387, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(387, 10, "Y")
    #账套+年度+对象+部门重复校验_0387
    def test_Zzv_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.CW_4001()
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame4)
        ZT_infor.ZK_Button()
        ZT_infor.Dep_ZH()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame4()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            self.WriteXlsx(388, 11, "PASS")
        else:
            self.WriteXlsx(388, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(388, 10, "Y")
    #账套+年度+对象+部门重复校验_0388
    def test_Zzw_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.CW_3001()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            self.WriteXlsx(389, 11, "PASS")
        else:
            self.WriteXlsx(389, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(389, 10, "Y")
    #账套+年度+对象+部门重复校验_0389
    def test_Zzx_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.CW_3001()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame4)
        ZT_infor.ZK_Button()
        ZT_infor.Dep_JCS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame4()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            self.WriteXlsx(390, 11, "PASS")
        else:
            self.WriteXlsx(390, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(390, 10, "Y")
    #账套+年度+对象+部门重复校验_0390
    def test_Zzy_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.ZK_Button()
        ZT_infor.Dep_BGS()
        ZT_infor.Dep_ZH()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame4)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame4()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(391, 11, "PASS")
        else:
            self.WriteXlsx(391, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(391, 10, "Y")
    #账套+年度+对象+部门重复校验_0391
    def test_Zzz_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.CW_3001()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame4)
        ZT_infor.ZK_Button()
        ZT_infor.Dep_JBGS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame4()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg == "保存成功!":
            self.WriteXlsx(392, 11, "PASS")
        else:
            self.WriteXlsx(392, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(392, 10, "Y")
    #账套+年度+对象+部门重复校验_0392
    def test_Zzza_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_All()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.switch_frame_default()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg1=="数据删除后将无法恢复,请确认是否进行删除!" and Msg2=="删除成功!":
            self.WriteXlsx(393, 11, "PASS")
        else:
            self.WriteXlsx(393, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(393, 10, "Y")
    #账套+年度+对象+部门重复校验_0393
    def test_Zzzb_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2020()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_All()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.switch_frame_default()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(394, 11, "PASS")
        else:
            self.WriteXlsx(394, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(394, 10, "Y")
    def Select_DJMJ_A_andAdd_new(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择多级枚举-单位枚举(单位A)，账套选择001，新增进入科目配置窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_001()  # 新增进入科目配置界面
        page.Add_Button()
        page.switch_frame_default()  # 切出来
        page.switch_frame(Data.frame1)
    #新增单位枚举科目配置_0394
    def test_Zzzc_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.YWMJ_A()
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_one_text()
        if Msg1=="保存成功!":
            self.WriteXlsx(395, 11, "PASS")
        else:
            self.WriteXlsx(395, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(395, 10, "Y")
    #新增单位枚举科目配置_0395
    def test_Zzzd_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_two_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(396, 11, "PASS")
        else:
            self.WriteXlsx(396, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(396, 10, "Y")

    def Dep_BGS_And_GWY(self):
        '''frame4'''
        '''部门名称选择办公室（包括财务室）、综合科、水政水资源科、国务院下的三级部门办公厅、综合司、条法司'''
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame4)
        ZT_infor.ZK_Button()
        ZT_infor.Dep_BGS()
        ZT_infor.Dep_ZH()
        ZT_infor.Dep_SZSZY()
        ZT_infor.Dep_ZKGWY()
        ZT_infor.Dep_CZB()
        ZT_infor.Dep_BGT()
        ZT_infor.Dep_ZHS()
        ZT_infor.Dep_TFS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame4()
        ZT_infor.switch_frame(Data.frame1)

    #新增单位枚举科目配置_0396
    def test_Zzze_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_five()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_four()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        self.Dep_BGS_And_GWY()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_three7_text()
        Lab2 = page.ZJM_first_three_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(397, 11, "PASS")
        else:
            self.WriteXlsx(397, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(397, 10, "Y")
    #新增单位枚举科目配置_0397
    def test_Zzzf_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_001()  #
        page.ZJM_List_three()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.ZK_Button()
        ZT_infor.Dep_JCS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_four7_text()
        Lab2 = page.ZJM_first_four_text()
        if Msg1=="保存成功!":
            self.WriteXlsx(398, 11, "PASS")
        else:
            self.WriteXlsx(398, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(398, 10, "Y")
    #新增业务枚举科目配置_0398
    def test_Zzzg_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.YWMJ_A()
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_one_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(399, 11, "PASS")
        else:
            self.WriteXlsx(399, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(399, 10, "Y")
    #新增业务枚举科目配置_0399
    def test_Zzzh_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_two_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(400, 11, "PASS")
        else:
            self.WriteXlsx(400, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(400, 10, "Y")
    #新增业务枚举科目配置_0400
    def test_Zzzi_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_SJ_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_five()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_four()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        self.Dep_BGS_And_GWY()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_three7_text()
        Lab2 = page.ZJM_first_three_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(401, 11, "PASS")
        else:
            self.WriteXlsx(401, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(401, 10, "Y")
    #新增业务枚举科目配置_0401
    def test_Zzzj_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_001()  #
        page.ZJM_List_three()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.Dep_JCS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_four7_text()
        Lab2 = page.ZJM_first_four_text()
        print("La1-Lab2:",Lab1,Lab2)
        if Msg1 == "保存成功!":
            self.WriteXlsx(402, 11, "PASS")
        else:
            self.WriteXlsx(402, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(402, 10, "Y")
    def QJMJ(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择请假类型-公共枚举，账套选择001，新增进入科目配置窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_QJ() #
        page.DF_ZT_and_001()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)

    #新增公共枚举科目配置_0402
    def test_Zzzk_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.QJMJ()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        self.CW_3001()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_one_text()
        Lab2 = ZT_infor.ZJM_QJ_one_text()
        if Msg1 == "保存成功!" and Lab1 == "3001[清算资金往来]" and Lab2 == "年休假":
            self.WriteXlsx(403, 11, "PASS")
        else:
            self.WriteXlsx(403, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(403, 10, "Y")
    #新增公共枚举科目配置_0403
    def test_Zzzl_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.QJMJ()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_five()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_138_span()
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_two_text()
        Lab2 = ZT_infor.ZJM_QJ_two_text()
        if Msg1 == "保存成功!" and Lab1 == "3002[货币兑换]" and Lab2 == "婚假":
            self.WriteXlsx(404, 11, "PASS")
        else:
            self.WriteXlsx(404, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(404, 10, "Y")
    #新增公共枚举科目配置_0404
    def test_Zzzm_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.QJMJ()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_seven()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_138_span()
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        self.Dep_BGS_And_GWY()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_three_text()
        Lab2 = ZT_infor.ZJM_QJ_three_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(405, 11, "PASS")
        else:
            self.WriteXlsx(405, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(405, 10, "Y")
    #新增公共枚举科目配置_0405
    def test_Zzzn_Subject_config_DA(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择请假类型-公共枚举，账套选择001，列表选择其他，新增进入科目配置窗口'''
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_QJ()  #
        page.DF_ZT_and_001()
        page.ZJM_List_three()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.Dep_JCS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_four_text()
        Lab2 = page.ZJM_first_four_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(406, 11, "PASS")
        else:
            self.WriteXlsx(406, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(406, 10, "Y")
    def Free_TextAnd2019(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择映射文本-自由文本(单位A)，账套选择001，新增进入科目配置窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_Free_Text()  #
        page.DF_ZT_and_001()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)

    def test_Zzzo_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Free_TextAnd2019()
        self.Glo()
        self.Replace(407,8)
        ZT_infor.Free_Input_Row2(C["映射文本-自由文本(单位A)"])
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_eigth()#2291
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()
        Lab2 = page.ZJM_first_one_text()
        if Msg == "保存成功!" and Lab1 == "收":
            self.WriteXlsx(407, 11, "PASS")
        else:
            self.WriteXlsx(407, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(407, 10, "Y")
    #新增自由文本科目配置_0407
    def test_Zzzp_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Free_TextAnd2019()
        self.Glo()
        self.Replace(408, 8)
        ZT_infor.Free_Input_Row2(C["映射文本-自由文本(单位A)"])
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_nine()  # 2292
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_two_text()
        Lab2 = page.ZJM_first_two_text()
        if Msg == "保存成功!" and Lab1 == "收付":
            self.WriteXlsx(408, 11, "PASS")
        else:
            self.WriteXlsx(408, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(408, 10, "Y")
    #新增自由文本科目配置_0408
    def test_Zzzq_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Free_TextAnd2019()
        self.Glo()
        self.Replace(409, 8)
        ZT_infor.Free_Input_Row2(C["映射文本-自由文本(单位A)"])
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_ten()  # 2293
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row5()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_one()  # 2001
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        self.Dep_BGS_And_GWY()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_three_text()
        Lab2 = page.ZJM_first_three7_text()
        Lab3 = page.ZJM_first_three_text()
        if Msg == "保存成功!" and Lab1 == "收付转":
            self.WriteXlsx(409, 11, "PASS")
        else:
            self.WriteXlsx(409, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(409, 10, "Y")
    #新增自由文本科目配置_0409
    def test_Zzzr_Subject_config_DA(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择映射文本-自由文本(单位A)，账套选择001，列表选择收付转，新增进入科目配置窗口'''
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_Free_Text()  #
        page.DF_ZT_and_001()
        page.ZJM_List_three()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.Dep_JCS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_four_text()
        Lab2 = page.ZJM_first_four7_text()
        Lab3 = page.ZJM_first_four_text()
        if Msg == "保存成功!" and Lab1 == "收付转":
            self.WriteXlsx(410, 11, "PASS")
        else:
            self.WriteXlsx(410, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(410, 10, "Y")

    def ZJM_Dele_DJMJ(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择多级枚举-单位枚举(单位A)，账套选择全部'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  #
        page.DF_ZT_and_All()

    #删除新增的单位枚举科目配置_0410#
    def test_Zzzs_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.ZJM_Dele_DJMJ()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(411, 11, "PASS")
        else:
            self.WriteXlsx(411, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(411, 10, "Y")
    #删除新增的业务枚举科目配置_0411
    def test_Zzzt_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWMJDJ()  # 业务枚举多级-业务枚举(单位A)
        page.DF_ZT_and_All()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.switch_frame_default()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(412, 11, "PASS")
        else:
            self.WriteXlsx(412, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(412, 10, "Y")
    #删除新增的公共枚举科目配置_0412
    def test_Zzzu_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_QJ()  # 请假
        page.DF_ZT_and_All()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.switch_frame_default()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(413, 11, "PASS")
        else:
            self.WriteXlsx(413, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(413, 10, "Y")
    #删除新增的自由文本科目配置_0413
    def test_Zzzv_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_Free_Text()  # 自由文本
        page.DF_ZT_and_All()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.switch_frame_default()
        page.OK_Msg()

        Msg2 = page.Dialog()
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(414, 11, "PASS")
        else:
            self.WriteXlsx(414, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(414, 10, "Y")
    def Select_N6N9And001(self):
        '''点击贷方科目配置，年度选择2019，映射类型选择N6/N9支出事项-系统档案(单位A)，账套选择001，新增进入科目配置窗口'''
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)

    #新增N9支出事项科目映射_0414
    def test_Zzzw_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_N6N9And001()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.N6_A()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_eigth()  # 2291
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()
        Lab2 = page.ZJM_first_one_text()
        if Msg == "保存成功!" and Lab1 == "支出事项A":
            self.WriteXlsx(415, 11, "PASS")
        else:
            self.WriteXlsx(415, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(415, 10, "Y")
    #新增N9支出事项科目映射_0415
    def test_Zzzx_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_N6N9And001()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.N6_B()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_nine()  # 2292
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_two_text()
        Lab2 = page.ZJM_first_two_text()
        if Msg == "保存成功!" and Lab1 == "支出事项B":
            self.WriteXlsx(416, 11, "PASS")
        else:
            self.WriteXlsx(416, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(416, 10, "Y")
    #新增N9支出事项科目映射_0416
    def test_Zzzy_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_N6N9And001()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.N6_C()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_ten()  # 2293
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        self.Dep_BGS_And_GWY()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_three_text()
        Lab2 = page.ZJM_first_three7_text()
        if Msg == "保存成功!" and Lab1 == "支出事项C":
            self.WriteXlsx(417, 11, "PASS")
        else:
            self.WriteXlsx(417, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(417, 10, "Y")
    #新增N9支出事项科目映射_0417
    def test_Zzzz_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()  #
        page.DF_ZT_and_001()
        page.ZJM_List_three()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.Dep_JCS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_four_text()
        Lab2 = page.ZJM_first_four7_text()
        Lab3 = page.ZJM_first_four_text()
        if Msg == "保存成功!" and Lab1 == "支出事项C":
            self.WriteXlsx(418, 11, "PASS")
        else:
            self.WriteXlsx(418, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(418, 10, "Y")

    #删除新增的N9支出事项科目配置_0418
    def test_Zzzza_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()  #
        page.DF_ZT_and_All()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.switch_frame_default()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(419, 11, "PASS")
        else:
            self.WriteXlsx(419, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(419, 10, "Y")
    #新增枚举类型科目配置_0419
    def test_Zzzzb_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.YWMJ_A()
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_one_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(420, 11, "PASS")
        else:
            self.WriteXlsx(420, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(420, 10, "Y")
    #新增枚举类型科目配置_0420
    def test_Zzzzc_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_two_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(421, 11, "PASS")
        else:
            self.WriteXlsx(421, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(421, 10, "Y")
    #新增枚举类型科目配置_0421
    def test_Zzzzd_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_five()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_four()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        self.Dep_BGS_And_GWY()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_three7_text()
        Lab2 = page.ZJM_first_three_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(422, 11, "PASS")
        else:
            self.WriteXlsx(422, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(422, 10, "Y")
    #不选择列表数据情况下修改_0423
    def test_Zzzze_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_001()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        Msg = page.Dialog()
        if Msg =="请选择一条需要修改的数据!":
            self.WriteXlsx(424, 11, "PASS")
        else:
            self.WriteXlsx(424, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(424, 10, "Y")
    #多选列表数据情况下修改_0424
    def test_Zzzzf_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_001()
        ZT_infor.CW_List_one()
        ZT_infor.CW_List_two()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        Msg = page.Dialog()
        if Msg == "请选择一条需要修改的数据!":
            self.WriteXlsx(425, 11, "PASS")
        else:
            self.WriteXlsx(425, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(425, 10, "Y")
    #修改账套_0428
    def test_Zzzzg_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_All()
        ZT_infor.CW_List_one()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row1()
        page.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab1 = ZT_infor.ZT_infor_input()#平行记帐测试账套002[002, 2019]
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        # ZT_infor.switch_frame(Data.myframe
        if Lab1 =="平行记帐测试账套002[002, 2019]" and Msg1=="保存成功!":
            self.WriteXlsx(429, 11, "PASS")
        else:
            self.WriteXlsx(429, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(429, 10, "Y")
    #修改账套_0429
    def test_Zzzzh_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_All()
        ZT_infor.CW_List_three()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row1()
        page.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        Lab1 = ZT_infor.ZT_infor_input()  # 测试帐套001[001, 2020]
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        page.select_2020()
        page.select_YWDJMJ_A()
        Lab2 = ZT_infor.CW_List_one1()
        if Lab1 =="测试帐套001[001, 2020]" and Msg1=="保存成功!" and Lab2=="测试帐套001":
            self.WriteXlsx(430, 11, "PASS")
        else:
            self.WriteXlsx(430, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(430, 10, "Y")
    #修改账套，清空_0430
    def test_Zzzzi_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()
        page.select_2020()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_All()
        ZT_infor.CW_List_one()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row1()
        page.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        if Msg =="请选择账套信息!":
            self.WriteXlsx(431, 11, "PASS")
        else:
            self.WriteXlsx(431, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(431, 10, "Y")

    #新增N9支出项目科目配置_0450
    def test_Zzzzj_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_N6N9And001()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.N6_A()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_eigth()  # 2291
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()
        Lab2 = page.ZJM_first_one_text()
        if Msg == "保存成功!" and Lab1 == "支出事项A":
            self.WriteXlsx(451, 11, "PASS")
        else:
            self.WriteXlsx(451, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(451, 10, "Y")
    #修改N9支出项目映射信息_0451
    def test_Zzzzk_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        ZT_infor.CW_List_one()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.N6_B()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()
        Lab2 = page.ZJM_first_one_text()
        if Msg == "保存成功!" and Lab1 == "支出事项B":
            self.WriteXlsx(452, 11, "PASS")
        else:
            self.WriteXlsx(452, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(452, 10, "Y")
    #修改N9支出项目映射信息_0452
    def test_Zzzzl_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        ZT_infor.CW_List_one()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.N6_C()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()
        Lab2 = page.ZJM_first_one_text()
        if Msg == "保存成功!" and Lab1 == "支出事项C":
            self.WriteXlsx(453, 11, "PASS")
        else:
            self.WriteXlsx(453, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(453, 10, "Y")
    #修改N9支出项目映射信息，请空_0453
    def test_Zzzzm_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        ZT_infor.CW_List_one()
        page.ZJM_XG_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        if Msg=="请选填N6/N9支出事项-系统档案(单位A)!":
            self.WriteXlsx(454, 11, "PASS")
        else:
            self.WriteXlsx(454, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(454, 10, "Y")
    #新增N9支出项目映射信息时，N9支出项目必填校验_0454
    def test_Zzzzn_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        self.Select_N6N9And001()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        if Msg == "请选填N6/N9支出事项-系统档案(单位A)!":
            self.WriteXlsx(455, 11, "PASS")
        else:
            self.WriteXlsx(455, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(455, 10, "Y")


    #删除科目配置，取消删除_0457
    def test_Zzzzo_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        ZT_infor.CW_List_one()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.NG_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()#支出事项C
        if Msg =="数据删除后将无法恢复,请确认是否进行删除!" and Lab1=="支出事项C":
            self.WriteXlsx(458, 11, "PASS")
        else:
            self.WriteXlsx(458, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(458, 10, "Y")
    #删除科目配置，单行删除_0458
    def test_Zzzzp_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        ZT_infor.CW_List_one()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(459, 11, "PASS")
        else:
            self.WriteXlsx(459, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(459, 10, "Y")
    #删除科目配置，多行删除_0459
    def test_Zzzzq_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Glo()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(460, 11, "PASS")
        else:
            self.WriteXlsx(460, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(460, 10, "Y")
        #增加科目配置做为测试数据_0462
    def test_Zzzzr_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.YWMJ_A()
        self.Glo()
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_one_text()
        if Msg1=="保存成功!":
            self.WriteXlsx(463, 11, "PASS")
        else:
            self.WriteXlsx(463, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(463, 10, "Y")
    #增加科目配置做为测试数据_0463
    def test_Zzzzs_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_three()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_two_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(464, 11, "PASS")
        else:
            self.WriteXlsx(464, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(464, 10, "Y")

    #增加科目配置做为测试数据_0464
    def test_Zzzzt_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_DJMJ_A_andAdd_new()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_five()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_four()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        self.Dep_BGS_And_GWY()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_three7_text()
        Lab2 = page.ZJM_first_three_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(465, 11, "PASS")
        else:
            self.WriteXlsx(465, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(465, 10, "Y")
    #增加科目配置做为测试数据_0465
    def test_Zzzzu_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_001()  #
        page.ZJM_List_three()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.switch_frame_default()
        ZT_infor.Clear_22frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row3()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.Dep_JCS()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = page.ZJM_first_four7_text()
        Lab2 = page.ZJM_first_four_text()
        if Msg1=="保存成功!":
            self.WriteXlsx(466, 11, "PASS")
        else:
            self.WriteXlsx(466, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(466, 10, "Y")
    #科目配置查询，对象名称查询_0472
    def test_Zzzzv_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_All()  #
        page.ZJM_Search_Button()
        page.ZJM_Search_Button_DXMC()
        self.Replace(473,8)
        page.ZJM_Search_Button_Input(C["对象名称"])
        page.ZJM_Search_Button_TB()
        Lab1 = page.ZJM_first_one_text()
        if Lab1 == "1632[累计折耗]":
            self.WriteXlsx(473, 11, "PASS")
        else:
            self.WriteXlsx(473, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(473, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(474,8)
        page.ZJM_Search_Button_Input(C["对象名称"])
        page.ZJM_Search_Button_TB()
        Lab2 = page.ZJM_first_one_text()
        Lab3 = page.ZJM_first_two_text()
        if Lab3 == Lab2 !="1193[三级枚举科目]":
            self.WriteXlsx(474, 11, "PASS")
        else:
            self.WriteXlsx(474, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(474, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(475, 8)
        page.ZJM_Search_Button_Input(C["对象名称"])
        page.ZJM_Search_Button_TB()
        Lab4 = page.ZJM_All_list()
        if Lab4 == "":
            self.WriteXlsx(475, 11, "PASS")
        else:
            self.WriteXlsx(475, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(475, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(476, 8)
        page.ZJM_Search_Button_Input(C["对象名称"])
        page.ZJM_Search_Button_TB()
        Lab5 = page.ZJM_first_one_text()
        Lab6 = page.ZJM_first_two_text()
        page.ZJM_Clear_Button_Input()
        page.ZJM_Search_Button_TB()
        Lab7 = page.ZJM_first_one_text()
        Lab8 = page.ZJM_first_two_text()
        if Lab5 == Lab6 !="1193[三级枚举科目]" and Lab7!="1191[一级枚举科目]" and Lab8!="1192[二级枚举科目]":
            self.WriteXlsx(476, 11, "PASS")
        else:
            self.WriteXlsx(476, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(476, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(477, 8)
        page.ZJM_Search_Button_Input(C["对象名称"])
        page.ZJM_Search_Button_TB()
        Lab9 = page.ZJM_All_list()
        if Lab9 == "":
            self.WriteXlsx(477, 11, "PASS")
        else:
            self.WriteXlsx(477, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(477, 10, "Y")
    #科目配置查询，财务科目名称查询_0477
    def test_Zzzzw_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_All()  #
        page.ZJM_Search_Button()
        page.ZJM_Search_Button_CW()
        self.Replace(478, 8)
        page.ZJM_Search_Button_Input(C["财务科目名称"])
        page.ZJM_Search_Button_TB()
        Lab1 = page.ZJM_first_one_text()
        if Lab1 != "1191[一级枚举科目]":
            self.WriteXlsx(478, 11, "PASS")
        else:
            self.WriteXlsx(478, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(478, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(479, 8)
        page.ZJM_Search_Button_Input(C["财务科目名称"])
        page.ZJM_Search_Button_TB()
        Lab2 = page.ZJM_first_one_text()
        Lab3 = page.ZJM_first_two_text()
        if Lab3 == Lab2 != "1193[三级枚举科目]":
            self.WriteXlsx(479, 11, "PASS")
        else:
            self.WriteXlsx(479, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(479, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(480, 8)
        page.ZJM_Search_Button_Input(C["财务科目名称"])
        page.ZJM_Search_Button_TB()
        Lab4 = page.ZJM_All_list()
        if Lab4 == "":
            self.WriteXlsx(480, 11, "PASS")
        else:
            self.WriteXlsx(480, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(480, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(481, 8)
        page.ZJM_Search_Button_Input(C["财务科目名称"])
        page.ZJM_Search_Button_TB()
        Lab5 = page.ZJM_first_one_text()
        Lab6 = page.ZJM_first_two_text()
        page.ZJM_Clear_Button_Input()
        page.ZJM_Search_Button_TB()
        Lab7 = page.ZJM_first_three7_text()
        Lab8 = page.ZJM_first_four7_text()
        if Lab5==Lab6==Lab7==Lab8=="1193[三级枚举科目]":
            self.WriteXlsx(481, 11, "PASS")
        else:
            self.WriteXlsx(481, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(481, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(482, 8)
        page.ZJM_Search_Button_Input(C["财务科目名称"])
        page.ZJM_Search_Button_TB()
        Lab9 = page.ZJM_All_list()
        if Lab9 == "":
            self.WriteXlsx(482, 11, "PASS")
        else:
            self.WriteXlsx(482, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(482, 10, "Y")
   # 科目配置查询，税金科目名称查询_0482
    def test_Zzzzx_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_All()  #
        page.ZJM_Search_Button()
        page.ZJM_Search_Button_SJ()
        self.Replace(484,8)
        page.ZJM_Search_Button_Input(C["税金科目名称"])
        page.ZJM_Search_Button_TB()
        Lab1 = page.ZJM_All_list()
        if Lab1 == "":
            self.WriteXlsx(484, 11, "PASS")
        else:
            self.WriteXlsx(484, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(484, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(486, 8)
        page.ZJM_Search_Button_Input(C["税金科目名称"])
        page.ZJM_Search_Button_TB()
        Lab2 = page.ZJM_All_list()
        if Lab2 == "":
            self.WriteXlsx(486, 11, "PASS")
        else:
            self.WriteXlsx(486, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(486, 10, "Y")
    #科目配置查询，部门名称查询_0486
    def test_Zzzzy_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()  # 多级枚举-单位枚举(单位A)
        page.DF_ZT_and_All()  #
        self.Replace(487,8)
        page.ZJM_Search_Button()
        page.ZJM_Search_Button_BM()
        page.ZJM_Search_Button_Input(C["部门名称"])
        page.ZJM_Search_Button_TB()
        Lab1 = page.ZJM_first_one_text_12()
        if Lab1 !="监查室":
            self.WriteXlsx(487, 11, "PASS")
        else:
            self.WriteXlsx(487, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(487, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(488, 8)
        page.ZJM_Search_Button_Input(C["部门名称"])
        page.ZJM_Search_Button_TB()
        Lab2 = page.ZJM_first_one_text_12()
        if Lab2 != "办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司":
            self.WriteXlsx(488, 11, "PASS")
        else:
            self.WriteXlsx(488, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(488, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(489, 8)
        page.ZJM_Search_Button_Input(C["部门名称"])
        page.ZJM_Search_Button_TB()
        Lab3 = page.ZJM_All_list()
        if Lab3 == "":
            self.WriteXlsx(489, 11, "PASS")
        else:
            self.WriteXlsx(489, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(489, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(490, 8)
        page.ZJM_Search_Button_Input(C["部门名称"])
        page.ZJM_Search_Button_TB()
        Lab4 = page.ZJM_first_one_text_12()
        Lab5 = page.ZJM_first_two_text_12()
        page.ZJM_Clear_Button_Input()
        page.ZJM_Search_Button_TB()
        Lab6 = page.ZJM_first_three_text()
        Lab7 = page.ZJM_first_four_text()
        if Lab4 ==Lab6!="监查室" and Lab5 ==Lab7!="办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司":
            self.WriteXlsx(490, 11, "PASS")
        else:
            self.WriteXlsx(490, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(490, 10, "Y")
        page.ZJM_Clear_Button_Input()
        self.Replace(491, 8)
        page.ZJM_Search_Button_Input(C["部门名称"])
        page.ZJM_Search_Button_TB()
        Lab8 = page.ZJM_All_list()
        if Lab8 == "":
            self.WriteXlsx(491, 11, "PASS")
        else:
            self.WriteXlsx(491, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(491, 10, "Y")
    #新增自由文本科目配置_0491
    def test_Zzzzz_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Free_TextAnd2019()
        self.Glo()
        self.Replace(492, 8)
        ZT_infor.Free_Input_Row2(C["映射文本-自由文本(单位A)"])
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_eigth()  # 2291
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()
        Lab2 = page.ZJM_first_one_text()
        if Msg == "保存成功!" and Lab1 == "收":
            self.WriteXlsx(492, 11, "PASS")
        else:
            self.WriteXlsx(492, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(492, 10, "Y")
    #新增N9支出项目科目配置_0492
    def test_Zzzzza_Subject_config_DA(self):
        self.Log_In_DFKM()
        self.Select_N6N9And001()
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.N6_A()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_93_span()  # 负债
        ZT_infor.CW_List_eigth()  # 2291
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        Msg = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab1 = ZT_infor.ZJM_QJ_one_text()
        Lab2 = page.ZJM_first_one_text()
        if Msg == "保存成功!" and Lab1 == "支出事项A":
            self.WriteXlsx(493, 11, "PASS")
        else:
            self.WriteXlsx(493, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(493, 10, "Y")
    #新增枚举类型科目配置_0493
    def test_Zzzzzb_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2020()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_2_span()  # 资产
        ZT_infor.CW_Last_page()
        ZT_infor.CW_List_two()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab = page.ZJM_first_one_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(494, 11, "PASS")
        else:
            self.WriteXlsx(494, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(494, 10, "Y")
    #新增枚举类型科目配置_0494
    def test_Zzzzzc_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_002()
        page.Add_Button()
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        self.Glo()
        ZT_infor.Click_Row2()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame2)
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame2()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.Click_Row4()
        ZT_infor.switch_frame_default()
        ZT_infor.switch_frame(Data.frame3)
        ZT_infor.CW_JZC()
        ZT_infor.CW_List_one()
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame3()
        ZT_infor.switch_frame(Data.frame1)
        ZT_infor.switch_frame_default()
        ZT_infor.Save_frame1()
        ZT_infor.switch_frame_default()
        Msg1 = ZT_infor.Dialog()
        page.OK_Msg()
        ZT_infor.switch_frame(Data.myframe)
        Lab2 = page.ZJM_first_one_text()
        if Msg1 == "保存成功!":
            self.WriteXlsx(495, 11, "PASS")
        else:
            self.WriteXlsx(495, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(495, 10, "Y")
    #年结操作提示_0504
    def test_Zzzzzd_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_All()
        page.NJ_Button()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1=="请确认是否进行【年结】操作!":
            self.WriteXlsx(505, 11, "PASS")
        else:
            self.WriteXlsx(505, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(505, 10, "Y")
        page.NG_Msg()
        self.WriteXlsx(506, 10, "Y")
        self.WriteXlsx(506, 11, "PASS")
        page.switch_frame(Data.myframe)
        page.NJ_Button()
        page.switch_frame_default()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg2=="请先选择账套":
            self.WriteXlsx(507, 11, "PASS")
        else:
            self.WriteXlsx(507, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(507, 10, "Y")
    #选择没有新年度的账套，进行年结_0507
    def test_Zzzzze_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_002()
        page.NJ_Button()
        page.switch_frame_default()
        page.OK_Msg()
        Msg1 = page.Dialog()
        if Msg1 =="已是最大年度，无法向下结转!":
            self.WriteXlsx(508, 11, "PASS")
        else:
            self.WriteXlsx(508, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(508, 10, "Y")
    #新年度有数据时，进行年结_0508
    def test_Zzzzzf_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.NJ_Button()
        page.switch_frame_default()
        page.OK_Msg()
        Msg1 = page.Dialog()
        if Msg1 == "新一年的中存在配置数据，无法进行结转...":
            self.WriteXlsx(509, 11, "PASS")
        else:
            self.WriteXlsx(509, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(509, 10, "Y")
    #旧年度没有数据时，进行年结_0509
    def test_Zzzzzg_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_QJ()
        page.DF_ZT_and_001()
        page.NJ_Button()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg2 == "没有数据可进行结转..." and Msg1=='请确认是否进行【年结】操作!':
            self.WriteXlsx(510, 11, "PASS")
        else:
            self.WriteXlsx(510, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(510, 10, "Y")
    #正常年结N9支出事项科目配置_0510
    def test_Zzzzzh_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        page.NJ_Button()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if  Msg1=='请确认是否进行【年结】操作!' and Msg2=="数据正在结转，结转完毕后会推送消息进行提醒...":
            self.WriteXlsx(511, 11, "PASS")
        else:
            self.WriteXlsx(511, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(511, 10, "Y")
    #正常年结自由文本科目配置_0512
    def test_Zzzzzi_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_Free_Text()
        page.DF_ZT_and_001()
        page.NJ_Button()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg1 == '请确认是否进行【年结】操作!' and Msg2 == "数据正在结转，结转完毕后会推送消息进行提醒...":
            self.WriteXlsx(513, 11, "PASS")
        else:
            self.WriteXlsx(513, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(513, 10, "Y")
    #删除2020年的枚举科目配置_0514
    def test_Zzzzzj_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2020()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_All()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(515, 11, "PASS")
        else:
            self.WriteXlsx(515, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(515, 10, "Y")
    #正常年结枚举科目配置_0515
    def test_Zzzzzk_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.NJ_Button()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        page.OK_Msg()
        Msg2 = page.Dialog()
        if Msg1 == '请确认是否进行【年结】操作!' and Msg2 == "数据正在结转，结转完毕后会推送消息进行提醒...":
            self.WriteXlsx(516, 11, "PASS")
        else:
            self.WriteXlsx(516, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(516, 10, "Y")
    #删除单位枚举科目配置数据_0517
    def test_Zzzzzl_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(518, 11, "PASS")
        else:
            self.WriteXlsx(518, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(518, 10, "Y")
    #删除单位枚举科目配置数据_0518
    def test_Zzzzzm_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2020()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(519, 11, "PASS")
        else:
            self.WriteXlsx(519, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(519, 10, "Y")
        #删除新增的自由文本科目配置_0519
    def test_Zzzzzn_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_YWDJMJ_A()
        page.DF_ZT_and_001()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.switch_frame_default()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(520, 11, "PASS")
        else:
            self.WriteXlsx(520, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(520, 10, "Y")
    #删除新增的自由文本科目配置_0520
    def test_Zzzzzo_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2020()
        page.select_Free_Text()
        page.DF_ZT_and_001()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(521, 11, "PASS")
        else:
            self.WriteXlsx(521, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(521, 10, "Y")
    #删除科目配置，单行删除_0521
    def test_Zzzzzp_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2019()
        page.select_N6N9()
        page.DF_ZT_and_001()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(522, 11, "PASS")
        else:
            self.WriteXlsx(522, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(522, 10, "Y")
    #删除科目配置，单行删除_0522
    def test_Zzzzzq_Subject_config_DA(self):
        self.Log_In_DFKM()
        page.DF_Button()  # 贷方科目配置
        page.select_2020()
        page.select_N6N9()
        page.DF_ZT_and_001()
        page.ALL_List()
        page.Dele()
        page.switch_frame_default()
        Msg = page.Dialog()
        page.OK_Msg()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        Lab1 = page.ZJM_All_list()  # 支出事项C
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!" and Lab1 == "":
            self.WriteXlsx(523, 11, "PASS")
        else:
            self.WriteXlsx(523, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(523, 10, "Y")



if __name__ == '__main__':
    unittest.main(verbosity=2)