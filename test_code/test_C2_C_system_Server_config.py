import unittest,re
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.support.select import Select
from test_code.JsOperate import addAttribute


class TestC2_01(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.33:8081"
        print("Test Start")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def WriteXlsx(self,row,column,data):
        Excel_path =r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\服务器配置.xlsx"
        # Excel_path = '../test_case\服务器配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["Sheet1"]
        name = worksheetname.title  # 获取表名
        # print(name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\服务器配置.xlsx"
        # Excel_path = '../test_case\服务器配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["Sheet1"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def System_config(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxa")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        # self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                # print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
    def test_a_Server_config(self):#U8数据库配置成功
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")#切换到数据库服务器页面中
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_name']").clear()
        sleep(1)
        self.Replace(2, 8)
        self.driver.find_element_by_xpath("//input[@id='db_name']").send_keys(C["数据库名称"])#服务器名称
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_hostip']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_hostip']").send_keys(C["服务器地址为U8的地址"])#服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_port']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_port']").send_keys(C["端口"])#端口
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_user']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_user']").send_keys(C["用户名"])
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_pwd']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_pwd']").send_keys(C["用户密码"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()#测试连接
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[2]").click()  # 确定按钮
        sleep(1)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg2)
        if msg1 =="当前数据库信息可用!":
            if msg2=='服务器数据保存成功!':
                self.WriteXlsx(2, 11, "Pass")
            else:
                self.WriteXlsx(2, 11, "Error")
        self.WriteXlsx(2, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
    #输入不存在的服务器地址
    def test_b_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        self.driver.find_element_by_xpath("//input[@id='db_hostip']").clear()#对服务器地址进行清除
        sleep(1)
        self.Replace(3, 8)
        self.driver.find_element_by_xpath("//input[@id='db_hostip']").send_keys(C["服务器地址为U8的地址"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        # self.driver.implicitly_wait(10)#隐式等待，10内不断查询元素，超时则报错
        sleep(10)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
        if msg1 == "ERP服务器不能正常连接，请检查!":
            self.WriteXlsx(3, 11, "Pass")
        else:
            self.WriteXlsx(3, 11, "Error")
        self.WriteXlsx(3, 10, "Y")
        self.driver.implicitly_wait(20)
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//input[@id='db_hostip']").clear()  # 对服务器地址进行清除
        sleep(1)
        self.Replace(4, 8)
        self.driver.find_element_by_xpath("//input[@id='db_hostip']").send_keys(C["服务器地址为U8的地址"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(10)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        if msg2 == "ERP服务器不能正常连接，请检查!":
            self.WriteXlsx(4, 11, "Pass")
        else:
            self.WriteXlsx(4, 11, "Error")
        self.WriteXlsx(4, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
    #输入错误的端口地址
    def test_c_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_port']").clear()#清除端口输入框
        sleep(1)
        self.Replace(5, 8)
        self.driver.find_element_by_xpath("//input[@id='db_port']").send_keys(C["端口"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(10)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        if msg1 == "ERP服务器不能正常连接，请检查!":
            self.WriteXlsx(5, 11, "Pass")
        else:
            self.WriteXlsx(5, 11, "Error")
        self.WriteXlsx(5, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        self.driver.find_element_by_xpath("//input[@id='db_port']").clear()  # 清除端口输入框
        sleep(1)
        self.Replace(6, 8)
        self.driver.find_element_by_xpath("//input[@id='db_port']").send_keys(C["端口"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(10)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        if msg2 == "数据库端口只能为数字!":
            self.WriteXlsx(6, 11, "Pass")
        else:
            self.WriteXlsx(6, 11, "Error")
        self.WriteXlsx(6, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
    #输入错误的用户名
    def test_d_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        self.driver.find_element_by_xpath("//input[@id='db_user']").clear()
        sleep(1)
        self.Replace(7,8)
        self.driver.find_element_by_xpath("//input[@id='db_user']").send_keys(C["用户名"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(4)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        if msg1 == "用户 'sa1' 登录失败。":
            self.WriteXlsx(7, 11, "Pass")
        else:
            self.WriteXlsx(7, 11, "Error")
        self.WriteXlsx(7, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//input[@id='db_user']").clear()
        sleep(1)
        self.Replace(8, 8)
        self.driver.find_element_by_xpath("//input[@id='db_user']").send_keys(C["用户名"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(4)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        if msg2 =="用户 '12345' 登录失败。":
            self.WriteXlsx(8, 11, "Pass")
        else:
            self.WriteXlsx(8, 11, "Error")
        self.WriteXlsx(8, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
    #输入错误的密码
    def test_e_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        self.driver.find_element_by_xpath("//input[@id='db_pwd']").clear()
        sleep(1)
        self.Replace(9,8)
        self.driver.find_element_by_xpath("//input[@id='db_pwd']").send_keys(C["用户密码"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(10)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()  # 弹框的确定按钮
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//input[@id='db_pwd']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_pwd']").send_keys(C["用户密码"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(10)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        if msg1 == "用户 'sa' 登录失败。" and msg2 == "用户 'sa' 登录失败。":
            self.WriteXlsx(9, 11, "Pass")
        else:
            self.WriteXlsx(9, 11, "Error")
        self.WriteXlsx(9, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
    #输入错误的数据库名称
    def test_f_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        self.driver.find_element_by_xpath("//input[@id='db_name']").clear()
        sleep(1)
        self.Replace(10, 8)
        self.driver.find_element_by_xpath("//input[@id='db_name']").send_keys(C["数据库名称1"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(10)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_name']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='db_name']").send_keys(C["数据库名称2"])
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()  # 测试连接
        sleep(10)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)
        if msg1 == '无法打开登录所请求的数据库 "ufsystemq"。 登录失败。' and msg2 == '无法打开登录所请求的数据库 "12345"。 登录失败。':
            self.WriteXlsx(10, 11, "Pass")
        else:
            self.WriteXlsx(10, 11, "Error")
        self.WriteXlsx(10, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()

    #点击确定按钮后再点击测试连接
    def test_g_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")  # 切换到数据库服务器页面中
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)
        if msg1 == "请先进行数据库测试连接!":
            self.WriteXlsx(11, 11, "Pass")
        else:
            self.WriteXlsx(11, 11, "Error")
        self.WriteXlsx(11, 10, "Y")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()
    #新建正确的OA服务器账套
    def test_h_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)

        #在Jenkins上找不到控件，暂时未找到解决方案   2020-3-20
        # main_content = Common().find_element(self.driver, ('id', 'main-content'))
        # JsOperate.addAttribute(self.driver, main_content, 'style', 'width:125%')
        # self.driver.get_screenshot_as_file('D:\AUTO_YF\Debugging_YF\\b.png')



        self.driver.switch_to.frame("templateSysMgr")

        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")  # 搜索结果显示条数
        Select(sel).select_by_value("002")
        sleep(1)
        sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        Select(sel1).select_by_value("2019")
        sleep(1)
        self.Replace(13, 8)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys(C["输入服务器地址"])  # 服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys(C["端口"])  # 端口
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()  # 保存
        sleep(1)
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)#保存成功
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()#确定
        sleep(3)
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")
        msg2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text
        print(msg2)
        if msg2 =="平行记帐测试账套002[002]" and msg1 == "保存成功!":
            self.WriteXlsx(13,11, "Pass")
        else:
            self.WriteXlsx(13,11, "Error")
        self.WriteXlsx(13, 10, "Y")
        self.driver.find_element_by_xpath("//input[@row='0']").click()  # 勾选.
        sleep(2)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除创建的002账套
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()

    #新建账套-不选择账套
    def test_i_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        self.Replace(14, 8)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys(C["输入服务器地址"])  # 服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys(C["端口"])  # 端口
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath(
            "//div[@id='layui-layer1']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()#保存
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)#年度不能为空
        # self.assertEqual(msg1,"年度不能为空")
        if msg1 == '年度不能为空':
            self.WriteXlsx(14, 11, "Pass")
        else:
            self.WriteXlsx(14, 11, "Error")
        self.WriteXlsx(14, 10, "Y")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定

    #新建账套-不输入服务器地址
    def test_j_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")  # 搜索结果显示条数
        Select(sel).select_by_value("002")
        sleep(1)
        sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        Select(sel1).select_by_value("2019")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath(
            "//div[@id='layui-layer1']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()#保存
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        # msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)#	OA服务器地址不能为空
        if msg1 == 'OA服务器地址不能为空':
            self.WriteXlsx(15, 11, "Pass")
        else:
            self.WriteXlsx(15, 11, "Error")
        self.WriteXlsx(15, 10, "Y")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
    #新建账套-输入错误的服务器地址
    def test_k_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")  # 搜索结果显示条数
        Select(sel).select_by_value("002")
        sleep(1)
        sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        Select(sel1).select_by_value("2019")
        sleep(1)
        self.Replace(16, 8)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys(C["输入服务器地址"])  # 服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys(C["端口"])  # 端口
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(30)
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 	测试连接未通过，请检查ip或端口是否正确！
        if msg1 == '测试连接未通过，请检查ip或端口是否正确！':
            self.WriteXlsx(16, 11, "Pass")
        else:
            self.WriteXlsx(16, 11, "Error")
        self.WriteXlsx(16, 10, "Y")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
    #新建账套-输入错误的端口号
    def test_l_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")  # 搜索结果显示条数
        Select(sel).select_by_value("002")
        sleep(1)
        sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        Select(sel1).select_by_value("2019")
        sleep(1)
        self.Replace(17, 8)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys(C["输入服务器地址"])  # 服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys(C["端口"])  # 端口
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath(
            "//div[@id='layui-layer1']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()  # 保存
        sleep(2)
        # self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 测试连接未通过，请检查ip或端口是否正确！
        if msg1 == '测试连接未通过，请检查ip或端口是否正确！':
            self.WriteXlsx(17, 11, "Pass")
        else:
            self.WriteXlsx(17, 11, "Error")
        self.WriteXlsx(17, 10, "Y")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
    #  新建账套-再次新建已存在的账套
    def test_m_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")  # 搜索结果显示条数
        Select(sel).select_by_value("002")
        sleep(1)
        sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        Select(sel1).select_by_value("2019")
        sleep(1)
        self.Replace(18, 8)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys(C["输入服务器地址"])  # 服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys(C["端口"])  # 端口
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath(
            "//div[@id='layui-layer1']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 保存成功！
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
        sleep(1)
        if msg1 == '该账套对应的年度已被绑定！':
            self.WriteXlsx(18, 11, "Pass")
        else:
            self.WriteXlsx(18,11, "Error")
        self.WriteXlsx(18, 10, "Y")
        # self.driver.switch_to.default_content()
        # self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
    #选择账套号001与002，点击修改
    def test_n_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")  # 搜索结果显示条数
        Select(sel).select_by_value("001")
        sleep(1)
        sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        Select(sel1).select_by_value("2019")
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys("192.168.20.33")  # 服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys("8081")  # 端口
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath(
            "//div[@id='layui-layer1']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")
        # self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 点击新建
        # sleep(1)
        # self.driver.switch_to.default_content()  # 切换出来
        # self.driver.switch_to.frame("layui-layer-iframe2")
        # sleep(1)
        # sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")  # 搜索结果显示条数
        # Select(sel).select_by_value("002")
        # sleep(1)
        # sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        # Select(sel1).select_by_value("2019")
        # sleep(1)
        # self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys("192.168.20.33")  # 服务器地址
        # sleep(1)
        # self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys("8081")  # 端口
        # sleep(1)
        # self.driver.switch_to.default_content()
        # self.driver.find_element_by_xpath(
        #     "//div[@id='layui-layer2']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()  # 保存
        # sleep(1)
        # self.driver.switch_to.default_content()
        # self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
        # sleep(1)
        # self.driver.switch_to.frame("myiframe")
        # self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//div[@class='hDiv']/div/table/thead/tr/th[1]/div/input").click()#全选
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='north1']/div[1]/a[2]").click()#修改
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 请选择一行数据！
        if msg1 == '请选择一行数据！':
            self.WriteXlsx(19, 11, "Pass")
        else:
            self.WriteXlsx(19, 11, "Error")
        self.WriteXlsx(19, 10, "Y")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
    # 选择账套号001与002，点击删除
    def test_o_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        # self.driver.find_element_by_xpath("////tbody[@id='list']/tr[1]/td[1]/div/input").click()单选
        self.driver.find_element_by_xpath("//div[@class='hDiv']/div/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='north1']/div[1]/a[3]").click()#删除
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 请选择一行数据！
        if msg1 == '请选择一行数据！':
            self.WriteXlsx(21, 11, "Pass")
        else:
            self.WriteXlsx(21, 11, "Error")
        self.WriteXlsx(21, 10, "Y")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()
    #选择账套号002点击删除
    def test_p_Server_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#单选
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='north1']/div[1]/a[3]").click()  # 删除
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg1)  # 是否删除该单据绑定数据!
        self.driver.switch_to.default_content()
        # self.driver.switch_to.frame("myiframe")
        # self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()#确定按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg2)# 删除成功！
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()
        if msg1 == "是否删除该单据绑定数据!" and msg2 == "删除成功！":
            self.WriteXlsx(22, 11, "Pass")
        else:
            self.WriteXlsx(22, 11, "Error")
        self.WriteXlsx(22, 10, "Y")
    #每页显示设置4
    # def test_q_Server_config(self):
    #     self.System_config()
    #     self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
    #     self.driver.switch_to.frame("myiframe")
    #     sleep(1)
    #     self.driver.find_element_by_xpath("//a[@id='btn2']").click()
    #     print("进入OA服务器配置")
    #     sleep(1)
    #     self.driver.switch_to.frame("templateSysMgr")
    #     self.driver.find_element_by_xpath("//div[@class='pDiv']/div[1]/span/input").send_keys(self.Replace(32,7,0))
    #     sleep(1)
    #     self.driver.find_element_by_xpath("//div[@class='pDiv']/div[1]/a[5]/a").click()
    #     sleep(1)
    #     msg = self.driver.find_element_by_xpath("//div[@class='pDiv']/div[1]/span/input").text
    #     print(msg)
    #
    #     self.WriteXlsx(30, 10, "Y")


if __name__ == '__main__':
    unittest.main(verbosity=2)
