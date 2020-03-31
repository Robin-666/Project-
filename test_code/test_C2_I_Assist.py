import unittest,re
from test_page.ZT_information import ZT_ZM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log
from test_page.People_Assist_JM import People_Assist


class TestC2_01(unittest.TestCase):
    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\辅助核算.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["辅助核算"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\辅助核算.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["辅助核算"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ZT_infor
        ZT_infor = ZT_ZM(self.driver, self.base_url, '')

    def People_FZHS(self):
        global PEOPLE
        PEOPLE = People_Assist(self.driver, self.base_url, '')

    def Log_In_People_Assist(self):
        '''从登陆界面进入人员辅助核算界面'''
        global page, Data ,PEOPLE
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        PEOPLE = People_Assist(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        PEOPLE.Assist_Button()
        page.switch_frame(Data.myframe)
        PEOPLE.People_FZHS_button()
        PEOPLE.switch_frame(Data.people_frame)

    def Select_Button_Clear_All(self):
        self.People_FZHS()
        PEOPLE.Bing_state_OK()
        PEOPLE.All_Select_Button()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        page.OK_Msg()
        self.driver.implicitly_wait(10)
        page.OK_Msg()



    def test_A_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.User_Name_Bind()
        self.driver.implicitly_wait(10)
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        Msg = PEOPLE.Msg_Frame()
        if "是否继续？？"in Msg:
            page.switch_frame_default()
            PEOPLE.Save_frame1()
            Msg2 = page.Dialog()
            if Msg2 == "成功自动绑定1条数据！":
                page.OK_Msg()
                PEOPLE.switch_frame(Data.myframe)
                PEOPLE.switch_frame(Data.people_frame)
                PEOPLE.Bing_state_OK()
                if PEOPLE.List_Display_two()==PEOPLE.List_Display_four():
                    self.WriteXlsx(2, 11, "PASS")
                    self.WriteXlsx(3, 11, "PASS")
                else:
                    self.WriteXlsx(2, 11, "ERROR")
                    self.WriteXlsx(3, 11, "ERROR")
                    do_log.error("报错：实际与预期不符")
            else:
                self.WriteXlsx(2, 11, "ERROR")
                self.WriteXlsx(3, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(2, 11, "ERROR")
            self.WriteXlsx(3, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(2, 10, "Y")
        self.WriteXlsx(3, 10, "Y")
        # self.Select_Button_Clear_All()

    def test_B_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_002()
        PEOPLE.User_Name_Bind()
        self.driver.implicitly_wait(10)
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        Msg = PEOPLE.Msg_Frame()
        if "是否继续？？"in Msg:
            page.switch_frame_default()
            PEOPLE.Save_frame1()
            Msg2 = page.Dialog()
            if Msg2 == "无数据可以绑定！":
                page.OK_Msg()
                self.WriteXlsx(4, 11, "PASS")
            else:
                self.WriteXlsx(4, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(4, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(4, 10, "Y")

    def test_C_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.User_Name_Bind()
        self.driver.implicitly_wait(10)
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        Msg = PEOPLE.Msg_Frame()
        if "是否继续？？" in Msg:
            page.switch_frame_default()
            PEOPLE.Not_Save_frame1()
            PEOPLE.switch_frame(Data.myframe)
            PEOPLE.switch_frame(Data.people_frame)
            PEOPLE.Bing_state_OK()
            if PEOPLE.All_Display()!="":
                self.WriteXlsx(5, 11, "PASS")
            else:
                self.WriteXlsx(5, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(5, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(5, 10, "Y")
        PEOPLE.Bing_state_All()
        PEOPLE.Year_2020_people()
        PEOPLE.Dep_JCS()
        PEOPLE.CXB()
        PEOPLE.CXD()
        PEOPLE.Save_Bind()

    def test_D_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_OK()
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 =="请至少选择一条数据进行清除绑定!":
            self.WriteXlsx(6, 11, "PASS")
        else:
            self.WriteXlsx(6, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(6, 10, "Y")
        PEOPLE.switch_frame_default()
        page.OK_Msg()

    def test_E_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_OK()
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.All_Select_Button()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 =="是否要清除当前所选的人员？？":
            PEOPLE.switch_frame_default()
            page.OK_Msg()
            Msg2 = page.Dialog()
            if Msg2 =="清除数据成功！请重新对照人员！":
                PEOPLE.switch_frame_default()
                page.OK_Msg()
                page.switch_frame(Data.myframe)
                PEOPLE.switch_frame(Data.people_frame)
                Lab = PEOPLE.All_Display()
                if Lab =="":
                    self.WriteXlsx(7, 11, "PASS")
                else:
                    self.WriteXlsx(7, 11, "ERROR")
                    do_log.error("报错：实际与预期不符")
            else:
                self.WriteXlsx(7, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(7, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(7, 10, "Y")

    def test_F_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_Not_OK()
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.OA_people_A()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 == "当前为未绑定人员，请选择已绑定人员进行清除绑定！！":
            self.WriteXlsx(8, 11, "PASS")
        else:
            self.WriteXlsx(8, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(8, 10, "Y")
        PEOPLE.switch_frame_default()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.ERP_people_006()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 == "当前为未绑定人员，请选择已绑定人员进行清除绑定！！":
            self.WriteXlsx(9, 11, "PASS")
        else:
            self.WriteXlsx(9, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(9, 10, "Y")

    def test_G_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.User_Name_Bind()
        self.driver.implicitly_wait(10)
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        page.switch_frame_default()
        PEOPLE.Save_frame1()
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Bing_state_OK()
        PEOPLE.All_Select_Button()
        PEOPLE.Clear_Button()
        page.switch_frame_default()
        page.NG_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        if PEOPLE.List_Display_two() == PEOPLE.List_Display_four():
            self.WriteXlsx(10, 11, "PASS")
        else:
            self.WriteXlsx(10, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(10, 10, "Y")

    def test_H_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_All()
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        self.Glo()
        Lab1 = ZT_infor.CW_XSYS()
        from MySQL_Check.people_assist_SQL_01 import Num_R
        sleep(3)
        if Lab1 =="条/共168条记录" and Num_R == "168":
            self.WriteXlsx(11, 11, "PASS")
        else:
            self.WriteXlsx(11, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(11, 10, "Y")

    def test_I_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_OK()
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        self.Glo()
        Lab1 = ZT_infor.CW_XSYS()
        from MySQL_Check.people_assist_SQL_02 import Num_2
        sleep(3)
        if Lab1 == "条/共1条记录" and Num_2 == "1":
            self.WriteXlsx(12, 11, "PASS")
        else:
            self.WriteXlsx(12, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(12, 10, "Y")
        PEOPLE.Bing_state_Not_OK()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共167条记录":
            self.WriteXlsx(13, 11, "PASS")
        else:
            self.WriteXlsx(13, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(13, 10, "Y")

    def test_J_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_All()
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001
        PEOPLE.Save_Bind()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 == "请至少选择一个OA人员进行绑定!":
            self.WriteXlsx(14, 11, "PASS")
        else:
            self.WriteXlsx(14, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(14, 10, "Y")
        page.switch_frame_default()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.OA_people_A()
        PEOPLE.Save_Bind()
        page.switch_frame_default()
        Msg2 = page.Dialog()
        if Msg2 =="请选择ERP人员！":
            self.WriteXlsx(15, 11, "PASS")
        else:
            self.WriteXlsx(15, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(15, 10, "Y")
        page.switch_frame_default()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.OA_people_BGSKY1()
        PEOPLE.OA_people_BGSKY2()
        PEOPLE.OA_people_BGSKY3()
        PEOPLE.Save_Bind()
        page.switch_frame_default()
        Msg2 = page.Dialog()
        if Msg2 == "请选择ERP人员！":
            self.WriteXlsx(16, 11, "PASS")
        else:
            self.WriteXlsx(16, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(16, 10, "Y")

    def test_K_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_All()
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.ERP_people_ZS()
        PEOPLE.Save_Bind()
        page.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 == "请至少选择一个OA人员进行绑定!":
            self.WriteXlsx(17, 11, "PASS")
        else:
            self.WriteXlsx(17, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(17, 10, "Y")
        page.switch_frame_default()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.OA_people_BGSZR()
        PEOPLE.Save_Bind()
        page.switch_frame_default()
        Msg2 = page.Dialog()
        if Msg2 == "人员绑定成功！":
            self.WriteXlsx(18, 11, "PASS")
        else:
            self.WriteXlsx(18, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(18, 10, "Y")
        page.switch_frame_default()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.ERP_people_LS()
        PEOPLE.OA_people_BGSKY1()
        PEOPLE.OA_people_BGSKY2()
        PEOPLE.OA_people_BGSKY3()
        PEOPLE.Save_Bind()
        page.switch_frame_default()
        Msg3 = page.Dialog()
        if Msg3 == "当前选择了多个OA人员绑定一个ERP业务员，导致个人往来辅助核算OA多个人员带出同一个ERP业务员,是否继续？":
            page.switch_frame_default()
            page.OK_Msg()
            Msg4 = page.Dialog()
            if Msg4 == "人员绑定成功！":
                self.WriteXlsx(19, 11, "PASS")
            else:
                self.WriteXlsx(19, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(19, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(19, 10, "Y")
        page.switch_frame_default()
        page.OK_Msg()
        page.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Bing_state_OK()

    def test_L_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_All()
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.OA_CXTJ()
        PEOPLE.OA_CXTJ_Name()
        self.Replace(20,8)
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        self.Glo()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 =="条/共8条记录":
            self.WriteXlsx(20, 11, "PASS")
        else:
            self.WriteXlsx(20, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(20, 10, "Y")
        PEOPLE.Clear_OA_Name_input_Send_key()
        self.Replace(21, 8)
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        self.Glo()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共1条记录":
            self.WriteXlsx(21, 11, "PASS")
        else:
            self.WriteXlsx(21, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(21, 10, "Y")
        PEOPLE.Clear_OA_Name_input_Send_key()
        PEOPLE.OA_CXTJ()
        PEOPLE.OA_CXTJ_YWY_BM()
        self.Replace(22, 8)
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        self.Glo()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共2条记录":
            self.WriteXlsx(22, 11, "PASS")
        else:
            self.WriteXlsx(22, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(22, 10, "Y")
        PEOPLE.Clear_OA_Name_input_Send_key()
        PEOPLE.OA_CXTJ()
        PEOPLE.OA_CXTJ_YWY_name()
        self.Replace(23, 8)
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        self.Glo()
        Lab4 = ZT_infor.CW_XSYS()
        if Lab4 == "条/共1条记录":
            self.WriteXlsx(23, 11, "PASS")
        else:
            self.WriteXlsx(23, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(23, 10, "Y")

    def test_M_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_All()
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.ERP_CXTJ()
        PEOPLE.ERP_CXTJ_YWY_BM()
        self.Replace(24,8)
        PEOPLE.ERP_input_Send_key(C["输入"])
        PEOPLE.ERP_Search_button()
        Lab1 = PEOPLE.ERP_Data_Display()
        if Lab1 == "条/共1条记录":
            self.WriteXlsx(24, 11, "PASS")
        else:
            self.WriteXlsx(24, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(24, 10, "Y")
        PEOPLE.Clear_ERP_input_Send_key()
        PEOPLE.ERP_CXTJ()
        self.Replace(25, 8)
        PEOPLE.ERP_input_Send_key(C["输入"])
        PEOPLE.ERP_Search_button()
        Lab2 = PEOPLE.ERP_Data_Display()
        if Lab2 == "条/共3条记录":
            self.WriteXlsx(25, 11, "PASS")
        else:
            self.WriteXlsx(25, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(25, 10, "Y")
        PEOPLE.Clear_ERP_input_Send_key()
        PEOPLE.ERP_CXTJ()
        PEOPLE.ERP_CXTJ_YWY_name()
        self.Replace(26, 8)
        PEOPLE.ERP_input_Send_key(C["输入"])
        PEOPLE.ERP_Search_button()
        Lab3 = PEOPLE.ERP_Data_Display()
        if Lab3 == "条/共1条记录":
            self.WriteXlsx(26, 11, "PASS")
        else:
            self.WriteXlsx(26, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(26, 10, "Y")
        PEOPLE.Clear_ERP_input_Send_key()
        PEOPLE.ERP_CXTJ()
        self.Replace(27, 8)
        PEOPLE.ERP_input_Send_key(C["输入"])
        PEOPLE.ERP_Search_button()
        Lab4 = PEOPLE.ERP_Data_Display()
        if Lab4 == "条/共1条记录":
            self.WriteXlsx(27, 11, "PASS")
        else:
            self.WriteXlsx(27, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(27, 10, "Y")

    def test_N_People_Assist(self):
        self.Log_In_People_Assist()
        self.driver.implicitly_wait(10)
        self.Replace(28,8)
        PEOPLE.DW_input_text(C["输入"])
        PEOPLE.DW_Button()
        Lab1 = PEOPLE.Unit_C()
        if Lab1 =="单位C":
            self.WriteXlsx(28, 11, "PASS")
        else:
            self.WriteXlsx(28, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(28, 10, "Y")
        PEOPLE.Clear_DW_input_text()
        self.Replace(29, 8)
        PEOPLE.DW_input_text(C["输入"])
        PEOPLE.DW_Button()
        Lab1 = PEOPLE.Administrator()
        if Lab1 == "办公室":
            self.WriteXlsx(29, 11, "PASS")
        else:
            self.WriteXlsx(29, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(29, 10, "Y")
    def test_O_People_Assist(self):
        global page, Data, PEOPLE
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        PEOPLE = People_Assist(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        PEOPLE.Assist_Button()
        self.driver.implicitly_wait(10)
        PEOPLE.prevpage()
        PEOPLE.switch_frame(Data.myframe)
        Lab1 = PEOPLE.Clear_Up()
        if Lab1 =='整理':
            self.WriteXlsx(32, 11, "PASS")
        else:
            self.WriteXlsx(32, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(32, 10, "Y")
        PEOPLE.switch_frame_default()
        PEOPLE.nextpage()
        PEOPLE.nextpage()
        PEOPLE.switch_frame(Data.myframe)
        Lab2 = PEOPLE.Year_End()
        if Lab2 == '年结':
            self.WriteXlsx(33, 11, "PASS")
        else:
            self.WriteXlsx(33, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(33, 10, "Y")

    '''验证枚举带科目功能，先配置：
    1，2020、账套001、多级枚举-单位枚举（借方/贷方科目配置年度、账套一致）
    2，借方科目配置：一级枚举B绑定科目1148、一级枚举C绑定科目1147，DDDDD绑定科目1146且科目都带了辅助核算
    3，贷方科目配置：DDDDD绑定科目100202、一级枚举B绑定科目1012、一级枚举C绑定科目1148'''
    def test_P_People_Assist(self):
        '''DDDD枚举带出1146科目'''
        global page, Data, PEOPLE
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        PEOPLE = People_Assist(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        ZJM = self.driver.current_window_handle  #切换窗口前的句柄
        PEOPLE.JF_MJ_DDDD()
        # page.Win_Split()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                # print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
        page.switch_frame(Data.zwIframe)
        Lab1 = PEOPLE.JF_KMBM_Text_Value()
        Lab2 = PEOPLE.JF_KM_name_Text_Value()
        PEOPLE.JF_MJ_Button()
        PEOPLE.JF_Select_MJ_DDDD()
        Lab3 = PEOPLE.JF_KMBM_Text_Value()
        Lab4 = PEOPLE.JF_KM_name_Text_Value()
        if Lab1 == Lab2 and Lab3=="1146" and Lab4 =="客项部":
            self.WriteXlsx(34, 11, "PASS")
        else:
            self.WriteXlsx(34, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(34, 10, "Y")
        sleep(3)
        '''一级枚举C带出1147个项银行'''
        PEOPLE.JF_MJ_Button()
        PEOPLE.JF_Select_MJ_C()
        Lab5 = PEOPLE.JF_KMBM_Text_Value()
        Lab6 = PEOPLE.JF_KM_name_Text_Value()
        if Lab5 == "1147" and Lab6 == "个项银行":
            self.WriteXlsx(35, 11, "PASS")
        else:
            self.WriteXlsx(35, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(35, 10, "Y")
        sleep(3)
        '''一级枚举B带出1148外币核算人民币'''
        PEOPLE.JF_MJ_Button()
        PEOPLE.JF_Select_MJ_B()
        Lab7 = PEOPLE.JF_KMBM_Text_Value()
        Lab8 = PEOPLE.JF_KM_name_Text_Value()
        if Lab7 == "1148" and Lab8 == "外币核算人民币":
            self.WriteXlsx(36, 11, "PASS")
        else:
            self.WriteXlsx(36, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(36, 10, "Y")
        sleep(3)
        '''贷方科目配置-DDDD带出100202 银行存款-现金流量科目2'''
        Lab9 = PEOPLE.DF_KMBM_Text_Value()
        Lab10 = PEOPLE.DF_KM_name_Text_Value()
        PEOPLE.DF_MJ_Button()
        PEOPLE.DF_Select_MJ_DDDD()
        Lab11 = PEOPLE.DF_KMBM_Text_Value()
        Lab12 = PEOPLE.DF_KM_name_Text_Value()
        if Lab9 == Lab10 and Lab11 =="100202" and Lab12 == "银行存款-现金流量科目2":
            self.WriteXlsx(37, 11, "PASS")
        else:
            self.WriteXlsx(37, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(37, 10, "Y")
        sleep(3)
        '''一级C带出1304贷款损失准备'''
        PEOPLE.DF_MJ_Button()
        PEOPLE.DF_Select_MJ_C()
        Lab13 = PEOPLE.DF_KMBM_Text_Value()
        Lab14 = PEOPLE.DF_KM_name_Text_Value()
        if Lab13 == "1304" and Lab14 =="贷款损失准备":
            self.WriteXlsx(38, 11, "PASS")
        else:
            self.WriteXlsx(38, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(38, 10, "Y")
        sleep(3)
        '''一级B带出1012其他货币资金-外币港元'''
        PEOPLE.DF_MJ_Button()
        PEOPLE.DF_Select_MJ_B()
        Lab15 = PEOPLE.DF_KMBM_Text_Value()
        Lab16 = PEOPLE.DF_KM_name_Text_Value()
        if Lab15 == "1012" and Lab16 == "其他货币资金-外币港元":
            self.WriteXlsx(39, 11, "PASS")
        else:
            self.WriteXlsx(39, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(39, 10, "Y")

    '''CXB带出U8业务员CXD'''
    def test_Q_People_Assist(self):
        global page, Data, PEOPLE
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        PEOPLE = People_Assist(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        ZJM = self.driver.current_window_handle  # 切换窗口前的句柄
        PEOPLE.JF_MJ_DDDD()#选择表单
        # page.Win_Split()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                sleep(1)
        page.switch_frame(Data.zwIframe)
        PEOPLE.Select_People_icon()
        PEOPLE.switch_frame_default()
        PEOPLE.switch_frame(Data.frame1)
        PEOPLE.Select_CXB()
        PEOPLE.Select_Right_icon()
        PEOPLE.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        page.switch_frame(Data.zwIframe)
        PEOPLE.JF_MJ_Button()
        PEOPLE.JF_Select_MJ_C()
        self.Replace(40,8)
        PEOPLE.JF_BXJE_Text_Value(C["报销金额"])
        PEOPLE.DF_MJ_Button()
        PEOPLE.DF_Select_MJ_B()
        PEOPLE.DF_BXJE_Text_Value(C["支付金额"])
        PEOPLE.switch_frame_default()
        PEOPLE.Form_Send()
        self.driver.switch_to.window(ZJM)
        page.Select_XT_Work()
        PEOPLE.Select_Wait_Do()
        PEOPLE.switch_frame(Data.mainIframe)
        PEOPLE.All_Wait_Do()
        PEOPLE.Wait_Do_List_Display_Button()
        Win1 = self.driver.window_handles  # 所有窗口句柄
        for windows1 in Win1:
            if windows1 != ZJM:
                self.driver.switch_to.window(windows1)
                sleep(1)
        PEOPLE.Agree_Button()
        PEOPLE.switch_frame_default()
        PEOPLE.switch_frame(Data.frame1)
        PEOPLE.Span_Title()
        Lab = PEOPLE.PZ_Display()
        if '崔雄D' in Lab:
            self.WriteXlsx(40, 11, "PASS")
        else:
            self.WriteXlsx(40, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(40, 10, "Y")

    def Log_In_Dep_Assist(self):
        '''从登陆界面进入部门辅助核算界面'''
        global page, Data ,PEOPLE
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        PEOPLE = People_Assist(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        PEOPLE.Assist_Button()
        page.switch_frame(Data.myframe)
        PEOPLE.Dep_FZHS_button()
        PEOPLE.switch_frame(Data.people_frame)

    def test_R_People_Assist(self):
        self.Log_In_Dep_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.User_Name_Bind()
        self.driver.implicitly_wait(10)
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        Msg = PEOPLE.Msg_Frame()
        if "是否继续？？" in Msg:
            page.switch_frame_default()
            PEOPLE.Save_frame1()
            Msg2 = page.Dialog()
            if Msg2 == "成功自动绑定4条数据！":
                page.OK_Msg()
                PEOPLE.switch_frame(Data.myframe)
                PEOPLE.switch_frame(Data.people_frame)
                PEOPLE.Bing_state_OK()
                if PEOPLE.List_Display_two() == PEOPLE.List_Display_four():
                    self.WriteXlsx(41, 11, "PASS")
                    self.WriteXlsx(42, 11, "PASS")
                else:
                    self.WriteXlsx(41, 11, "ERROR")
                    self.WriteXlsx(42, 11, "ERROR")
                    do_log.error("报错：实际与预期不符")
            else:
                self.WriteXlsx(41, 11, "ERROR")
                self.WriteXlsx(42, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(41, 11, "ERROR")
            self.WriteXlsx(42, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(41, 10, "Y")
        self.WriteXlsx(42, 10, "Y")
    def test_S_People_Assist(self):
        self.Log_In_Dep_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_002()
        PEOPLE.User_Name_Bind()
        self.driver.implicitly_wait(10)
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        Msg = PEOPLE.Msg_Frame()
        if "是否继续？？" in Msg:
            page.switch_frame_default()
            PEOPLE.Save_frame1()
            Msg2 = page.Dialog()
            if Msg2 == "无数据可以绑定！":
                page.OK_Msg()
                self.WriteXlsx(43, 11, "PASS")
            else:
                self.WriteXlsx(43, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(43, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(43, 10, "Y")
        # self.Select_Button_Clear_All()

    def test_T_People_Assist(self):
        self.Log_In_Dep_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2019_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.User_Name_Bind()
        self.driver.implicitly_wait(10)
        page.switch_frame_default()
        page.switch_frame(Data.frame1)
        Msg = PEOPLE.Msg_Frame()
        if "是否继续？？" in Msg:
            page.switch_frame_default()
            PEOPLE.Not_Save_frame1()
            PEOPLE.switch_frame(Data.myframe)
            PEOPLE.switch_frame(Data.people_frame)
            PEOPLE.Bing_state_OK()
            if PEOPLE.All_Display() == "":
                self.WriteXlsx(44, 11, "PASS")
            else:
                self.WriteXlsx(44, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(44, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(44, 10, "Y")

    def test_U_People_Assist(self):
        self.Log_In_Dep_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_OK()
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 == "请至少选择一条数据进行清除绑定!":
            self.WriteXlsx(45, 11, "PASS")
        else:
            self.WriteXlsx(45, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(45, 10, "Y")
        PEOPLE.switch_frame_default()
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Bing_state_OK()
        PEOPLE.All_Select_Button()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg2 = page.Dialog()
        if Msg2 =="是否要清除当前所选的部门？？":
            PEOPLE.switch_frame_default()
            page.OK_Msg()
            Msg3 = page.Dialog()
            if Msg3 == "清除数据成功！请重新对照部门！":
                PEOPLE.switch_frame_default()
                page.OK_Msg()
                PEOPLE.switch_frame(Data.myframe)
                PEOPLE.switch_frame(Data.people_frame)
                if PEOPLE.All_Display() == "":
                    self.WriteXlsx(46, 11, "PASS")
                else:
                    self.WriteXlsx(46, 11, "ERROR")
                    do_log.error("报错：实际与预期不符")
            else:
                self.WriteXlsx(46, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(46, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(46, 10, "Y")
        PEOPLE.Bing_state_All()
        PEOPLE.Dep_BGS()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg4 = page.Dialog()
        if Msg4 =="是否要清除当前所选的部门？？":
            PEOPLE.switch_frame_default()
            page.OK_Msg()
            Msg5 = page.Dialog()
            if Msg5 =="清除数据成功！请重新对照部门！":
                PEOPLE.switch_frame_default()
                page.OK_Msg()
                self.WriteXlsx(47, 11, "PASS")
            else:
                self.WriteXlsx(47, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(47, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(47, 10, "Y")
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Dep_ERP_TestDep()
        PEOPLE.Clear_Button()
        PEOPLE.switch_frame_default()
        Msg6 = page.Dialog()
        if Msg6 == "请至少选择一条数据进行清除绑定!":
            self.WriteXlsx(48, 11, "PASS")
        else:
            self.WriteXlsx(48, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(48, 10, "Y")
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.User_Name_Bind()
        PEOPLE.switch_frame_default()
        PEOPLE.switch_frame(Data.frame1)
        PEOPLE.switch_frame_default()
        PEOPLE.Dialog_OK_Button()
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Bing_state_OK()
        PEOPLE.All_Select_Button()
        PEOPLE.Clear_Button()
        page.switch_frame_default()
        Msg7 = page.Dialog()
        page.switch_frame_default()
        page.NG_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        if PEOPLE.List_Display_two() == PEOPLE.List_Display_four():
            self.WriteXlsx(49, 11, "PASS")
        else:
            self.WriteXlsx(49, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(49, 10, "Y")

    def test_V_People_Assist(self):
        self.Log_In_Dep_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Bing_state_OK()
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001()
        self.Glo()
        Lab1 = ZT_infor.CW_XSYS()
        if Lab1 =="条/共8条记录":
            self.WriteXlsx(51, 11, "PASS")
        else:
            self.WriteXlsx(51, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(51, 10, "Y")
        PEOPLE.Bing_state_Not_OK()
        Lab2 = ZT_infor.CW_XSYS()
        if Lab2 == "条/共18条记录":
            self.WriteXlsx(52, 11, "PASS")
        else:
            self.WriteXlsx(52, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(52, 10, "Y")
        PEOPLE.Bing_state_All()
        Lab3 = ZT_infor.CW_XSYS()
        if Lab3 == "条/共26条记录":
            self.WriteXlsx(50, 11, "PASS")
        else:
            self.WriteXlsx(50, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(50, 10, "Y")

    def test_W_People_Assist(self):
        self.Log_In_Dep_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Save_Bind()
        PEOPLE.switch_frame_default()
        Msg1 = page.Dialog()
        if Msg1 =="请至少选择一个OA部门进行绑定!":
            self.WriteXlsx(53, 11, "PASS")
        else:
            self.WriteXlsx(53, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(53, 10, "Y")
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Dep_ZHK()
        PEOPLE.Save_Bind()
        PEOPLE.switch_frame_default()
        Msg2 = page.Dialog()
        if Msg2 =="请选择ERP部门！":
            self.WriteXlsx(54, 11, "PASS")
        else:
            self.WriteXlsx(54, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(54, 10, "Y")
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Dep_ZHS()
        PEOPLE.Dep_TFS()
        PEOPLE.Save_Bind()
        PEOPLE.switch_frame_default()
        Msg3 = page.Dialog()
        if Msg3 == "请选择ERP部门！":
            self.WriteXlsx(55, 11, "PASS")
        else:
            self.WriteXlsx(55, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(55, 10, "Y")
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Clear_Button()#先清除一下选中的部门
        PEOPLE.switch_frame_default()
        PEOPLE.switch_frame_default()
        page.OK_Msg()
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Dep_ERP_TestDep()
        PEOPLE.Save_Bind()
        PEOPLE.switch_frame_default()
        Msg5 = page.Dialog()
        if Msg5 == "请至少选择一个OA部门进行绑定!":
            self.WriteXlsx(56, 11, "PASS")
        else:
            self.WriteXlsx(56, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(56, 10, "Y")
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Dep_GWY()
        PEOPLE.Dep_JSB()
        PEOPLE.Save_Bind()
        PEOPLE.switch_frame_default()
        Msg6 = page.Dialog()
        if Msg6 == "部门绑定成功！":
            self.WriteXlsx(57, 11, "PASS")
        else:
            self.WriteXlsx(57, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(57, 10, "Y")
        page.OK_Msg()
        PEOPLE.switch_frame(Data.myframe)
        PEOPLE.switch_frame(Data.people_frame)
        PEOPLE.Dep_TFS()
        PEOPLE.Dep_ZHS()
        PEOPLE.Dep_ZHK()
        PEOPLE.Dep_ERP_TestDep()
        PEOPLE.Save_Bind()
        PEOPLE.switch_frame_default()
        Msg4 = page.Dialog()
        if Msg4 == "当前选择了多个OA部门绑定一个ERP部门，导致部门辅助核算OA多个部门带出同一个ERP部门，是否继续？？":
            page.OK_Msg()
            Msg7 = page.Dialog()
            if Msg7 =="部门绑定成功！":
                page.OK_Msg()
                self.WriteXlsx(58, 11, "PASS")
            else:
                self.WriteXlsx(58, 11, "ERROR")
                do_log.error("报错：实际与预期不符")
        else:
            self.WriteXlsx(58, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(58, 10, "Y")

    def test_X_People_Assist(self):
        self.Log_In_Dep_Assist()
        self.driver.implicitly_wait(10)
        PEOPLE.Year_2020_people()
        PEOPLE.PEOPLE_ZT_001()
        PEOPLE.OA_CXTJ()
        self.Replace(59,8)
        PEOPLE.OA_CXTJ_Name()
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        Lab1 = PEOPLE.All_Display()
        if Lab1 =="":
            self.WriteXlsx(59, 11, "PASS")
        else:
            self.WriteXlsx(59, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(59, 10, "Y")
        PEOPLE.Clear_OA_Name_input_Send_key()
        self.Replace(60, 8)
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        Lab2 = PEOPLE.List_Display_two()
        if Lab2=="综合科":
            self.WriteXlsx(60, 11, "PASS")
        else:
            self.WriteXlsx(60, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(60, 10, "Y")
        PEOPLE.Clear_OA_Name_input_Send_key()
        PEOPLE.OA_CXTJ()
        PEOPLE.OA_CXTJ_YWY_BM()
        self.Replace(61, 8)
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        if PEOPLE.List_Display_two() =="综合科" and PEOPLE.List_Display_four()=="测试部":
            self.WriteXlsx(61, 11, "PASS")
        else:
            self.WriteXlsx(61, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(61, 10, "Y")
        PEOPLE.Clear_OA_Name_input_Send_key()
        PEOPLE.OA_CXTJ()
        PEOPLE.OA_CXTJ_YWY_name()
        self.Replace(62, 8)
        PEOPLE.OA_Name_input_Send_key(C["输入"])
        PEOPLE.OA_Search_OA_NAME()
        if PEOPLE.List_Display_two() =="国务院" and PEOPLE.List_Display_four()=="技术部":
            self.WriteXlsx(62, 11, "PASS")
        else:
            self.WriteXlsx(62, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(62, 10, "Y")
        PEOPLE.ERP_CXTJ()
        PEOPLE.ERP_CXTJ_YWY_BM()
        self.Replace(63, 8)
        PEOPLE.ERP_input_Send_key(C["输入"])
        PEOPLE.ERP_Search_button()
        Lab3 = PEOPLE.ERP_Data_Display()
        if Lab3 == "条/共1条记录":
            self.WriteXlsx(63, 11, "PASS")
        else:
            self.WriteXlsx(63, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(63, 10, "Y")
        PEOPLE.Clear_ERP_input_Send_key()
        PEOPLE.ERP_CXTJ()
        PEOPLE.ERP_CXTJ_YWY_name()
        self.Replace(64, 8)
        PEOPLE.ERP_input_Send_key(C["输入"])
        PEOPLE.ERP_Search_button()
        Lab4 = PEOPLE.ERP_Data_Display()
        if Lab4 == "条/共1条记录":
            self.WriteXlsx(64, 11, "PASS")
        else:
            self.WriteXlsx(64, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(64, 10, "Y")

    def test_Y_People_Assist(self):
        global page, Data, PEOPLE
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        PEOPLE = People_Assist(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        ZJM = self.driver.current_window_handle  # 切换窗口前的句柄
        PEOPLE.JF_MJ_DDDD()  # 选择表单
        # page.Win_Split()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                sleep(1)
        page.switch_frame(Data.zwIframe)
        PEOPLE.Select_People_icon()
        PEOPLE.switch_frame_default()
        PEOPLE.switch_frame(Data.frame1)
        PEOPLE.Select_CXB()
        PEOPLE.Select_Right_icon()
        PEOPLE.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        page.switch_frame(Data.zwIframe)
        PEOPLE.JF_MJ_Button()
        PEOPLE.JF_Select_MJ_A()
        self.Replace(65, 8)
        PEOPLE.JF_BXJE_Text_Value(C["报销金额"])
        PEOPLE.DF_MJ_Button()
        PEOPLE.DF_Select_MJ_B()
        PEOPLE.DF_BXJE_Text_Value(C["支付金额"])
        PEOPLE.switch_frame_default()
        PEOPLE.Form_Send()
        self.driver.switch_to.window(ZJM)
        page.Select_XT_Work()
        PEOPLE.Select_Wait_Do()
        PEOPLE.switch_frame(Data.mainIframe)
        PEOPLE.All_Wait_Do()
        PEOPLE.Wait_Do_List_Display_Button()
        Win1 = self.driver.window_handles  # 所有窗口句柄
        for windows1 in Win1:
            if windows1 != ZJM:
                self.driver.switch_to.window(windows1)
                sleep(1)
        PEOPLE.Agree_Button()
        PEOPLE.switch_frame_default()
        PEOPLE.switch_frame(Data.frame1)
        PEOPLE.Span_Title_1131()
        Lab = PEOPLE.PZ_Display()
        if '测试部' in Lab:
            self.WriteXlsx(65, 11, "PASS")
        else:
            self.WriteXlsx(65, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(65, 10, "Y")

if __name__ == '__main__':
    unittest.main(verbosity=2)






















