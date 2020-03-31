import unittest,re
from test_page.ERP_config_JM import ERP_JM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log


class TestC2_01(unittest.TestCase):

    def Log_In_ERP(self):
        '''从登陆界面进入ERP配置'''
        global page, Data
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        page.CWJC()
        page.Win_XF()
        page.Win_Split()
        page.ERP_config_Button()
        page.switch_frame(Data.myframe)
        sleep(1)
    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\ERP配置.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["ERP配置"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\ERP配置.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["ERP配置"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)
    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ERP_infor
        ERP_infor = ERP_JM(self.driver, self.base_url, '')

    @unittest.skip("直接跳过")
    def test_A_Button_ON(self):
        self.Log_In_ERP()
        self.Glo()
        # ERP_infor.ERP_JM_ON()    #判断必填按钮的初始状态
        ERP_infor.OK_button()
        page.switch_frame(Data.ERP_frame1)
        ERP_infor.ERP_Dialog()
        page.switch_frame_default()
        ERP_infor.ERP_OK_Msg()
        



















































































if __name__ == '__main__':
    unittest.main(verbosity=2)