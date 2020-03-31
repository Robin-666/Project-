import unittest,re
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.support.select import Select

class TestC2_01(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.33:8081"
        print("Test Start")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        # Excel_path = '../test_case\人员对照.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)



    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)


    def System_config(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxa")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        # self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                # print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
        self.driver.find_element_by_xpath("//div[@title = '科目配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
    #新增功能窗口展示
    def test_a_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()#点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()#2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()#账套选择-全部
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()#新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Label1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/th").text#账套信息:
        Label2 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/th").text#业务枚举多级-业务枚举(单位A):
        Label3 = self.driver.find_element_by_xpath("//tr[@id='DeptTR']/th").text#部门名称:
        Label4 = self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/th").text#财务科目:
        Label5 = self.driver.find_element_by_xpath("//tr[@id='TaxTR']/th").text#税金科目:
        Label6 = self.driver.find_element_by_xpath("//tr[@id='BudgetTR']/th").text#预算科目:
        print("Label1~Label6:",Label1,Label2,Label3,Label4,Label5,Label6)
        if Label1 == "账套信息:" and Label2 =="业务枚举多级-业务枚举(单位A):" and Label3 =="部门名称:" :
            if Label4=="财务科目:"and Label5 =="税金科目:" and Label6 =="预算科目:":
                self.WriteXlsx(2,11,"PASS")
            else:
                self.WriteXlsx(2, 11, "ERROR")
        else:
            self.WriteXlsx(2, 11, "ERROR")
        self.WriteXlsx(2, 10, "Y")
    #不选账套时，新增功能默认账套显示
    def test_b_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 账套选择-全部
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        TEXT = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").text
        print('TEXT:',TEXT)
        if TEXT == "":
            self.WriteXlsx(3, 11, "PASS")
        else:
            self.WriteXlsx(3, 11, "ERROR")
        self.WriteXlsx(3, 10, "Y")
    #选择账套时，新增功能默认账套显示
    def test_c_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        sel = self.driver.find_element_by_xpath("//select[@id='account1']")  # 账套选择-全部
        Select(sel).select_by_value("001[@]U8")#测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        TEXT = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").get_attribute('value')#获取输入框内容
        print('TEXT:',TEXT)#测试帐套001[001, 2019]
        if TEXT == "测试帐套001[001, 2019]":
            self.WriteXlsx(4, 11, "PASS")
        else:
            self.WriteXlsx(4, 11, "ERROR")
        self.WriteXlsx(4, 10, "Y")
    def Frame1(self):
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 账套选择-全部
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)

    def Frame2(self):
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 账套选择-全部
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)
    #科目映射窗口账套参照窗口展示
    def test_d_Subject_config(self):
        self.System_config()
        self.Frame2()
        Lab1 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[2]/div").text#年度
        Lab2 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[3]/div").text#账套号
        Lab3 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[4]/div").text#账套名称
        print("Lab1~Lab3:",Lab1,Lab2,Lab3)
        if Lab1 =="年度" and Lab2 =="账套号" and Lab3 =="账套名称":
            self.WriteXlsx(5, 11, "PASS")
        else:
            self.WriteXlsx(5, 11, "ERROR")
        self.WriteXlsx(5, 10, "Y")
    #科目映射窗口账套参照窗口数据展示
    def test_e_Subject_config(self):
        self.System_config()
        self.Frame2()
        Lab1_1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[2]").text#2019
        Lab1_2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text#001
        Lab2_1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[2]").text  # 2020
        Lab2_2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text  # 001
        Lab3_1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[2]").text  # 2019
        Lab3_2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[3]").text  # 002
        print('Lab1_1~Lab3_2:',Lab1_1,Lab1_2,Lab2_1,Lab2_2,Lab3_1,Lab3_2)
        if Lab1_1 == Lab3_1 =="2019":
            if Lab2_2 == Lab1_2=="001":
                if Lab2_1=="2020" and Lab3_2=="002":
                    self.WriteXlsx(6, 11, "PASS")
                else:
                    self.WriteXlsx(6, 11, "ERROR")
            else:
                self.WriteXlsx(6, 11, "ERROR")
        else:
            self.WriteXlsx(6, 11, "ERROR")
        self.WriteXlsx(6, 10, "Y")
    #科目映射窗口账套查询1
    def test_f_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()#--查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(7, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套名称查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()#查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text#测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 测试帐套001
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[4]/div").text  # 平行记帐测试账套002
        print('Lab1~Lab3:',Lab1,Lab2,Lab3)
        if Lab1==Lab2=="测试帐套001" and Lab3=="平行记帐测试账套002":
            self.WriteXlsx(7, 11, "PASS")
        else:
            self.WriteXlsx(7, 11, "ERROR")
        self.WriteXlsx(7, 10, "Y")
    #科目映射窗口账套查询2
    def test_g_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(8, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套名称查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text  # 测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 测试帐套001
        print('Lab1~Lab2:', Lab1, Lab2)
        if Lab1 == Lab2 == "测试帐套001" :
            self.WriteXlsx(8, 11, "PASS")
        else:
            self.WriteXlsx(8, 11, "ERROR")
        self.WriteXlsx(8, 10, "Y")
    #科目映射窗口账套查询3
    def test_h_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(9, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套名称查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//table[@id='MyTable']").text
        print('Lab:',Lab)
        if Lab =="":
            self.WriteXlsx(9, 11, "PASS")
        else:
            self.WriteXlsx(9, 11, "ERROR")
        self.WriteXlsx(9, 10, "Y")
    #科目映射窗口账套查询4
    def test_i_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(10, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套名称查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text  # 测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 测试帐套001
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[4]/div").text  # 平行记帐测试账套002
        print('Lab1~Lab3:', Lab1, Lab2, Lab3)
        if Lab1 == Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002":
            self.WriteXlsx(10, 11, "PASS")
        else:
            self.WriteXlsx(10, 11, "ERROR")
        self.WriteXlsx(10, 10, "Y")
    #科目映射窗口账套查询5
    def test_j_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(11, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套号查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text  # 测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 测试帐套001
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[4]/div").text  # 平行记帐测试账套002
        print('Lab1~Lab3:', Lab1, Lab2, Lab3)
        if Lab1 == Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002":
            self.WriteXlsx(11, 11, "PASS")
        else:
            self.WriteXlsx(11, 11, "ERROR")
        self.WriteXlsx(11, 10, "Y")
    #科目映射窗口账套查询6
    def test_k_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(12, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套号查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]/div").text  # 002
        print(Lab1)
        if Lab1 =="002":
            self.WriteXlsx(12, 11, "PASS")
        else:
            self.WriteXlsx(12, 11, "ERROR")
        self.WriteXlsx(12, 10, "Y")
    #科目映射窗口账套查询7
    def test_l_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(13, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套号查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//table[@id='MyTable']").text
        print('Lab:', Lab)
        if Lab == "":
            self.WriteXlsx(13, 11, "PASS")
        else:
            self.WriteXlsx(13, 11, "ERROR")
        self.WriteXlsx(13, 10, "Y")
    #科目映射窗口账套查询8
    def test_m_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按账套名称查询']").click()
        sleep(1)
        self.Replace(14, 8)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").send_keys(C["按账套号查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='acctnm']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text  # 测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 测试帐套001
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[4]/div").text  # 平行记帐测试账套002
        print('Lab1~Lab3:', Lab1, Lab2, Lab3)
        if Lab1 == Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002":
            self.WriteXlsx(14, 11, "PASS")
        else:
            self.WriteXlsx(14, 11, "ERROR")
        self.WriteXlsx(14, 10, "Y")
    # 科目映射窗口账套查询9
    def test_n_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按年度查询']").click()
        sleep(1)
        self.Replace(15, 8,)
        self.driver.find_element_by_xpath("//input[@id='year']").send_keys(C["按年度查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text  # 测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 测试帐套001
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[4]/div").text  # 平行记帐测试账套002
        print('Lab1~Lab3:', Lab1, Lab2, Lab3)
        if Lab1 == Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002":
            self.WriteXlsx(15, 11, "PASS")
        else:
            self.WriteXlsx(15, 11, "ERROR")
        self.WriteXlsx(15, 10, "Y")
    #科目映射窗口账套查询10
    def test_o_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按年度查询']").click()
        sleep(1)
        self.Replace(16, 8)
        self.driver.find_element_by_xpath("//input[@id='year']").send_keys(C["按年度查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text  # 测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[2]/div").text  #2020
        print("Lab1-Lab2:",Lab1,Lab2)
        if Lab1=="测试帐套001" and Lab2=="2020":
            self.WriteXlsx(16, 11, "PASS")
        else:
            self.WriteXlsx(16, 11, "ERROR")
        self.WriteXlsx(16, 10, "Y")
    #科目映射窗口账套查询11
    def test_p_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按年度查询']").click()
        sleep(1)
        self.Replace(17, 8)
        self.driver.find_element_by_xpath("//input[@id='year']").send_keys(C["按年度查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//table[@id='MyTable']").text
        print('Lab:', Lab)
        if Lab == "":
            self.WriteXlsx(17, 11, "PASS")
        else:
            self.WriteXlsx(17, 11, "ERROR")
        self.WriteXlsx(17, 10, "Y")
    #科目映射窗口账套查询12
    def test_q_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='ellipsis_table']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='按年度查询']").click()
        sleep(2)
        self.Replace(18, 8)
        self.driver.find_element_by_xpath("//input[@id='year']").send_keys(C["按年度查询"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab4 = self.driver.find_element_by_xpath("//table[@id='MyTable']").text
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='year']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text  # 测试帐套001
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[4]/div").text  # 测试帐套001
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[4]/div").text  # 平行记帐测试账套002
        print('Lab1~Lab3:', Lab1, Lab2, Lab3)
        print("Lab4:",Lab4)
        if Lab4=="":
            if Lab1 == Lab2 == "测试帐套001" and Lab3 == "平行记帐测试账套002":
                self.WriteXlsx(18, 11, "PASS")
            else:
                self.WriteXlsx(18, 11, "ERROR")
        else:
            self.WriteXlsx(18, 11, "ERROR")
        self.WriteXlsx(18, 10, "Y")
    #科目映射窗口账套选择
    def test_r_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        TEXT = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").get_attribute('value')  # 获取输入框内容
        print('TEXT:', TEXT)  # 测试帐套001[001, 2019]
        if TEXT =="测试帐套001[001, 2019]":
            self.WriteXlsx(19, 11, "PASS")
        else:
            self.WriteXlsx(19, 11, "ERROR")
        self.WriteXlsx(19, 10, "Y")
    #科目映射窗口账套选择2
    def test_s_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        TEXT = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").get_attribute('value')  # 获取输入框内容
        print('TEXT:', TEXT)  # 平行记帐测试账套002[002, 2019]
        if TEXT == "平行记帐测试账套002[002, 2019]":
            self.WriteXlsx(21, 11, "PASS")
        else:
            self.WriteXlsx(21, 11, "ERROR")
        self.WriteXlsx(21, 10, "Y")
    #科目映射窗口账套清空
    def test_t_Subject_config(self):
        self.System_config()
        self.Frame2()
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        TEXT1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").get_attribute('value')  # 获取输入框内容
        print('TEXT1:', TEXT1)  # 测试帐套001[001, 2019]
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[2]").click()  # 清空按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        TEXT2 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").get_attribute('value')  # 获取输入框内容
        print('TEXT2:', TEXT2)  # 平行记帐测试账套002[002, 2019]
        if TEXT1 == "测试帐套001[001, 2019]" and TEXT2=="":
            self.WriteXlsx(22, 11, "PASS")
        else:
            self.WriteXlsx(22, 11, "ERROR")
        self.WriteXlsx(22, 10, "Y")
    #科目映射窗口账套必选验证1
    def test_u_Subject_config(self):
        self.System_config()
        self.Frame1()
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#业务枚举多级-业务枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_body left']").text
        print(msg)
        if msg =="请选择账套信息!":
            self.WriteXlsx(23, 11, "PASS")
        else:
            self.WriteXlsx(23, 11, "ERROR")
        self.WriteXlsx(23, 10, "Y")
    #科目映射窗口账套必选验证2
    def test_v_Subject_config(self):
        self.System_config()
        self.Frame1()
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_body left']").text
        print(msg)
        if msg == "请选择账套信息!":
            self.WriteXlsx(24, 11, "PASS")
        else:
            self.WriteXlsx(24, 11, "ERROR")
        self.WriteXlsx(24, 10, "Y")
    #科目映射窗口账套必选验证3
    def test_w_Subject_config(self):
        self.System_config()
        self.Frame1()
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_body left']").text
        print(msg)
        if msg == "请选择账套信息!":
            self.WriteXlsx(25, 11, "PASS")
        else:
            self.WriteXlsx(25, 11, "ERROR")
        self.WriteXlsx(25, 10, "Y")
    #科目映射窗口账套必选验证4
    def test_x_Subject_config(self):
        self.System_config()
        self.Frame1()
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='BudgetTR']/td/input[2]").click()  # 预算科目:
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_body left']").text
        print(msg)
        if msg == "请选择账套信息!":
            self.WriteXlsx(26, 11, "PASS")
        else:
            self.WriteXlsx(26, 11, "ERROR")
        self.WriteXlsx(26, 10, "Y")
    #科目映射窗口账套必选验证5
    def test_y_Subject_config(self):
        self.System_config()
        self.Frame1()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_body left']").text
        print(msg)
        if msg == "请选择账套信息!":
            self.WriteXlsx(27, 11, "PASS")
        else:
            self.WriteXlsx(27, 11, "ERROR")
        self.WriteXlsx(27, 10, "Y")
class TestC2_02(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.33:8081"
        print("Test Start")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def System_config(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxa")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        # self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                # print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
        self.driver.find_element_by_xpath("//div[@title = '科目配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
    def Frame1_YW_Deve1(self):
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)

    #科目配置枚举窗口展示
    def test_a_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[2]/div").text#编码
        Lab2 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[3]/div").text  # 名称
        Lab3 = self.driver.find_element_by_xpath("//a[@id='mytree_1_a']/span[2]").text#业务枚举多级-业务枚举(单位A)
        print("Lab1~Lab3:",Lab1,Lab2,Lab3)
        if Lab1 =="编码" and Lab2=="名称":
            if Lab3=="业务枚举多级-业务枚举(单位A)":
                self.WriteXlsx(28, 11, "PASS")
            else:
                self.WriteXlsx(28, 11, "ERROR")
        else:
            self.WriteXlsx(28, 11, "ERROR")
        self.WriteXlsx(28, 10, "Y")
    #科目映射枚举窗口数据展示
    def test_b_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text#业务枚举一级A
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  #业务枚举一级B
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]").text  #业务枚举二级AA
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[3]").text  #业务枚举二级AB
        Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[5]/td[3]").text  #业务枚举三级AAA
        Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[6]/td[3]").text  #业务枚举三级AAB
        print("Lab1~Lab3",Lab1,Lab2,Lab3)
        print("Lab4~Lab6", Lab4, Lab5, Lab6)
        if Lab1=="业务枚举一级A" and Lab2=="业务枚举一级B" and Lab3=="业务枚举二级AA":
            if Lab4=="业务枚举二级AB" and Lab5=="业务枚举三级AAA" and Lab6=="业务枚举三级AAB":
                self.WriteXlsx(29, 11, "PASS")
            else:
                self.WriteXlsx(29, 11, "ERROR")
        else:
            self.WriteXlsx(29, 11, "ERROR")
        self.WriteXlsx(29, 10, "Y")
    #科目配置枚举窗口定位1
    def test_c_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.Replace(30, 8)
        self.driver.find_element_by_xpath("//div[@id='north']/div/ul/li/input").send_keys(C["定位栏输入"])#输入框
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()#查询
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//a[@id='mytree_3_a']/span[2]").text
        print(Lab)#业务枚举二级AA
        if Lab=="业务枚举二级AA":
            self.WriteXlsx(30, 11, "PASS")
        else:
            self.WriteXlsx(30, 11, "ERROR")
        self.WriteXlsx(30, 10, "Y")
    #科目配置枚举窗口定位2
    def test_d_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.Replace(31, 8)
        self.driver.find_element_by_xpath("//div[@id='north']/div/ul/li/input").send_keys(C["定位栏输入"])  # 输入框
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举二级AA
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  # 业务枚举三级AAA
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]").text  # 业务枚举三级AAB
        print("Lab1~Lab3:",Lab1,Lab2,Lab3)
        if Lab1=="业务枚举二级AA" and Lab2=="业务枚举三级AAA" and Lab3=="业务枚举三级AAB":
            self.WriteXlsx(31, 11, "PASS")
        else:
            self.WriteXlsx(31, 11, "ERROR")
        self.WriteXlsx(31, 10, "Y")
    #科目配置枚举窗口定位3
    def test_e_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.Replace(32, 8)
        self.driver.find_element_by_xpath("//div[@id='north']/div/ul/li/input").send_keys(C["定位栏输入"])  # 输入框
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//a[@id='mytree_6_a']/span[2]").text
        print(Lab)#业务枚举二级AB
        if Lab =="业务枚举二级AB":
            self.WriteXlsx(32, 11, "PASS")
        else:
            self.WriteXlsx(32, 11, "ERROR")
        self.WriteXlsx(32, 10, "Y")
    #科目配置枚举窗口定位4
    def test_f_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.Replace(33, 8)
        self.driver.find_element_by_xpath("//div[@id='north']/div/ul/li/input").send_keys(C["定位栏输入"])  # 输入框
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举二级AB
        print(Lab1)
        if Lab1 =="业务枚举二级AB":
            self.WriteXlsx(33, 11, "PASS")
        else:
            self.WriteXlsx(33, 11, "ERROR")
        self.WriteXlsx(33, 10, "Y")
    #科目配置枚举窗口定位5
    def test_g_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_7_a']").click()#业务枚举一级B
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//a[@id='mytree_7_a']/span[2]").text
        print(Lab1)
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举一级B
        print(Lab2)
        if Lab2 ==Lab1== "业务枚举一级B" :
            self.WriteXlsx(34, 11, "PASS")
        else:
            self.WriteXlsx(34, 11, "ERROR")
        self.WriteXlsx(34, 10, "Y")
    #科目配置枚举窗口定位6
    def test_h_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_2_a']").click()#业务枚举一级A
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举一级A
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  # 业务枚举二级AA
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]").text  # 业务枚举二级AB
        print("Lab1~Lab3:",Lab1,Lab2,Lab3)
        if Lab1 =="业务枚举一级A" and Lab2=="业务枚举二级AA" and Lab3=="业务枚举二级AB":
            self.WriteXlsx(35, 11, "PASS")
        else:
            self.WriteXlsx(35, 11, "ERROR")
        self.WriteXlsx(35, 10, "Y")
    #科目配置枚举窗口定位7
    def test_i_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_2_a']").click()  # 业务枚举二级AA
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_switch']").click()#展开下级
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_3_span']").click()#业务枚举二级AA
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举一级B
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  # 业务枚举三级AAA
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]").text  # 业务枚举三级AAB
        print("Lab1~Lab3:", Lab1, Lab2, Lab3)
        if Lab1 == "业务枚举二级AA" and Lab2 == "业务枚举三级AAA" and Lab3 == "业务枚举三级AAB":
            self.WriteXlsx(36, 11, "PASS")
        else:
            self.WriteXlsx(36, 11, "ERROR")
        self.WriteXlsx(36, 10, "Y")
    #科目配置枚举窗口定位8
    def test_j_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举一级A
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  # 业务枚举一级B
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]").text  # 业务枚举二级AA
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[3]").text  # 业务枚举二级AB
        Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[5]/td[3]").text  # 业务枚举三级AAA
        Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[6]/td[3]").text  # 业务枚举三级AAB
        print("Lab1~Lab3", Lab1, Lab2, Lab3)
        print("Lab4~Lab6", Lab4, Lab5, Lab6)
        if Lab1 == "业务枚举一级A" and Lab2 == "业务枚举一级B" and Lab3 == "业务枚举二级AA":
            if Lab4 == "业务枚举二级AB" and Lab5 == "业务枚举三级AAA" and Lab6 == "业务枚举三级AAB":
                self.WriteXlsx(37, 11, "PASS")
            else:
                self.WriteXlsx(37, 11, "ERROR")
        else:
            self.WriteXlsx(37, 11, "ERROR")
        self.WriteXlsx(37, 10, "Y")
    #科目配置枚举窗口查询1
    def test_k_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td[1]").click()#查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='名称']").click()
        sleep(1)
        self.Replace(38, 8)
        self.driver.find_element_by_xpath("//li[@id='name_container']/input").send_keys(C["选择名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()#查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举二级AA
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  # 业务枚举二级AB
        print('Lab1~Lab2:',Lab1,Lab2)
        if Lab1 =="业务枚举二级AA" and Lab2=="业务枚举二级AB":
            self.WriteXlsx(38, 11, "PASS")
        else:
            self.WriteXlsx(38, 11, "ERROR")
        self.WriteXlsx(38, 10, "Y")
    #科目配置枚举窗口查询2
    def test_l_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td[1]").click()#查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='名称']").click()
        sleep(1)
        self.Replace(39, 8)
        self.driver.find_element_by_xpath("//li[@id='name_container']/input").send_keys(C["选择名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()#查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举二级AA
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  # 业务枚举三级AAA
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]").text  # 业务枚举三级AAB
        print('Lab1~Lab2:',Lab1,Lab2,Lab3)
        if Lab1 =="业务枚举二级AA" and Lab2=="业务枚举三级AAA" and Lab3=="业务枚举三级AAB":
            self.WriteXlsx(39, 11, "PASS")
        else:
            self.WriteXlsx(39, 11, "ERROR")
        self.WriteXlsx(39, 10, "Y")
    # 科目配置枚举窗口查询3
    def test_m_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@class='common_drop_list_title']/a/table/tbody/tr/td[1]").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='名称']").click()
        sleep(1)
        self.Replace(40, 8)
        self.driver.find_element_by_xpath("//li[@id='name_container']/input").send_keys(C["选择名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']").text  # 空字符串
        print('Lab1:', Lab1)
        if Lab1 == "":
            self.WriteXlsx(40, 11, "PASS")
        else:
            self.WriteXlsx(40, 11, "ERROR")
        self.WriteXlsx(40, 10, "Y")
    # 科目配置枚举窗口查询4
    def test_n_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath(
            "//div[@class='common_drop_list_title']/a/table/tbody/tr/td[1]").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='名称']").click()
        sleep(1)
        self.Replace(41, 8)
        self.driver.find_element_by_xpath("//li[@id='name_container']/input").send_keys(C["选择名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询按钮
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//table[@id='mytable']").text  # 空字符串
        sleep(1)
        self.driver.find_element_by_xpath("//li[@id='name_container']/input").clear()#
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text  # 业务枚举一级A
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]").text  # 业务枚举一级B
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]").text  # 业务枚举二级AA
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[3]").text  # 业务枚举二级AB
        Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[5]/td[3]").text  # 业务枚举三级AAA
        Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[6]/td[3]").text  # 业务枚举三级AAB
        print("Lab1~Lab3", Lab1, Lab2, Lab3)
        print("Lab4~Lab6", Lab4, Lab5, Lab6)
        if Lab1 == "业务枚举一级A" and Lab2 == "业务枚举一级B" and Lab3 == "业务枚举二级AA" and Lab=="":
            if Lab4 == "业务枚举二级AB" and Lab5 == "业务枚举三级AAA" and Lab6 == "业务枚举三级AAB":
                self.WriteXlsx(41, 11, "PASS")
            else:
                self.WriteXlsx(41, 11, "ERROR")
        else:
            self.WriteXlsx(41, 11, "ERROR")
        self.WriteXlsx(41, 10, "Y")
    #科目配置枚举窗口选择1
    def test_o_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").text
        print('Lab1:',Lab1)
        inputBox = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]")
        # 开始模拟鼠标双击操作
        ActionChains(self.driver).double_click(inputBox).perform()
        self.driver.switch_to.default_content()
        sleep(1)
        # self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#确定按钮
        # sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").get_attribute("value")
        print(Lab2)
        if Lab1==Lab2=="业务枚举一级A":
            self.WriteXlsx(42, 11, "PASS")
        else:
            self.WriteXlsx(42, 11, "ERROR")
        self.WriteXlsx(42, 10, "Y")
    #科目配置枚举窗口选择3
    def test_p_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").click()#业务枚举一级A
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").get_attribute("value")
        print(Lab2)#业务枚举一级A
        if Lab2 =="业务枚举一级A":
            self.WriteXlsx(44, 11, "PASS")
        else:
            self.WriteXlsx(44, 11, "ERROR")
        self.WriteXlsx(44, 10, "Y")
    #科目配置枚举窗口清空
    def test_q_Subject_config(self):
        self.System_config()
        self.Frame1_YW_Deve1()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        inputBox = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]")#业务枚举一级A
        # 开始模拟鼠标双击操作
        ActionChains(self.driver).double_click(inputBox).perform()
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").get_attribute("value")
        print(Lab1)  # 业务枚举一级A
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        # self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]").click()  # 业务枚举一级A
        # sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[2]").click()  # 清空按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").get_attribute("value")
        print('Lab2:',Lab2)#空
        if Lab1 =="业务枚举一级A" and Lab2=="":
            self.WriteXlsx(45, 11, "PASS")
        else:
            self.WriteXlsx(45, 11, "ERROR")
        self.WriteXlsx(45, 10, "Y")
    def Deve_Frame1(self):
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        # self.driver.find_element_by_xpath("//select[@id='SubMapping1']/option[3]").click()#业务枚举多级-业务枚举(单位A)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)
    #科目配置枚举必填验证
    def test_r_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.default_content()
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg)#请选填业务枚举多级-业务枚举(单位A)!
        if msg =="请选填业务枚举多级-业务枚举(单位A)!":
            self.WriteXlsx(46, 11, "PASS")
        else:
            self.WriteXlsx(46, 11, "ERROR")
        self.WriteXlsx(46, 10, "Y")
    #科目配置部门参照1
    def test_s_Subject_config(self):
        self.System_config()
        self.Deve_Frame1()
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        Lab1 = self.driver.find_element_by_xpath("//span[@id='myTree_1_span']").text#全部
        Lab2 = self.driver.find_element_by_xpath("//span[@id='myTree_30_span']").text#办公室
        sleep(1)
        Lab3 = self.driver.find_element_by_xpath("//span[@id='myTree_37_span']").text#国务院
        Lab4 = self.driver.find_element_by_xpath("//span[@id='myTree_54_span']").text#文化和旅游部
        print("Lab1~Lab4:",Lab1,Lab2,Lab3,Lab4)
        if Lab1=="全部" and Lab2=="办公室":
            if Lab3=="国务院" and Lab4=="文化和旅游部":
                self.WriteXlsx(47, 11, "PASS")
            else:
                self.WriteXlsx(47, 11, "ERROR")
        else:
            self.WriteXlsx(47, 11, "ERROR")
        self.WriteXlsx(47, 10, "Y")
    #科目配置部门参照2
    def test_t_Subject_config(self):
        self.System_config()
        self.Deve_Frame1()
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()#展开按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//span[@id='myTree_38_span']").text
        print(Lab1)#财政部
        if Lab1=="财政部":
            self.WriteXlsx(48, 11, "PASS")
        else:
            self.WriteXlsx(48, 11, "ERROR")
        self.WriteXlsx(48, 10, "Y")
    #科目配置部门参照3
    def test_u_Subject_config(self):
        self.System_config()
        self.Deve_Frame1()
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()  # 展开按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//span[@id='myTree_38_span']").text
        print(Lab1)  # 财政部
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_38_switch']").click()# 展开按钮
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//span[@id='myTree_39_span']").text#办公厅
        Lab3 = self.driver.find_element_by_xpath("//span[@id='myTree_45_span']").text#国防司
        Lab4 = self.driver.find_element_by_xpath("//span[@id='myTree_47_span']").text#部署事业单位
        print("Lab2~Lab4:",Lab2,Lab3,Lab4)
        if Lab1=="财政部":
            if Lab2=="办公厅" and Lab3=="国防司" and Lab4=="部署事业单位":
                self.WriteXlsx(49, 11, "PASS")
            else:
                self.WriteXlsx(49, 11, "ERROR")
        else:
            self.WriteXlsx(49, 11, "ERROR")
        self.WriteXlsx(49, 10, "Y")
    #科目配置部门选择返回
    def test_v_Subject_config(self):
        self.System_config()
        self.Deve_Frame1()
        self.driver.find_element_by_xpath("//span[@id='myTree_2_switch']").click()#展开办公室
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_9_switch']").click()#展开国务院
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_10_switch']").click()#展开财政部
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_11_switch']").click()  # 展开财政部
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_20_switch']").click()#展开部署事业单位
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_3_check']").click()#财务室
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_11_check']").click()#办公厅
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_12_check']").click()#综合司
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_20_span']").click()#部署事业单位
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab = self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").get_attribute("value")
        print(Lab)
        if Lab=="办公室,财务室,财政部,综合司,条法司,税政司,关税司,预算司,国防司,会计司,部署事业单位,国库支付中心,关税政策研究中心,信息网络中心":
            self.WriteXlsx(57, 11, "PASS")
        else:
            self.WriteXlsx(57, 11, "ERROR")
        self.WriteXlsx(57, 10, "Y")
    #科目配置部门选择状态记录
    def test_w_Subject_config(self):
        self.System_config()
        self.Deve_Frame1()
        self.driver.find_element_by_xpath("//span[@id='myTree_1_check']").click()#全部
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        # self.driver.find_element_by_xpath("//span[@id='myTree_8_check']").click()#规划设计科
        sleep(1)
        # self.driver.find_element_by_xpath("//span[@id='myTree_26_check']").click()#文化和旅游部
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab = self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").get_attribute("value")
        print(Lab)
        if Lab == "办公室,财务室,综合科,监查室,局办公室,水政水资源科,规划设计科,国务院,财政部,办公厅,综合司," \
                  "条法司,税政司,关税司,预算司,国防司,会计司,部署事业单位,国库支付中心,关税政策研究中心,信息网络中心," \
                  "住建部,生态环境部,商务部,文化和旅游部,国家文物局,办公室,财务室,综合科,监查室,局办公室,水政水资源科," \
                  "规划设计科,国务院,财政部,办公厅,综合司,条法司,税政司,关税司,预算司,国防司,会计司,部署事业单位,国库支付中心," \
                  "关税政策研究中心,信息网络中心,住建部,生态环境部,商务部,文化和旅游部,国家文物局,办公室,财务室,综合科,监查室," \
                  "局办公室,水政水资源科,规划设计科,国务院,财政部,办公厅,综合司,条法司,税政司,关税司,预算司,国防司,会计司," \
                  "部署事业单位,国库支付中心,关税政策研究中心,信息网络中心,住建部,生态环境部,商务部,文化和旅游部,国家文物局":
            self.WriteXlsx(59, 11, "PASS")
        else:
            self.WriteXlsx(59, 11, "ERROR")
        self.WriteXlsx(59, 10, "Y")
    #科目配置部门清空
    def test_x_Subject_config(self):
        self.System_config()
        self.Deve_Frame1()
        # self.driver.find_element_by_xpath("//span[@id='myTree_2_check']").click()#办公室
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").get_attribute("value")
        print(Lab1)#办公室,财务室,综合科,监查室
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").get_attribute("value")
        print("Lab2:",Lab2)
        if Lab1=="办公室,财务室,综合科,监查室" and Lab2=="":
            self.WriteXlsx(61, 11, "PASS")
        else:
            self.WriteXlsx(61, 11, "ERROR")
        self.WriteXlsx(61, 10, "Y")

class TestC2_03(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.33:8081"
        print("Test Start")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def System_config(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxa")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        # self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                # print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
        self.driver.find_element_by_xpath("//div[@title = '科目配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
    def Frame2_CW_Deve1(self):
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)

    #科目配置财务科目参照页面展示
    def test_a_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//a[@id='mytree_2_a']").text#资产
        Lab2 = self.driver.find_element_by_xpath("//a[@id='mytree_93_a']").text#负债
        Lab3 = self.driver.find_element_by_xpath("//a[@id='mytree_138_a']").text#共同
        Lab4 = self.driver.find_element_by_xpath("//a[@id='mytree_146_a']").text#权益
        Lab5 = self.driver.find_element_by_xpath("//a[@id='mytree_154_a']").text#成本
        Lab6 = self.driver.find_element_by_xpath("//a[@id='mytree_162_a']").text#损益
        print("Lab1~Lab6:",Lab1,Lab2,Lab3,Lab4,Lab5,Lab6)
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[2]/div").text#科目编码
        Msg2 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[3]/div").text#科目名称
        Msg3 = self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[4]/div").text#自动编号
        print("Mg1~Mg3:",Msg1,Msg2,Msg3)
        if Lab1=="资产" and Lab2=="负债" and Lab3=="共同" and Lab4=="权益" and Lab5=="成本" and Lab6=="损益":
            if Msg1=="科目编码" and Msg2=="科目名称" and Msg3=="自动编号":
                self.WriteXlsx(62,11,"PASS")
            else:
                self.WriteXlsx(62,11,"ERROR")
        else:
            self.WriteXlsx(62, 11, "ERROR")
        self.WriteXlsx(62, 10, "Y")
    #科目配置财务科目定位1
    def test_b_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.Replace(63,8)
        self.driver.find_element_by_xpath("//li[@class='common_search_input']/input").send_keys(C["定位框中录入"])
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()#查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//span[@id='mytree_20_span']").text
        print(Lab1)#1101 交易性金融资产
        # Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']").text#空
        # print("Lab2:",Lab2)
        if Lab1=="1101 交易性金融资产":
            self.WriteXlsx(63, 11, "PASS")
        else:
            self.WriteXlsx(63, 11, "ERROR")
        self.WriteXlsx(63, 10, "Y")
    #科目配置财务科目定位2
    def test_c_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.Replace(64, 8)
        self.driver.find_element_by_xpath("//li[@class='common_search_input']/input").send_keys(C["定位框中录入"])
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[2]").text#2221
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text#应交税费
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[2]").text  # 222101
        Lab4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text  # 应交增值税
        Lab5 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[2]").text  # 22210101
        Lab6 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[3]").text  # 进项税额
        Lab7 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[2]").text  # 22210102
        Lab8 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[3]").text  # 销项税额
        print("Lab1~Lab8:",Lab1,Lab2,Lab3,Lab4,Lab5,Lab6,Lab7,Lab8)
        if Lab1 =="2221" and Lab2 =="应交税费" and Lab3=="222101" and Lab4=="应交增值税":
            if Lab5=="22210101" and Lab6=="进项税额" and Lab7=="22210102" and Lab8=="销项税额":
                self.WriteXlsx(64, 11, "PASS")
            else:
                self.WriteXlsx(64, 11, "ERROR")
        else:
            self.WriteXlsx(64, 11, "ERROR")
        self.WriteXlsx(64, 10, "Y")
    #科目配置财务科目定位，手工定位1
    def test_d_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//a[@id='mytree_1_a']").click()#科目
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text#条/共56条记录
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[2]").text  #共3页
        print("Lab1~Lab2:",Lab1,Lab2)
        if Lab1=="条/共188条记录" and Lab2=="共10页":
            self.WriteXlsx(65, 11, "PASS")
        else:
            self.WriteXlsx(65, 11, "ERROR")
        self.WriteXlsx(65, 10, "Y")
    #科目配置财务科目定位，手工定位2
    def test_e_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_span']").click()#资产
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共26条记录
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[2]").text  # 共2页
        print("Lab1~Lab2:", Lab1, Lab2)
        if Lab1 == "条/共90条记录" and Lab2 == "共5页":
            self.WriteXlsx(66, 11, "PASS")
        else:
            self.WriteXlsx(66, 11, "ERROR")
        self.WriteXlsx(66, 10, "Y")
    #科目配置财务科目定位，手工定位3
    def test_f_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共14条记录
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[2]").text  # 共1页
        print("Lab1~Lab2:", Lab1, Lab2)
        if Lab1 == "条/共44条记录" and Lab2 == "共3页":
            self.WriteXlsx(67, 11, "PASS")
        else:
            self.WriteXlsx(67, 11, "ERROR")
        self.WriteXlsx(67, 10, "Y")
    #科目配置财务科目定位，手工定位4
    def test_g_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共6条记录
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[2]").text  # 共1页
        print("Lab1~Lab2:", Lab1, Lab2)
        if Lab1 == "条/共7条记录" and Lab2 == "共1页":
            self.WriteXlsx(68, 11, "PASS")
        else:
            self.WriteXlsx(68, 11, "ERROR")
        self.WriteXlsx(68, 10, "Y")
    # 科目配置财务科目定位，手工定位5
    def test_h_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()  # 权益
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共2条记录
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[2]").text  # 共1页
        print("Lab1~Lab2:", Lab1, Lab2)
        if Lab1 == "条/共7条记录" and Lab2 == "共1页":
            self.WriteXlsx(69, 11, "PASS")
        else:
            self.WriteXlsx(69, 11, "ERROR")
        self.WriteXlsx(69, 10, "Y")
    # 科目配置财务科目定位，手工定位6
    def test_i_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_162_span']").click()  # 损益
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共3条记录
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[2]").text  # 共1页
        print("Lab1~Lab2:", Lab1, Lab2)
        if Lab1 == "条/共33条记录" and Lab2 == "共2页":
            self.WriteXlsx(70, 11, "PASS")
        else:
            self.WriteXlsx(70, 11, "ERROR")
        self.WriteXlsx(70, 10, "Y")
    #科目配置财务科目定位，手工定位7
    def test_j_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_switch']").click()#展开资产列表
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_4_span']").click()#1002银行存款
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text#银行存款
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text  # 现金流量科目一
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[3]").text  # 现金流量科目2
        Lab4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[3]").text  # 银行科目A
        print("Lab1~Lab4:",Lab1,Lab2,Lab3,Lab4)
        if Lab1 =="银行存款" and Lab2=="现金流量科目一" and Lab3=="现金流量科目2" and Lab4=="银行科目A":
            self.WriteXlsx(71, 11, "PASS")
        else:
            self.WriteXlsx(71, 11, "ERROR")
        self.WriteXlsx(71, 10, "Y")
    #科目配置财务科目定位，手工定位8
    def test_k_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_switch']").click()  # 展开资产列表
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_22_span']").click()#1121 应收票据
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']").text#空字符
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共0条记录
        Lab3 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[2]").text  # 共1页
        print("Lab1:",Lab1)
        print("Lab2~Lab3:",Lab2,Lab3)
        if Lab2=="条/共1条记录" and Lab3=="共1页":
            self.WriteXlsx(72, 11, "PASS")
        else:
            self.WriteXlsx(72, 11, "ERROR")
        self.WriteXlsx(72, 10, "Y")
    #科目配置财务科目查询1
    def test_l_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()#科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()#查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(73, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()#查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text  #现金流量科目一
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text  # 现金流量科目2
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[3]").text  # 银行科目A
        print("Lab1~Lab3:",Lab1,Lab2,Lab3)
        if Lab1=="现金流量科目一" and Lab2=="现金流量科目2" and Lab3=="银行科目A":
            self.WriteXlsx(73, 11, "PASS")
        else:
            self.WriteXlsx(73, 11, "ERROR")
        self.WriteXlsx(73, 10, "Y")
    #科目配置财务科目查询1
    def test_m_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(74,8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共11条记录
        print(Lab1)
        if Lab1=="条/共12条记录":
            self.WriteXlsx(74, 11, "PASS")
        else:
            self.WriteXlsx(74, 11, "ERROR")
        self.WriteXlsx(74, 10, "Y")
    #科目配置财务科目查询3
    def test_n_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(75, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']").text  # 空字符
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共0条记录
        print(Lab2)
        print("Lab1:",Lab1)
        if Lab1=="" and Lab2=="条/共0条记录":
            self.WriteXlsx(75, 11, "PASS")
        else:
            self.WriteXlsx(75, 11, "ERROR")
        self.WriteXlsx(75, 10, "Y")
    #科目配置财务科目查询4
    def test_o_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(76,8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共0条记录
        print("Lab1:",Lab1)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共56条记录
        print("Lab2:",Lab2)
        if Lab1=="条/共0条记录" and Lab2=="条/共188条记录":
            self.WriteXlsx(76, 11, "PASS")
        else:
            self.WriteXlsx(76, 11, "ERROR")
        self.WriteXlsx(76, 10, "Y")
    #科目配置财务科目查询5
    def test_p_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(77, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text  #应交税费
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[3]").text  # 分保费用
        Lab3 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text#条/共5条记录
        print("Lab1~Lab3:",Lab1,Lab2,Lab3)
        if Lab3=="条/共18条记录":
            self.WriteXlsx(77, 11, "PASS")
        else:
            self.WriteXlsx(77, 11, "ERROR")
        self.WriteXlsx(77, 10, "Y")
    #科目配置财务科目查询6
    def test_q_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(78, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text  # 应交税费
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text  # 应交增值税
        print("Lab1,Lab2:",Lab1,Lab2)
        Lab3 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text
        if Lab3=="条/共4条记录":
            self.WriteXlsx(78, 11, "PASS")
        else:
            self.WriteXlsx(78, 11, "ERROR")
        self.WriteXlsx(78, 10, "Y")
    #科目配置财务科目查询7
    def test_r_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(79, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共0条记录
        print("Lab1:", Lab1)
        if Lab1=="条/共0条记录":
            self.WriteXlsx(79, 11, "PASS")
        else:
            self.WriteXlsx(79, 11, "ERROR")
        self.WriteXlsx(79, 10, "Y")
    #科目配置财务科目查询8
    def test_s_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 科目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(80, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共0条记录
        print("Lab1:", Lab1)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text  # 条/共56条记录
        print("Lab2:", Lab2)
        if Lab1=="条/共0条记录" and Lab2=="条/共188条记录":
            self.WriteXlsx(80, 11, "PASS")
        else:
            self.WriteXlsx(80, 11, "ERROR")
        self.WriteXlsx(80, 10, "Y")
    #科目配置财务科目查询9
    def test_t_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(81, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text  # 应付账款
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text  # 销项税额
        print("Lab1~Lab2:",Lab1,Lab2)
        Lab3 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text
        if Lab3=="条/共7条记录":
            self.WriteXlsx(81, 11, "PASS")
        else:
            self.WriteXlsx(81, 11, "ERROR")
        self.WriteXlsx(81, 10, "Y")
    #科目配置财务科目查询10
    def test_u_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(82, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text  # 货币兑换
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text  # 被套期项目
        print("Lab1~Lab2:", Lab1, Lab2)
        if Lab1 == "货币兑换" and Lab2 == "被套期项目":
            self.WriteXlsx(82, 11, "PASS")
        else:
            self.WriteXlsx(82, 11, "ERROR")
        self.WriteXlsx(82, 10, "Y")
    #科目配置财务科目查询11
    def test_v_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_162_span']").click()  # 损益
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(83,8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text  # 分保费用
        print(Lab1)
        Lab3 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text
        if Lab3 == "条/共11条记录":
            self.WriteXlsx(83, 11, "PASS")
        else:
            self.WriteXlsx(83, 11, "ERROR")
        self.WriteXlsx(83, 10, "Y")
    #科目配置财务科目查询12
    def test_w_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_span']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='common_drop_list_title']/a/table/tbody/tr/td").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(84, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text  # 银行存款
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[3]").text#应收账款
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[3]").text#其他应收款
        print("Lab1~Lab3:",Lab1,Lab2,Lab3)
        Lab4 = self.driver.find_element_by_xpath("//div[@class='pDiv']/div/span/span[1]").text
        if Lab4 == "条/共10条记录":
            self.WriteXlsx(84, 11, "PASS")
        else:
            self.WriteXlsx(84, 11, "ERROR")
        self.WriteXlsx(84, 10, "Y")
    #科目配置财务科目选择1
    def test_x_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text#清算资金往来
        print("Lab1:",Lab1)
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()
        # inputBox = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]")#选择清算资金往来 /div/input
        # # 开始模拟鼠标双击操作
        # ActionChains(self.driver).double_click(inputBox).perform()
        # ActionChains(self.driver).double_click(inputBox).perform()
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").get_attribute("value")
        print("Lab2:",Lab2)
        if Lab1== "清算资金往来" and Lab2=="3001[清算资金往来]":
            self.WriteXlsx(85, 11, "PASS")
        else:
            self.WriteXlsx(85, 11, "ERROR")
        self.WriteXlsx(85, 10, "Y")
    #科目配置财务科目选择2
    def test_y_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()  # 权益
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 实收资本
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]").text # 实收资本
        print("Lab1:", Lab1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").get_attribute("value")
        print("Lab2:",Lab2)#4001[实收资本]
        if Lab1=="实收资本" and Lab2=="4001[实收资本]":
            self.WriteXlsx(86, 11, "PASS")
        else:
            self.WriteXlsx(86, 11, "ERROR")
        self.WriteXlsx(86, 10, "Y")
        #科目配置财务科目清空_0086
    def test_z_Subject_config(self):
        self.System_config()
        self.Frame2_CW_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_162_span']").click()#损益
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()#下一页
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[1]/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").get_attribute("value")#6542[分保费用]
        print(Lab1)#6521[保单红利支出]
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[2]").click()
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").get_attribute(
            "value")  # 空
        print("Lab2:",Lab2)  # 空
        if Lab1=='6521[保单红利支出]' and Lab2=='':
            self.WriteXlsx(87, 11, "PASS")
        else:
            self.WriteXlsx(87, 11, "ERROR")
        self.WriteXlsx(87, 10, "Y")
    #科目配置财务科目必填验证_0087
    def test_zb_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#业务枚举多级-业务枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").get_attribute("value")
        print(Lab1)#业务枚举一级A
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()
        self.driver.switch_to.default_content()
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(msg)#财务科目信息和预算科目信息不能同时为空!请选择...
        sleep(1)
        if Lab1=='业务枚举一级A' and msg=='财务科目信息和预算科目信息不能同时为空!请选择...':
            self.WriteXlsx(88, 11, "PASS")
        else:
            self.WriteXlsx(88, 11, "ERROR")
        self.WriteXlsx(88, 10, "Y")

class TestC2_04(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.33:8081"
        print("Test Start")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\科目配置.xlsx"
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["科目配置"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def System_config(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxa")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        # self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        sleep(1)
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                # print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
        self.driver.find_element_by_xpath("//div[@title = '科目配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
    def Frame2_SJ_Deve1(self):
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()#税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)
    #科目配置税金科目参照页面展示_0088
    def test_A_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        Lab1 = self.driver.find_element_by_xpath("//a[@id='mytree_2_a']").text  # 资产
        Lab2 = self.driver.find_element_by_xpath("//a[@id='mytree_93_a']").text  # 负债
        Lab3 = self.driver.find_element_by_xpath("//a[@id='mytree_138_a']").text  # 共同
        Lab4 = self.driver.find_element_by_xpath("//a[@id='mytree_146_a']").text  # 权益
        Lab5 = self.driver.find_element_by_xpath("//a[@id='mytree_154_a']").text  # 成本
        Lab6 = self.driver.find_element_by_xpath("//a[@id='mytree_162_a']").text  # 损益
        print("Lab1~Lab6:", Lab1, Lab2, Lab3, Lab4, Lab5, Lab6)
        Lab7 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print(Lab7)#条/共56条记录
        if Lab1 == "资产" and Lab2 == "负债" and Lab3 == "共同" and Lab4 == "权益" and Lab5 == "成本" and Lab6 == "损益":
            if Lab7=='条/共188条记录':
                self.WriteXlsx(89, 11, "PASS")
            else:
                self.WriteXlsx(89, 11, "ERROR")
        else:
            self.WriteXlsx(89, 11, "ERROR")
        self.WriteXlsx(89,10,"Y")
    #科目配置税金科目定位_0089
    def test_B_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.Replace(90,8)
        self.driver.find_element_by_xpath("//div[@class='layout_north']/div/ul/li/input").send_keys(C["定位框中录入"])
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()#定位
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//a[@id='mytree_20_a']").text#1101 交易性金融资产
        # Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']").text#空
        print(Lab1)
        # print("Lab2:",Lab2)
        if Lab1=="1101 交易性金融资产" :
            self.WriteXlsx(90, 11, "PASS")
        else:
            self.WriteXlsx(90, 11, "ERROR")
        self.WriteXlsx(90, 10, "Y")
    #科目配置税金科目定位_0090
    def test_C_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.Replace(91, 8)
        self.driver.find_element_by_xpath("//div[@class='layout_north']/div/ul/li/input").send_keys(C["定位框中录入"])
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 定位
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 定位
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='search_btn']").click()  # 定位
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//a[@id='mytree_107_a']").text#2221 应交税费
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text#2221
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text#应交税费
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 222101
        Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 应交增值税
        Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[2]/div").text  # 22210101
        Lab7 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]/div").text  # 进项税额
        Lab8 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[2]/div").text  # 22210102
        Lab9 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[3]/div").text  # 销项税额
        print("Lab1:",Lab1)
        print("Lab2~Lab5:",Lab2,Lab3,Lab4,Lab5)
        print("Lab6~Lab9:",Lab6,Lab7,Lab8,Lab9)
        if Lab1=="2221 应交税费":
            if Lab2=='2221' and Lab3=="应交税费" and Lab4=="222101" and Lab5=="应交增值税" and Lab6=="22210101":
                if Lab7=="进项税额" and Lab8=="22210102" and Lab9=="销项税额":
                    self.WriteXlsx(91, 11, "PASS")
                else:
                    self.WriteXlsx(91, 11, "ERROR")
            else:
                self.WriteXlsx(91, 11, "ERROR")
        else:
            self.WriteXlsx(91, 11, "ERROR")
        self.WriteXlsx(91, 10, "Y")
    #科目配置税金科目定位，手工定位_0091
    def test_D_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()#科目
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[2]").text
        print("Lab1:",Lab1)#条/共56条记录
        print("Lab2:",Lab2)#共3页
        if Lab1=="条/共188条记录" and Lab2=="共10页":
            self.WriteXlsx(92, 11, "PASS")
        else:
            self.WriteXlsx(92, 11, "ERROR")
        self.WriteXlsx(92, 10, "Y")
    #c科目配置税金科目定位，手工定位_0092
    def test_E_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_span']").click()#资产
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[2]").text
        print("Lab1:", Lab1)  # 条/共26条记录
        print("Lab2:", Lab2)  # 共2页
        if Lab1=="条/共90条记录" and Lab2=="共5页":
            self.WriteXlsx(93, 11, "PASS")
        else:
            self.WriteXlsx(93, 11, "ERROR")
        self.WriteXlsx(93, 10, "Y")
    #科目配置税金科目定位，手工定位_0093
    def test_F_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[2]").text
        print("Lab1:", Lab1)  # 条/共14条记录
        print("Lab2:", Lab2)  # 共1页
        if Lab1 == "条/共44条记录" and Lab2 == "共3页":
            self.WriteXlsx(94, 11, "PASS")
        else:
            self.WriteXlsx(94, 11, "ERROR")
        self.WriteXlsx(94, 10, "Y")
    #科目配置税金科目定位，手工定位_0094
    def test_G_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[2]").text
        print("Lab1:", Lab1)  # 条/共6条记录
        print("Lab2:", Lab2)  # 共1页
        if Lab1 == "条/共7条记录" and Lab2 == "共1页":
            self.WriteXlsx(95, 11, "PASS")
        else:
            self.WriteXlsx(95, 11, "ERROR")
        self.WriteXlsx(95, 10, "Y")
    #科目配置税金科目定位，手工定位_0095
    def test_H_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()  # 权益
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[2]").text
        print("Lab1:", Lab1)  # 条/共2条记录
        print("Lab2:", Lab2)  # 共1页
        if Lab1 == "条/共7条记录" and Lab2 == "共1页":
            self.WriteXlsx(96, 11, "PASS")
        else:
            self.WriteXlsx(96, 11, "ERROR")
        self.WriteXlsx(96, 10, "Y")
    #科目配置税金科目定位，手工定位_0096
    def test_I_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_162_span']").click()  # 损益
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[2]").text
        print("Lab1:", Lab1)  # 条/共3条记录
        print("Lab2:", Lab2)  # 共1页
        if Lab1 == "条/共33条记录" and Lab2 == "共2页":
            self.WriteXlsx(97, 11, "PASS")
        else:
            self.WriteXlsx(97, 11, "ERROR")
        self.WriteXlsx(97, 10, "Y")
    #科目配置税金科目定位，手工定位_0097
    def test_J_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_switch']").click()#展开资产栏
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_4_span']").click()#1002 银行存款
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text#1002
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text#银行存款
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 100201
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 现金流量科目一
        Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[2]/div").text  # 100202
        Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]/div").text  # 现金流量科目2
        Lab7 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[2]/div").text  # 100203
        Lab8 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[3]/div").text  # 银行科目A
        print("Lab1~Lab4:",Lab1,Lab2,Lab3,Lab4)
        print("Lab5~Lab8:",Lab5,Lab6,Lab7,Lab8)
        if Lab1 =="1002" and Lab2 =="银行存款" and Lab3 =="100201" and Lab4=="现金流量科目一":
            if Lab5=="100202" and Lab6=="现金流量科目2" and Lab7=="100203" and Lab8=="银行科目A":
                self.WriteXlsx(98, 11, "PASS")
            else:
                self.WriteXlsx(98, 11, "ERROR")
        else:
            self.WriteXlsx(98, 11, "ERROR")
        self.WriteXlsx(98, 10, "Y")
    #科目配置税金科目定位，手工定位_0098
    def test_K_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_switch']").click()  # 展开资产栏
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_22_span']").click()  # 1121 应收票据
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//a[@id='mytree_22_a']").text
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']").text  # 空
        print(Lab1)#1121 应收票据
        # print("Lab2:", Lab2)
        if Lab1=="1121 应收票据":
            self.WriteXlsx(99, 11, "PASS")
        else:
            self.WriteXlsx(99, 11, "ERROR")
        self.WriteXlsx(99, 10, "Y")
    #科目配置税金科目查询_0099
    def test_L_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()#查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(100,8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text  # 100201
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text  # 现金流量科目一
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 100202
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 现金流量科目2
        Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[2]/div").text  # 100203
        Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]/div").text  # 银行科目A
        print("Lab1~Lab6:",Lab1,Lab2,Lab3,Lab4,Lab5,Lab6)
        if Lab1=="100201" and Lab2=="现金流量科目一" and Lab3=="100202" and Lab4=="现金流量科目2":
            if Lab5=="100203" and Lab6=="银行科目A":
                self.WriteXlsx(100, 11, "PASS")
            else:
                self.WriteXlsx(100, 11, "ERROR")
        else:
            self.WriteXlsx(100, 11, "ERROR")
        self.WriteXlsx(100, 10, "Y")
    #科目配置税金科目查询_0100
    def test_M_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(101, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab1:",Lab1)#条/共11条记录
        if Lab1=="条/共12条记录":
            self.WriteXlsx(101, 11, "PASS")
        else:
            self.WriteXlsx(101, 11, "ERROR")
        self.WriteXlsx(101, 10, "Y")
    #科目配置税金科目查询_0101
    def test_N_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(102, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()#查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab1:", Lab1)  # 条/共0条记录
        if Lab1 == "条/共0条记录":
            self.WriteXlsx(102, 11, "PASS")
        else:
            self.WriteXlsx(102, 11, "ERROR")
        self.WriteXlsx(102, 10, "Y")
    #科目配置税金科目查询_0102
    def test_O_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(103, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab1:", Lab1)  # 条/共0条记录
        sleep(1)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab2:", Lab2)  # 条/共56条记录
        if Lab1=="条/共0条记录" and Lab2=="条/共188条记录":
            self.WriteXlsx(103, 11, "PASS")
        else:
            self.WriteXlsx(103, 11, "ERROR")
        self.WriteXlsx(103, 10, "Y")
    #科目配置税金科目查询_0103
    def test_P_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(104, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        # Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text  # 2221
        # Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text  # 应交税费
        # Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 2991
        # Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 差旅费科目
        # Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[2]/div").text  # 2992
        # Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]/div").text  # 维修费科目
        # Lab7 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[2]/div").text  # 2993
        # Lab8 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[4]/td[3]/div").text  # 招待费科目
        # Lab9 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[5]/td[2]/div").text  # 6542
        # Lab10 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[5]/td[3]/div").text  # 分保费用
        # Lab11 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        # print("Lab11:",Lab11)#条/共5条记录
        # print("Lab1~Lab6:", Lab1, Lab2, Lab3, Lab4, Lab5, Lab6)
        # print("Lab7~Lab10:",Lab7,Lab8,Lab9,Lab10)
        Lab11 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        if Lab11=="条/共18条记录":
            # if Lab1=="2221" and Lab2=="应交税费" and Lab3=="2991" and Lab4=="差旅费科目" and Lab5=="2992":
            #     if Lab6=="维修费科目" and Lab7=="2993" and Lab8=="招待费科目" and Lab9=="6542" and Lab10=="分保费用":
            self.WriteXlsx(104, 11, "PASS")
        else:
            self.WriteXlsx(104, 11, "ERROR")
        #     else:
        #         self.WriteXlsx(104, 11, "ERROR")
        # else:
        #     self.WriteXlsx(104, 11, "ERROR")
        self.WriteXlsx(104, 10, "Y")
    #科目配置税金科目查询_0104
    def test_Q_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(105, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text  # 2221
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text  # 应交税费
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 222101
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 应交增值税
        Lab11 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab1~Lab4:",Lab1,Lab2,Lab3,Lab4)
        print("Lab11:", Lab11)  # 条/共2条记录
        if Lab11=="条/共4条记录":
            # if Lab1=="2221" and Lab2=="应交税费" and Lab3=="222101" and Lab4=="应交增值税":
            self.WriteXlsx(105, 11, "PASS")
        else:
            self.WriteXlsx(105, 11, "ERROR")
        # else:
        #     self.WriteXlsx(105, 11, "ERROR")
        self.WriteXlsx(105, 10, "Y")
    #科目配置税金科目查询_0105
    def test_R_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(106, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab11 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab11:", Lab11)  # 条/共0条记录
        if Lab11=="条/共0条记录":
            self.WriteXlsx(106, 11, "PASS")
        else:
            self.WriteXlsx(106, 11, "ERROR")
        self.WriteXlsx(106, 10, "Y")
    #科目配置税金科目查询_0106
    def test_S_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(107, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab1:", Lab1)  # 条/共0条记录
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab2:", Lab2)  # 条/共56条记录
        if Lab1=="条/共0条记录" and Lab2=="条/共188条记录":
            self.WriteXlsx(107, 11, "PASS")
        else:
            self.WriteXlsx(107, 11, "ERROR")
        self.WriteXlsx(107, 10, "Y")
    #科目配置税金科目查询_0107
    def test_T_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(108, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab5 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab5:", Lab5)  # 条/共2条记录
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text  # 2202
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text  # 应付账款
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 22210102
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 销项税额
        print("Lab1~Lab4:",Lab1,Lab2,Lab3,Lab4)
        if Lab5=="条/共7条记录" :
            # if Lab1=="2202" and Lab2=="应付账款" and Lab3=="22210102" and Lab4=="销项税额":
            self.WriteXlsx(108, 11, "PASS")
        else:
            self.WriteXlsx(108, 11, "ERROR")
        # else:
        #     self.WriteXlsx(108, 11, "ERROR")
        self.WriteXlsx(108, 10, "Y")
    #科目配置税金科目查询_0108
    def test_U_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目编码']").click()
        sleep(1)
        self.Replace(109, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_container']/input").send_keys(C["选择科目编码"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab5 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab5:", Lab5)  # 条/共2条记录
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text  # 3002
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text  # 货币兑换
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 3202
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 被套期项目
        print("Lab1~Lab4:", Lab1, Lab2, Lab3, Lab4)
        if Lab5 == "条/共2条记录":
            if Lab1 == "3002" and Lab2 == "货币兑换" and Lab3 == "3202" and Lab4 == "被套期项目":
                self.WriteXlsx(109, 11, "PASS")
            else:
                self.WriteXlsx(109, 11, "ERROR")
        else:
            self.WriteXlsx(109, 11, "ERROR")
        self.WriteXlsx(109, 10, "Y")
    #科目配置税金科目查询_0109
    def test_V_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_162_span']").click()  # 损益
        sleep(1)
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(110, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text  # 6542
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text  # 分保费用
        Lab3 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        print("Lab1~Lab3:", Lab1,Lab2,Lab3)  # 条/共1条记录
        if Lab3=="条/共11条记录":
            # if Lab1=="6542" and Lab2=="分保费用":
            self.WriteXlsx(110, 11, "PASS")
        else:
            self.WriteXlsx(110, 11, "ERROR")
        # else:
        #     self.WriteXlsx(110, 11, "ERROR")
        self.WriteXlsx(110, 10, "Y")
    #科目配置税金科目查询_0110
    def test_W_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_2_span']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//td[@class='common_drop_list_text font_size12']").click()  # 查询条件
        sleep(1)
        self.driver.find_element_by_xpath("//a[@title='科目名称']").click()
        sleep(1)
        self.Replace(111, 8)
        self.driver.find_element_by_xpath("//li[@id='ccode_name_container']/input").send_keys(C["选择科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//li[@class='margin_l_5 search_btn']/a").click()  # 查询按钮
        sleep(1)
        Lab10 = self.driver.find_element_by_xpath("//div[@class='layout_center']/div/div[6]/div/span/span[1]").text
        Lab1 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[2]/div").text  # 1002
        Lab2 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[3]/div").text  # 银行存款
        Lab3 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[2]/div").text  # 1122
        Lab4 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[3]/div").text  # 应收账款
        Lab5 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[2]/div").text  # 1221
        Lab6 = self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[3]/div").text  # 其他应收款
        print("Lab10:",Lab10)#条/共3条记录
        print("Lab1~Lab6:",Lab1,Lab2,Lab3,Lab4,Lab5,Lab6)
        if Lab10 =="条/共10条记录":
            # if Lab1=="1002" and Lab2=="银行存款" and Lab3=="1122" and Lab4=="应收账款" and Lab5=="1221" and Lab6=="其他应收款":
            self.WriteXlsx(111, 11, "PASS")
        else:
            self.WriteXlsx(111, 11, "ERROR")
        # else:
        #     self.WriteXlsx(111, 11, "ERROR")
        self.WriteXlsx(111, 10, "Y")
    #科目配置税金科目选择_0111
    def test_X_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()#清算资金往来
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").get_attribute("value")
        print(Lab)#3001[清算资金往来]
        if Lab=="3001[清算资金往来]":
            self.WriteXlsx(112, 11, "PASS")
        else:
            self.WriteXlsx(112, 11, "ERROR")
        self.WriteXlsx(112, 10, "Y")
    #科目配置税金科目选择_0112
    def test_Y_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()  # 权益
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 实收资本
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").get_attribute("value")
        print(Lab)  # 4001[实收资本]
        if Lab == "4001[实收资本]":
            self.WriteXlsx(113, 11, "PASS")
        else:
            self.WriteXlsx(113, 11, "ERROR")
        self.WriteXlsx(113, 10, "Y")
    #科目配置税金科目清空_0113
    def test_Z_Subject_config(self):
        self.System_config()
        self.Frame2_SJ_Deve1()
        self.driver.find_element_by_xpath("//span[@id='mytree_162_span']").click()  # 损益
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[1]/div/input").click()#分保费用
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定按钮
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").get_attribute("value")
        print(Lab1)  # 6011[利息收入]
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame3)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[2]").click()  # 清空按钮
        sleep(1)
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").get_attribute("value")
        print("Lab2:",Lab2) # 空
        if Lab1 == "6011[利息收入]" and Lab2=="":
            self.WriteXlsx(114, 11, "PASS")
        else:
            self.WriteXlsx(114, 11, "ERROR")
        self.WriteXlsx(114, 10, "Y")
    #
    def Frame1_to_Select_A(self):#选择业务枚举多级-业务枚举(单位A)
        '''年度选择2019，映射类型选择业务枚举多级-业务枚举(单位A),账套选择001，新增进入科目配置窗口'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()#业务枚举一级A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.frame(frame1)

    #账套+年度+对象+部门重复校验_0114
    def test_a_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()#共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 清算资金往来
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)#保存成功
        if Msg=="保存成功!":
            self.WriteXlsx(115, 11, "PASS")
        else:
            self.WriteXlsx(115, 11, "ERROR")
        self.WriteXlsx(115, 10, "Y")
    #账套+年度+对象+部门重复校验_0115
    def test_b_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()  # 权益
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 实收资本
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)#数据重复!请确认...
        sleep(1)
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(116, 11, "PASS")
        else:
            self.WriteXlsx(116, 11, "ERROR")
        self.WriteXlsx(116, 10, "Y")
    #账套+年度+对象+部门重复校验_0116
    def test_c_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()#部门名称:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()#综合科
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()#确定
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe4")
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()  # 权益
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 实收资本
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据重复!请确认...
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(117, 11, "PASS")
        else:
            self.WriteXlsx(117, 11, "ERROR")
        self.WriteXlsx(117, 10, "Y")
    #账套+年度+对象+部门重复校验_0117
    def test_d_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()#2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 业务枚举一级A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 清算资金往来
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(118, 11, "PASS")
        else:
            self.WriteXlsx(118, 11, "ERROR")
        self.WriteXlsx(118, 10, "Y")
    #账套+年度+对象+部门重复校验_0118
    def test_e_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[3]").click()  # 平行记帐测试账套002[002]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.switch_to.default_content()
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 业务枚举一级A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_63_span']").click()  # 净资产
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 限定性净资产
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(119, 11, "PASS")
        else:
            self.WriteXlsx(119, 11, "ERROR")
        self.WriteXlsx(119, 10, "Y")
    #账套+年度+对象+部门重复校验_0119
    def test_f_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr/td[1]/div/input").click()#
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()#删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)#数据删除后将无法恢复,请确认是否进行删除!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()#确定
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg2)#删除成功!
        if Msg1=="数据删除后将无法恢复,请确认是否进行删除!" and Msg2=="删除成功!":
            self.WriteXlsx(120, 11, "PASS")
        else:
            self.WriteXlsx(120, 11, "ERROR")
        self.WriteXlsx(120, 10, "Y")
    #账套+年度+对象+部门重复校验_0120
    def test_g_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门名称:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe4")
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()  # 权益
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 4001实收资本
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        if Msg == "保存成功!":
            self.WriteXlsx(121, 11, "PASS")
        else:
            self.WriteXlsx(121, 11, "ERROR")
        self.WriteXlsx(121, 10, "Y")
    #账套+年度+对象+部门重复校验_0121
    def test_h_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 3001清算资金往来
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 	数据重复!请确认...
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(122, 11, "PASS")
        else:
            self.WriteXlsx(122, 11, "ERROR")
        self.WriteXlsx(122, 10, "Y")
    #账套+年度+对象+部门重复校验_0122
    def test_i_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门名称:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()#监察室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe4")
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 3001清算资金往来
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        if Msg == "保存成功!":
            self.WriteXlsx(123, 11, "PASS")
        else:
            self.WriteXlsx(123, 11, "ERROR")
        self.WriteXlsx(123, 10, "Y")
    #账套+年度+对象+部门重复校验_0123
    def test_j_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门名称:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()  # 办公室
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe4")
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 2001短期借款
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # "数据重复!请确认..."
        if Msg == "数据重复!请确认...":
            self.WriteXlsx(124, 11, "PASS")
        else:
            self.WriteXlsx(124, 11, "ERROR")
        self.WriteXlsx(124, 10, "Y")
    #账套+年度+对象+部门重复校验_0124
    def test_k_Subject_config(self):
        self.System_config()
        self.Frame1_to_Select_A()
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门名称:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe3")
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_34_check']").click()  # 局办公室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe4")
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 3001清算资金往来
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        if Msg == "保存成功!":
            self.WriteXlsx(125, 11, "PASS")
        else:
            self.WriteXlsx(125, 11, "ERROR")
        self.WriteXlsx(125, 10, "Y")
    #账套+年度+对象+部门重复校验_0125
    def test_l_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()#全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()#删除
        self.driver.switch_to.default_content()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text#
        print(Msg1)#数据删除后将无法恢复,请确认是否进行删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        # self.driver.switch_to.default_content()
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text#
        print(Msg2)#删除成功!
        if Msg1=="数据删除后将无法恢复,请确认是否进行删除!" and Msg2=="删除成功!":
            self.WriteXlsx(126, 11, "PASS")
        else:
            self.WriteXlsx(126, 11, "ERROR")
        self.WriteXlsx(126, 10, "Y")
    #账套+年度+对象+部门重复校验_0126
    def test_m_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()#2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        self.driver.switch_to.default_content()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(127, 11, "PASS")
        else:
            self.WriteXlsx(127, 11, "ERROR")
        self.WriteXlsx(127, 10, "Y")
    def DeM_A(self):#选择多级枚举-单位枚举(单位A)
        '''年度选择2019，映射类型选择多级枚举-单位枚举(单位A),账套选择001，新增进入科目配置窗口'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()#新增
        sleep(1)
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
    #新增单位枚举科目配置_0127
    def test_n_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#第一行第一个 一级枚举A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()#资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()#最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()# 一级枚举科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()#确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[6]").text
        print(Lab1)#1191
        if Msg1=="保存成功!":
            self.WriteXlsx(128, 11, "PASS")
        else:
            self.WriteXlsx(128, 11, "ERROR")
        self.WriteXlsx(128, 10, "Y")
    #新增单位枚举科目配置_0128
    def test_o_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举AA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[6]").text
        print(Lab1)  # 1192
        if Msg1 == "保存成功!":
            self.WriteXlsx(129, 11, "PASS")
        else:
            self.WriteXlsx(129, 11, "ERROR")
        self.WriteXlsx(129, 10, "Y")
    #新增单位枚举科目配置_0129
    def test_p_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[1]/div/input").click()  # 三级枚举AAA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        Frame4 = "layui-layer-iframe4"
        Frame5 = "layui-layer-iframe5"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()#部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_2_check']").click()#办公室
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()#综合科
        self.driver.find_element_by_xpath("//span[@id='myTree_35_check']").click()#水政水资源科
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()#展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_38_switch']").click()#展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_39_check']").click()#办公厅
        self.driver.find_element_by_xpath("//span[@id='myTree_40_check']").click()#综合司
        self.driver.find_element_by_xpath("//span[@id='myTree_41_check']").click()#条法司
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame4)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[1]/div/input").click()  # 1193[三级枚举科目]
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()#税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame5)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#2001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer5']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[6]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print(Lab1,Lab2)  # 1193  办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Msg1 == "保存成功!" :
            self.WriteXlsx(130, 11, "PASS")
        else:
            self.WriteXlsx(130, 11, "ERROR")
        self.WriteXlsx(130, 10, "Y")
    #新增单位枚举科目配置_0130
    def test_q_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()#三级枚举AAA
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增
        sleep(1)
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()#部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()  # 监察室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()#确定
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[6]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[12]").text
        print(Lab1, Lab2)  # 1193  监察室
        if Msg1 == "保存成功!" :
            self.WriteXlsx(131, 11, "PASS")
        else:
            self.WriteXlsx(131, 11, "ERROR")
        self.WriteXlsx(131, 10, "Y")
    def Frame1_to(self):
        '''年度选择2019，映射类型选择业务枚举多级-业务枚举(单位A),账套选择001，新增进入科目配置窗口'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()#2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)

    #新增业务枚举科目配置_0131
    def test_r_Subject_config(self):
        self.System_config()
        self.Frame1_to()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 业务枚举多级-业务枚举s
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一行第一个 一级枚举A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()  # 一级枚举科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[6]").text
        print(Lab1)  # 1191
        if Msg1 == "保存成功!":
            self.WriteXlsx(132, 11, "PASS")
        else:
            self.WriteXlsx(132, 11, "ERROR")
        self.WriteXlsx(132, 10, "Y")
    #新增业务枚举科目配置_0132
    def test_s_Subject_config(self):
        self.System_config()
        self.Frame1_to()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 业务枚举多级-业务枚举s
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举AA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[6]").text
        print(Lab1)  # 1192
        if Msg1 == "保存成功!":
            self.WriteXlsx(133, 11, "PASS")
        else:
            self.WriteXlsx(133, 11, "ERROR")
        self.WriteXlsx(133, 10, "Y")
    #新增业务枚举科目配置_0133
    def test_t_Subject_config(self):
        self.System_config()
        self.Frame1_to()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]" ).click()  #
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[1]/div/input").click()  # 三级枚举AAA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        Frame4 = "layui-layer-iframe4"
        Frame5 = "layui-layer-iframe5"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 办公室
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()  # 办公室
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        self.driver.find_element_by_xpath("//span[@id='myTree_35_check']").click()  # 水政水资源科
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_38_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_39_check']").click()  # 办公厅
        self.driver.find_element_by_xpath("//span[@id='myTree_40_check']").click()  # 综合司
        self.driver.find_element_by_xpath("//span[@id='myTree_41_check']").click()  # 条法司
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame4)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[1]/div/input").click()  # 1193[三级枚举科目]
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame5)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 2001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer5']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[6]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print(Lab1, Lab2)  # 1193  办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Msg1 == "保存成功!":
            self.WriteXlsx(134, 11, "PASS")
        else:
            self.WriteXlsx(134, 11, "ERROR")
        self.WriteXlsx(134, 10, "Y")
    #新增业务枚举科目配置_0134
    def test_u_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()#
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()#部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()#清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()#展开

        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()#选中监察室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[6]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[12]").text
        print(Lab1, Lab2)  # 1193  监察室
        if Msg1 == "保存成功!":
            self.WriteXlsx(135, 11, "PASS")
        else:
            self.WriteXlsx(135, 11, "ERROR")
        self.WriteXlsx(135, 10, "Y")
    def FrAME_MJ(self):
        '''年度选择2019，映射类型选择请假类型-公共枚举,账套选择001，新增进入科目配置窗口'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='请假类型-公共枚举']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
    #新增公共枚举科目配置_0135
    def test_v_Subject_config(self):
        self.System_config()
        self.FrAME_MJ()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#请假类型-公共枚举:
        sleep(1)
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()#年休假
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td[1]/div/input").click()  # 3001清算资金往来
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[7]").text
        print(Lab1, Lab2)  # 年休假  3001[清算资金往来]
        if Msg1 == "保存成功!" and Lab1 == "年休假" and Lab2 == "3001[清算资金往来]":
            self.WriteXlsx(136, 11, "PASS")
        else:
            self.WriteXlsx(136, 11, "ERROR")
        self.WriteXlsx(136, 10, "Y")
    #新增公共枚举科目配置_0136
    def test_w_Subject_config(self):
        self.System_config()
        self.FrAME_MJ()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 请假类型-公共枚举:
        sleep(1)
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[1]/div/input").click()  # 婚假
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[2]/td[1]/div/input").click()  # 3002货币兑换
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[7]").text
        print(Lab1, Lab2)  # 婚假  3002[货币兑换]
        if Msg1 == "保存成功!" and Lab1 == "婚假" and Lab2 == "3002[货币兑换]":
            self.WriteXlsx(137, 11, "PASS")
        else:
            self.WriteXlsx(137, 11, "ERROR")
        self.WriteXlsx(137, 10, "Y")
    #新增公共枚举科目配置_0137
    def test_x_Subject_config(self):
        self.System_config()
        self.FrAME_MJ()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 请假类型-公共枚举:
        sleep(1)
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        frame4 = "layui-layer-iframe4"
        frame5 = "layui-layer-iframe5"
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[7]/td[1]/div/input").click()  # 其他
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 办公室
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()  # 办公室
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        self.driver.find_element_by_xpath("//span[@id='myTree_35_check']").click()  # 水政水资源科
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_38_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_39_check']").click()  # 办公厅
        self.driver.find_element_by_xpath("//span[@id='myTree_40_check']").click()  # 综合司
        self.driver.find_element_by_xpath("//span[@id='myTree_41_check']").click()  # 条法司
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame4)
        self.driver.find_element_by_xpath("//span[@id='mytree_138_span']").click()  # 共同
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[3]/td[1]/div/input").click()  # 3007科目有权限
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame5)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 2001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer5']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print(Lab1, Lab2)  # 其他  办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Msg1 == "保存成功!":
            self.WriteXlsx(138, 11, "PASS")
        else:
            self.WriteXlsx(138, 11, "ERROR")
        self.WriteXlsx(138, 10, "Y")
    #新增公共枚举科目配置_0138
    def test_y_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='请假类型-公共枚举']").click()  # 业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()#选中其它
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()#展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()  # 选中监察室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[12]").text
        print(Lab1, Lab2)  # 其他  监查室
        if Msg1 == "保存成功!":
            self.WriteXlsx(139, 11, "PASS")
        else:
            self.WriteXlsx(139, 11, "ERROR")
        self.WriteXlsx(139, 10, "Y")
    def free_Frame1(self):
        '''年度选择2019，映射类型选择映射文本-自由文本(单位A),账套选择001，新增进入科目配置窗口'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  # 映射文本-自由文本(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
    #新增自由文本科目配置_0139
    def test_z_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        self.Replace(140,8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)栏输入"])
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[8]/td[1]/div/input").click()#科目A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[7]").text
        print(Lab1, Lab2)  # 收  2291[科目A]
        if Msg1 == "保存成功!" and Lab1 == "收" and Lab2 == "2101[交易性金融负债]":
            self.WriteXlsx(140, 11, "PASS")
        else:
            self.WriteXlsx(140, 11, "ERROR")
        self.WriteXlsx(140, 10, "Y")
    #新增自由文本科目配置_0140
    def test_zb_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        self.Replace(141, 8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)栏输入"])
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[9]/td[1]/div/input").click()  # 科目B
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[7]").text
        print(Lab1, Lab2)  # 收  2292[科目B]
        if Msg1 == "保存成功!" and Lab1 == "收付" and Lab2 == "2111[卖出回购金融资产款]":
            self.WriteXlsx(141, 11, "PASS")
        else:
            self.WriteXlsx(141, 11, "ERROR")
        self.WriteXlsx(141, 10, "Y")
    #新增自由文本科目配置_0141
    def test_zc_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        self.Replace(142, 8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)栏输入"])
        sleep(1)
        frame1 = 'layui-layer-iframe1'
        frame2 = 'layui-layer-iframe2'
        frame3 = 'layui-layer-iframe3'
        frame4 = 'layui-layer-iframe4'
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        #
        # self.driver.find_element_by_xpath("//span[@id='myTree_2_check']").click()  # 办公室
        # self.driver.find_element_by_xpath("//span[@id='myTree_4_check']").click()  # 综合科
        # self.driver.find_element_by_xpath("//span[@id='myTree_7_check']").click()  # 水政水资源科
        # self.driver.find_element_by_xpath("//span[@id='myTree_9_switch']").click()  # 展开
        # sleep(1)
        # self.driver.find_element_by_xpath("//span[@id='myTree_10_switch']").click()  # 展开
        # sleep(1)
        # self.driver.find_element_by_xpath("//span[@id='myTree_11_check']").click()  # 办公厅
        # self.driver.find_element_by_xpath("//span[@id='myTree_12_check']").click()  # 综合司
        # self.driver.find_element_by_xpath("//span[@id='myTree_13_check']").click()  # 条法司
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()  # 办公室
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        self.driver.find_element_by_xpath("//span[@id='myTree_35_check']").click()  # 水政水资源科
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_38_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_39_check']").click()  # 办公厅
        self.driver.find_element_by_xpath("//span[@id='myTree_40_check']").click()  # 综合司
        self.driver.find_element_by_xpath("//span[@id='myTree_41_check']").click()  # 条法司
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[10]/td[1]/div/input").click()  # 科目C
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame4)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 2001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print(Lab1, Lab2)  # 收付转  办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Msg1 == "保存成功!":
            self.WriteXlsx(142, 11, "PASS")
        else:
            self.WriteXlsx(142, 11, "ERROR")
        self.WriteXlsx(142, 10, "Y")
    #新增自由文本科目配置_0142
    def test_zd_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  # 映射文本-自由文本(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()#收付转
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()  # 选中监察室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print(Lab1, Lab2)  # 收付转  监查室
        if Msg1 == "保存成功!":
            self.WriteXlsx(143, 11, "PASS")
        else:
            self.WriteXlsx(143, 11, "ERROR")
        self.WriteXlsx(143, 10, "Y")
    #删除新增的单位枚举科目配置_0143
    def test_ze_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()#多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        self.driver.switch_to.default_content()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        # self.driver.switch_to.default_content()
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(144, 11, "PASS")
        else:
            self.WriteXlsx(144, 11, "ERROR")
        self.WriteXlsx(144, 10, "Y")
    #删除新增的业务枚举科目配置_0144
    def test_zf_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()#业务枚举多级-业务枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        self.driver.switch_to.default_content()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        # self.driver.switch_to.default_content()
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(145, 11, "PASS")
        else:
            self.WriteXlsx(145, 11, "ERROR")
        self.WriteXlsx(145, 10, "Y")
    #删除新增的公共枚举科目配置_0145
    def test_zh_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='请假类型-公共枚举']").click()  # 请假类型-公共枚举
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        self.driver.switch_to.default_content()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        # self.driver.switch_to.default_content()
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(146, 11, "PASS")
        else:
            self.WriteXlsx(146, 11, "ERROR")
        self.WriteXlsx(146, 10, "Y")
    #删除新增的自由文本科目配置_0146
    def test_zi_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  # 映射文本-自由文本(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        self.driver.switch_to.default_content()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        # self.driver.switch_to.default_content()
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(147, 11, "PASS")
        else:
            self.WriteXlsx(147, 11, "ERROR")
        self.WriteXlsx(147, 10, "Y")
    def N9_N6(self):
        '''年度选择2019，映射类型选择N6/N9支出事项-系统档案(单位A),账套选择001，新增进入科目配置窗口'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()#N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)

    #新增N9支出事项科目映射_0147
    def test_zj_Subject_config(self):
        self.System_config()
        self.N9_N6()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_3_span']").click()#支出事项A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[8]/td[1]/div/input").click()#2291科目A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  #保存成功!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text#支出事项A
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[7]").text  # 2291[科目A]
        print("Lab1,Lab2:",Lab1,Lab2)
        if Msg1 == "保存成功!":
            self.WriteXlsx(148, 11, "PASS")
        else:
            self.WriteXlsx(148, 11, "ERROR")
        self.WriteXlsx(148, 10, "Y")
    #新增N9支出事项科目映射_0148
    def test_zk_Subject_config(self):
        self.System_config()
        self.N9_N6()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_4_span']").click()  # 支出事项B
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[9]/td[1]/div/input").click()  # 2292科目B
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text # 支出事项B
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[7]").text # 2292[科目B]
        print("Lab1,Lab2:", Lab1, Lab2)
        if Msg1 == "保存成功!":
            self.WriteXlsx(149, 11, "PASS")
        else:
            self.WriteXlsx(149, 11, "ERROR")
        self.WriteXlsx(149, 10, "Y")
    #新增N9支出事项科目映射_0149
    def test_zl_Subject_config(self):
        self.System_config()
        self.N9_N6()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        frame4 = "layui-layer-iframe4"
        frame5 = "layui-layer-iframe5"
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_5_span']").click()  # 支出事项C
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()  # 办公室
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        self.driver.find_element_by_xpath("//span[@id='myTree_35_check']").click()  # 水政水资源科
        self.driver.find_element_by_xpath("//span[@id='myTree_54_check']").click()  # 文化和旅游部
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame4)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[10]/td[1]/div/input").click()  # 2293科目C
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame5)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 短期借款
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer5']/div[3]/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text  # 支出事项B
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text  #
        print("Lab1,Lab2:", Lab1, Lab2)#办公室,财务室,综合科,水政水资源科,文化和旅游部,国家文物局
        if Msg1 == "保存成功!":
            self.WriteXlsx(150, 11, "PASS")
        else:
            self.WriteXlsx(150, 11, "ERROR")
        self.WriteXlsx(150, 10, "Y")
    #新增N9支出事项科目映射_0150
    def test_zm_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()#支出事项C
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()#部门
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame(frame2)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()#清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()#监查室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text  # 支出事项B
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[12]").text  #
        print("Lab1,Lab2:", Lab1, Lab2)  #监查室
        if Msg1 == "保存成功!" and Lab1 == "支出事项C" and Lab2 == '监查室':
            self.WriteXlsx(151, 11, "PASS")
        else:
            self.WriteXlsx(151, 11, "ERROR")
        self.WriteXlsx(151, 10, "Y")
    #删除新增的N9支出事项科目配置_0151
    def test_zn_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        self.driver.switch_to.default_content()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(152, 11, "PASS")
        else:
            self.WriteXlsx(152, 11, "ERROR")
        self.WriteXlsx(152, 10, "Y")
    #新增枚举类型科目配置_0152
    def test_zo_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一行第一个 一级枚举A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()  # 一级枚举科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[6]").text
        print(Lab1)  # 1191
        if Msg1 == "保存成功!":
            self.WriteXlsx(153, 11, "PASS")
        else:
            self.WriteXlsx(153, 11, "ERROR")
        self.WriteXlsx(153, 10, "Y")
    # 新增枚举类型科目配置_0153
    def test_zp_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举AA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[6]").text
        print(Lab1)  # 1192
        if Msg1 == "保存成功!":
            self.WriteXlsx(154, 11, "PASS")
        else:
            self.WriteXlsx(154, 11, "ERROR")
        self.WriteXlsx(154, 10, "Y")
    # 新增枚举类型科目配置_0154
    def test_zq_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[1]/div/input").click()  # 三级枚举AAA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        Frame4 = "layui-layer-iframe4"
        Frame5 = "layui-layer-iframe5"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()  # 办公室
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        self.driver.find_element_by_xpath("//span[@id='myTree_35_check']").click()  # 水政水资源科
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_38_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_39_check']").click()  # 办公厅
        self.driver.find_element_by_xpath("//span[@id='myTree_40_check']").click()  # 综合司
        self.driver.find_element_by_xpath("//span[@id='myTree_41_check']").click()  # 条法司
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame4)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[1]/div/input").click()  # 1193[三级枚举科目]
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame5)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 2001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer5']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[6]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print(Lab1, Lab2)  # 1193  办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Msg1 == "保存成功!":
            self.WriteXlsx(155, 11, "PASS")
        else:
            self.WriteXlsx(155, 11, "ERROR")
        self.WriteXlsx(155, 10, "Y")
    #新增枚举类型科目配置_0155
    def test_zr_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()#三级枚举AAA
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增
        sleep(1)
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame(frame2)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()  # 监查室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text  # 三级枚举AAA
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[12]").text  #
        print("Lab1,Lab2:", Lab1, Lab2)  # 监查室
        if Msg1 == "保存成功!" and Lab1 == "三级枚举AAA" and Lab2 == '监查室' or Lab2=="办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司":
            self.WriteXlsx(156, 11, "PASS")
        else:
            self.WriteXlsx(156, 11, "ERROR")
        self.WriteXlsx(156, 10, "Y")
    #不选择列表数据情况下修改_0156
    def test_zs_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()#修改
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)#请选择一条需要修改的数据!
        if Msg=='请选择一条需要修改的数据!':
            self.WriteXlsx(157, 11, "PASS")
        else:
            self.WriteXlsx(157, 11, "ERROR")
        self.WriteXlsx(157, 10, "Y")
    #多选列表数据情况下修改_0157
    def test_zt_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#第一行第一个
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()#第一行第二个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请选择一条需要修改的数据!
        if Msg == '请选择一条需要修改的数据!':
            self.WriteXlsx(158, 11, "PASS")
        else:
            self.WriteXlsx(158, 11, "ERROR")
        self.WriteXlsx(158, 10, "Y")
    #修改页面显示_0158
    def test_zu_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").get_attribute('value')
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").get_attribute('value')
        print("Lab1,Lab2:",Lab1,Lab2)#一级枚举A  1191[一级枚举科目]
        if Lab1=="一级枚举A":
            self.WriteXlsx(159, 11, "PASS")
        else:
            self.WriteXlsx(159, 11, "ERROR")
        self.WriteXlsx(159, 10, "Y")
    #修改页面显示_0159
    def test_zv_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 第3个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").get_attribute('value')
        Lab2 = self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").get_attribute('value')
        Lab3 = self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").get_attribute('value')
        Lab4 = self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").get_attribute('value')
        print("Lab1~Lab4:",Lab1,Lab2,Lab3,Lab4)#三级枚举AAA 办公室,财务室,综合科,水政水资源科,办公厅,
        # 综合司,条法司 1193[三级枚举科目]  2001[短期借款]
        if Lab1=="二级枚举AAA" and Lab2=="办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司":
            # if Lab3=="1193[三级枚举科目]" and Lab4=="2001[短期借款]":
            self.WriteXlsx(160, 11, "PASS")
        else:
            self.WriteXlsx(160, 11, "ERROR")
        # else:
        #     self.WriteXlsx(160, 11, "ERROR")
        self.WriteXlsx(160, 10, "Y")
    #修改账套_0161
    def test_zw_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第1个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").click()#点击账套
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe2")
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 第3个
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").get_attribute('value')#
        print(Lab1)#平行记帐测试账套002[002, 2019]
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[3]").click()  # 平行记帐测试账套002[002]
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]/div").text  # 第1个
        print(Lab2)
        if Lab1== "平行记帐测试账套002[002, 2019]" and Lab2=='平行记帐测试账套002':
            self.WriteXlsx(162, 11, "PASS")
        else:
            self.WriteXlsx(162, 11, "ERROR")
        self.WriteXlsx(162, 10, "Y")
    #修改账套_0162
    def test_zx_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 第4个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").click()  # 点击账套
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe2")
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()  # 第2个
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("layui-layer-iframe1")
        Lab1 = self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").get_attribute('value')  #
        print(Lab1)  # 测试帐套001[001, 2020]
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        # self.driver.find_element_by_xpath("//select[@id='account1']/option[3]").click()  # 平行记帐测试账套002[002]
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[3]/div").text  # 第1个
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]/div").text
        print(Lab2,Lab3)
        if Lab1 == "测试帐套001[001, 2020]":
            self.WriteXlsx(163, 11, "PASS")
        else:
            self.WriteXlsx(163, 11, "ERROR")
        self.WriteXlsx(163, 10, "Y")
    #修改账套，清空_0163
    def test_zy_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第1个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR1']/td/input[5]").click()#测试帐套001[001, 2020]
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()#清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 保存
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)#请选择账套信息!
        if Msg == "请选择账套信息!":
            self.WriteXlsx(164, 11, "PASS")
        else:
            self.WriteXlsx(164, 11, "ERROR")
        self.WriteXlsx(164, 10, "Y")
    #修改枚举_0164
    def test_zz_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第1个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()#多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()#第2个
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)#保存成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()# 确定
        self.driver.switch_to.frame(frame)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个  一级枚举B
        print(Lab1)#一级枚举B
        if Msg=="保存成功!" and Lab1=="一级枚举B":
            self.WriteXlsx(165, 11, "PASS")
        else:
            self.WriteXlsx(165, 11, "ERROR")
        self.WriteXlsx(165, 10, "Y")
    #修改枚举_0165
    def test_zza_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第1个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  #
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='mytree_1_span']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[1]/div/input").click()  # 第4个
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
        self.driver.switch_to.frame(frame)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个  一级枚举B
        print(Lab1)  # 二级枚举AB
        if Msg == "保存成功!":
            self.WriteXlsx(166, 11, "PASS")
        else:
            self.WriteXlsx(166, 11, "ERROR")
        self.WriteXlsx(166, 10, "Y")
    def AB_2019(self):
        '''年度选择2019，映射类型选择多级枚举-单位枚举(单位A),账套选择全部，列表选择对象名为二级枚举AB的数据记录，点击修改进入科目配置窗口'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第1个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)

    #修改枚举，清空_0166
    def test_zzb_Subject_config(self):
        self.System_config()
        self.AB_2019()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请选填多级枚举-单位枚举(单位A)!
        if Msg=="请选填多级枚举-单位枚举(单位A)!":
            self.WriteXlsx(167, 11, "PASS")
        else:
            self.WriteXlsx(167, 11, "ERROR")
        self.WriteXlsx(167, 10, "Y")
    #修改部门_0167
    def test_zzc_Subject_config(self):
        self.System_config()
        self.AB_2019()
        frame  = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()#部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()#综合科
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  #保存成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()#保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[12]").text  # 第1个
        print(Lab)#综合科
        if Msg=="保存成功!" and Lab=="综合科":
            self.WriteXlsx(168, 11, "PASS")
        else:
            self.WriteXlsx(168, 11, "ERROR")
        self.WriteXlsx(168, 10, "Y")
    #修改部门_0168
    def test_zzd_Subject_config(self):
        self.System_config()
        self.AB_2019()
        frame = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()  # 监查室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[12]").text  # 第1个
        print(Lab)  # 综合科 监查室
        if Msg == "保存成功!":
            self.WriteXlsx(169, 11, "PASS")
        else:
            self.WriteXlsx(169, 11, "ERROR")
        self.WriteXlsx(169, 10, "Y")
    #修改部门，清空_0169
    def test_zze_Subject_config(self):
        self.System_config()
        self.AB_2019()
        frame = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 确定
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[12]").text  # 第1个
        print("Lab:",Lab)  # 空
        if Msg == "保存成功!" and Lab == " ":
            self.WriteXlsx(170, 11, "PASS")
        else:
            self.WriteXlsx(170, 11, "ERROR")
        self.WriteXlsx(170, 10, "Y")
    #修改财务科目_0170
    def test_zzf_Subject_config(self):
        self.System_config()
        self.AB_2019()
        frame = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_146_span']").click()#权益
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#4001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[7]").text  # 第3个
        print("Lab:", Lab)  #4001[实收资本]
        if Msg == "保存成功!":
            self.WriteXlsx(171, 11, "PASS")
        else:
            self.WriteXlsx(171, 11, "ERROR")
        self.WriteXlsx(171, 10, "Y")
    #修改财务科目_0171
    def test_zzg_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 第3个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        frame = "myiframe"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_162_span']").click()  # 损益
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 6001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[7]").text  # 第3个
        print("Lab:", Lab)  # 6001[主营业务收入]
        if Msg == "保存成功!":
            self.WriteXlsx(172, 11, "PASS")
        else:
            self.WriteXlsx(172, 11, "ERROR")
        self.WriteXlsx(172, 10, "Y")
    #修改财务科目，清空_0172
    def test_zzh_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 第3个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        frame = "myiframe"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()#清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 财务科目信息和预算科目信息不能同时为空!请选择...
        if Msg == "财务科目信息和预算科目信息不能同时为空!请选择...":
            self.WriteXlsx(173, 11, "PASS")
        else:
            self.WriteXlsx(173, 11, "ERROR")
        self.WriteXlsx(173, 10, "Y")
    #修改税金科目_0173
    def test_zzi_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第1个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame = "myiframe"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()#税金
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[1]/div/input").click()#22210101进项税额
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[9]").text  # 第1个
        print("Lab:", Lab)  # 22210101[应交税费-应交增值税-进项税额]
        if Msg == "保存成功!":
            self.WriteXlsx(174, 11, "PASS")
        else:
            self.WriteXlsx(174, 11, "ERROR")
        self.WriteXlsx(174, 10, "Y")
    #修改税金科目_0174
    def test_zzj_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()  # 第2个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame = "myiframe"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 22210101进项税额
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[9]").text  # 第1个
        print("Lab:", Lab)  # 2001[短期借款]
        if Msg == "保存成功!":
            self.WriteXlsx(175, 11, "PASS")
        else:
            self.WriteXlsx(175, 11, "ERROR")
        self.WriteXlsx(175, 10, "Y")
    #修改税金科目，清空_0175
    def test_zzk_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()  # 第2个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame = "myiframe"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[9]").text  # 第1个
        print("Lab:", Lab)  # 空
        if Msg == "保存成功!":
            self.WriteXlsx(176, 11, "PASS")
        else:
            self.WriteXlsx(176, 11, "ERROR")
        self.WriteXlsx(176, 10, "Y")
    #新增自由文本科目配置_0176
    def test_zzl_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        self.Replace(177,8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)"])
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目
        sleep(1)
        frame  = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[8]/td[1]/div/input").click()  # 2291
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[7]").text  # 第1个
        print("Lab:", Lab)  # 2291[科目A]
        if Msg == "保存成功!":
            self.WriteXlsx(177, 11, "PASS")
        else:
            self.WriteXlsx(177, 11, "ERROR")
        self.WriteXlsx(177, 10, "Y")
    #修改自由文本映射信息_0177
    def test_zzm_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  # 映射文本-自由文本(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").clear()
        sleep(1)
        self.Replace(178, 8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)"])
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个
        print("Lab:", Lab)  # 收付
        if Msg == "保存成功!" and Lab == "收付":
            self.WriteXlsx(178, 11, "PASS")
        else:
            self.WriteXlsx(178, 11, "ERROR")
        self.WriteXlsx(178, 10, "Y")
    #修改自由文本映射信息_0178
    def test_zzn_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  # 映射文本-自由文本(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").clear()
        sleep(1)
        self.Replace(179, 8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)"])
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个
        print("Lab:", Lab)  # 收付
        if Msg == "保存成功!" and Lab == "收付转":
            self.WriteXlsx(179, 11, "PASS")
        else:
            self.WriteXlsx(179, 11, "ERROR")
        self.WriteXlsx(179, 10, "Y")
    #修改自由文本映射信息，请空_0179
    def test_zzo_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  # 映射文本-自由文本(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").clear()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请选填映射文本-自由文本(单位A)!！
        if Msg == "请选填映射文本-自由文本(单位A)!" :
            self.WriteXlsx(180, 11, "PASS")
        else:
            self.WriteXlsx(180, 11, "ERROR")
        self.WriteXlsx(180, 10, "Y")
    #新增自由文本映射信息时，自由文本必填校验_0180
    def test_zzp_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请选填映射文本-自由文本(单位A)!！
        if Msg == "请选填映射文本-自由文本(单位A)!":
            self.WriteXlsx(181, 11, "PASS")
        else:
            self.WriteXlsx(181, 11, "ERROR")
        self.WriteXlsx(181, 10, "Y")
    #新增自由文本映射信息时，自由文本特殊字符验证_0181
    def test_zzq_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        self.Replace(182, 8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)"])
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[8]/td[1]/div/input").click()#2291
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个
        print("Lab:", Lab)  # !@#$%^&*()_+`-=
        if Msg == "保存成功!" and Lab == "!@#$%^&*()_+`-=":
            self.WriteXlsx(182, 11, "PASS")
        else:
            self.WriteXlsx(182, 11, "ERROR")
        self.WriteXlsx(182, 10, "Y")
    #新增自由文本映射信息时，自由文本特殊字符验证_0182
    def test_zzr_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        self.Replace(183, 8)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)"])
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[8]/td[1]/div/input").click()  # 2291
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text  # 第3个
        print("Lab:", Lab)  # +-;,.18|:<>?
        if Msg == "保存成功!":
            self.WriteXlsx(183, 11, "PASS")
        else:
            self.WriteXlsx(183, 11, "ERROR")
        self.WriteXlsx(183, 10, "Y")
    #新增N9支出项目科目配置_0183
    def test_zzs_Subject_config(self):
        self.System_config()
        self.N9_N6()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_3_span']").click()  # 支出事项A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[8]/td[1]/div/input").click()  # 2291
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个
        print("Lab:", Lab)  # 支出事项A
        if Msg == "保存成功!" and Lab == "支出事项A":
            self.WriteXlsx(184, 11, "PASS")
        else:
            self.WriteXlsx(184, 11, "ERROR")
        self.WriteXlsx(184, 10, "Y")
    #修改N9支出项目映射信息_0184
    def test_zzt_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_4_span']").click()  # 支出事项B
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个
        print("Lab:", Lab)  # 支出事项B
        if Msg == "保存成功!" and Lab == "支出事项B":
            self.WriteXlsx(185, 11, "PASS")
        else:
            self.WriteXlsx(185, 11, "ERROR")
        self.WriteXlsx(185, 10, "Y")
    #修改N9支出项目映射信息_0185
    def test_zzu_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_5_span']").click()  # 支出事项C
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功！
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 保存
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 第1个
        print("Lab:", Lab)  # 支出事项C
        if Msg == "保存成功!" and Lab == "支出事项C":
            self.WriteXlsx(186, 11, "PASS")
        else:
            self.WriteXlsx(186, 11, "ERROR")
        self.WriteXlsx(186, 10, "Y")
    #修改N9支出项目映射信息，请空_0186
    def test_zzv_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请选填N6/N9支出事项-系统档案(单位A)!
        if Msg == "请选填N6/N9支出事项-系统档案(单位A)!" :
            self.WriteXlsx(187, 11, "PASS")
        else:
            self.WriteXlsx(187, 11, "ERROR")
        self.WriteXlsx(187, 10, "Y")
    #新增N9支出项目映射信息时，N9支出项目必填校验_0187
    def test_zzw_Subject_config(self):
        self.System_config()
        self.N9_N6()
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  #请选填N6/N9支出事项-系统档案(单位A)!
        if Msg=="请选填N6/N9支出事项-系统档案(单位A)!":
            self.WriteXlsx(188, 11, "PASS")
        else:
            self.WriteXlsx(188, 11, "ERROR")
        self.WriteXlsx(188, 10, "Y")
    #删除科目配置，单行删除提示_0188
    def test_zzx_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()#删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            self.WriteXlsx(189, 11, "PASS")
        else:
            self.WriteXlsx(189, 11, "ERROR")
        self.WriteXlsx(189, 10, "Y")
    #删除科目配置，多行删除提示_0189
    def test_zzy_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            self.WriteXlsx(190, 11, "PASS")
        else:
            self.WriteXlsx(190, 11, "ERROR")
        self.WriteXlsx(190, 10, "Y")
    #删除科目配置，取消删除_0190
    def test_zzz_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[2]").click()#取消
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            self.WriteXlsx(191, 11, "PASS")
        else:
            self.WriteXlsx(191, 11, "ERROR")
        self.WriteXlsx(191, 10, "Y")
    #年度选择2019，映射类型选择N6/N9支出事项-系统档案(单位A)，账套选择001
    def test_zzza_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # N6/N9支出事项-系统档案(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 第一个
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  #确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2=="删除成功!":
            self.WriteXlsx(192, 11, "PASS")
        else:
            self.WriteXlsx(192, 11, "ERROR")
        self.WriteXlsx(192, 10, "Y")
    #删除科目配置，多行删除_0192
    def test_zzzb_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2=="删除成功!":
            self.WriteXlsx(193, 11, "PASS")
        else:
            self.WriteXlsx(193, 11, "ERROR")
        self.WriteXlsx(193, 10, "Y")
    #删除用例数据_0194
    def test_zzzc_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()#映射文本-自由文本(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr[1]/th/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        Msg2 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg2)  # 删除成功!
        if Msg1 == "数据删除后将无法恢复,请确认是否进行删除!" and Msg2 == "删除成功!":
            self.WriteXlsx(195, 11, "PASS")
        else:
            self.WriteXlsx(195, 11, "ERROR")
        self.WriteXlsx(195, 10, "Y")
    #增加科目配置做为测试数据_0195
    def test_zzzd_Subject_config(self):
        self.System_config()
        self.DeM_A()
        frame  = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#多级枚举-单位枚举
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#第一个
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()  # 一级枚举科目1191
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()#保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()
        self.driver.switch_to.frame(frame)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[7]/div").text
        print(Lab1)#1191[一级枚举科目]
        if Msg=="保存成功!":
            self.WriteXlsx(196, 11, "PASS")
        else:
            self.WriteXlsx(196, 11, "ERROR")
        self.WriteXlsx(196, 10, "Y")
    #增加科目配置做为测试数据_0196
    def test_zzze_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举AA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 二级枚举科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[6]").text
        print(Lab1)  # 1192
        if Msg1 == "保存成功!":
            self.WriteXlsx(197, 11, "PASS")
        else:
            self.WriteXlsx(197, 11, "ERROR")
        self.WriteXlsx(197, 10, "Y")
    #增加科目配置做为测试数据_0197
    def test_zzzf_Subject_config(self):
        self.System_config()
        self.DeM_A()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  # 多级枚举-单位枚举(单位A):
        sleep(1)
        self.driver.switch_to.default_content()
        Frame2 = "layui-layer-iframe2"
        self.driver.switch_to.frame(Frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[5]/td[1]/div/input").click()  # 三级枚举AAA
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Frame1 = "layui-layer-iframe1"
        Frame3 = "layui-layer-iframe3"
        Frame4 = "layui-layer-iframe4"
        Frame5 = "layui-layer-iframe5"
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()  # 展开
        self.driver.find_element_by_xpath("//span[@id='myTree_30_check']").click()  # 办公室
        self.driver.find_element_by_xpath("//span[@id='myTree_32_check']").click()  # 综合科
        self.driver.find_element_by_xpath("//span[@id='myTree_35_check']").click()  # 水政水资源科
        self.driver.find_element_by_xpath("//span[@id='myTree_37_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_38_switch']").click()  # 展开
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_39_check']").click()  # 办公厅
        self.driver.find_element_by_xpath("//span[@id='myTree_40_check']").click()  # 综合司
        self.driver.find_element_by_xpath("//span[@id='myTree_41_check']").click()  # 条法司
        sleep(1)
        # self.driver.find_element_by_xpath("//span[@id='myTree_2_check']").click()  # 办公室
        # self.driver.find_element_by_xpath("//span[@id='myTree_4_check']").click()  # 综合科
        # self.driver.find_element_by_xpath("//span[@id='myTree_7_check']").click()  # 水政水资源科
        # self.driver.find_element_by_xpath("//span[@id='myTree_9_switch']").click()  # 展开
        # sleep(1)
        # self.driver.find_element_by_xpath("//span[@id='myTree_10_switch']").click()  # 展开
        # sleep(1)
        # self.driver.find_element_by_xpath("//span[@id='myTree_11_check']").click()  # 办公厅
        # self.driver.find_element_by_xpath("//span[@id='myTree_12_check']").click()  # 综合司
        # self.driver.find_element_by_xpath("//span[@id='myTree_13_check']").click()  # 条法司
        # sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame4)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_ico']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[1]/div/input").click()  # 1193[三级枚举科目]
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.find_element_by_xpath("//tr[@id='TaxTR']/td/input[2]").click()  # 税金科目:
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(Frame5)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 2001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer5']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(Frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()  # 确定
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[6]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print(Lab1, Lab2)  # 1193  办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Msg1 == "保存成功!":
            self.WriteXlsx(198, 11, "PASS")
        else:
            self.WriteXlsx(198, 11, "ERROR")
        self.WriteXlsx(198, 10, "Y")
    #增加科目配置做为测试数据_0198
    def test_zzzg_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[1]/div/input").click()  # 三级枚举AAA
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增
        sleep(1)
        self.driver.switch_to.default_content()
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame(frame2)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[2]").click()  # 清空
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='DeptTR']/td/input[2]").click()  # 部门
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        self.driver.switch_to.frame(frame3)
        self.driver.find_element_by_xpath("//span[@id='myTree_29_switch']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_33_check']").click()  # 监查室
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 保存
        sleep(1)
        self.driver.switch_to.default_content()
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text  #
        print(Msg1)  # 保存成功!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text  # 三级枚举AAA
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[12]").text  #
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text
        print("Lab1,Lab2:", Lab1, Lab2)  # 监查室
        if Msg1 == "保存成功!":
            self.WriteXlsx(199, 11, "PASS")
        else:
            self.WriteXlsx(199, 11, "ERROR")
        self.WriteXlsx(199, 10, "Y")
    def All_DJ(self):
        '''年度选择2019，映射类型选择多级枚举-单位枚举(单位A)，账套选择全部'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)

    #科目配置查询，对象名称查询_0205
    def test_zzzh_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()#--查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[2]").click()#对象名称
        sleep(1)
        self.Replace(206,8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["对象名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()#查询
        sleep(1)
        Lab = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        print(Lab)#一级枚举A
        if  Lab =="一级枚举A":
            self.WriteXlsx(206, 11, "PASS")
        else:
            self.WriteXlsx(206, 11, "ERROR")
        self.WriteXlsx(206, 10, "Y")
    #科目配置查询，对象名称查询_0206
    def test_zzzi_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[2]").click()  # 对象名称
        sleep(1)
        self.Replace(207, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["对象名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        print(Lab1,Lab2)  # 三级枚举AAA
        if Lab1 ==Lab2== "三级枚举AAA":
            self.WriteXlsx(207, 11, "PASS")
        else:
            self.WriteXlsx(207, 11, "ERROR")
        self.WriteXlsx(207, 10, "Y")
    #科目配置查询，对象名称查询_0207
    def test_zzzj_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[2]").click()  # 对象名称
        sleep(1)
        self.Replace(208, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["对象名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:",Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(208, 11, "PASS")
        else:
            self.WriteXlsx(208, 11, "ERROR")
        self.WriteXlsx(208, 10, "Y")
    #科目配置查询，对象名称查询_0208
    def test_zzzk_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[2]").click()  # 对象名称
        sleep(1)
        self.Replace(209, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["对象名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        print(Lab1, Lab2)  # 三级枚举AAA
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        Lab5 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text
        Lab6 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text
        print("Lab3~Lab4:",Lab3,Lab4)#一级枚举A   二级枚举AA
        if Lab1 == Lab2 ==Lab5==Lab6== "三级枚举AAA":
            if Lab3=="一级枚举A" and Lab4=="二级枚举AA":
                self.WriteXlsx(209, 11, "PASS")
            else:
                self.WriteXlsx(209, 11, "ERROR")
        else:
            self.WriteXlsx(209, 11, "ERROR")
        self.WriteXlsx(209, 10, "Y")
    #科目配置查询，对象名称查询，特殊符号查询_0209
    def test_zzzl_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[2]").click()  # 对象名称
        sleep(1)
        self.Replace(210, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["对象名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:", Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(210, 11, "PASS")
        else:
            self.WriteXlsx(210, 11, "ERROR")
        self.WriteXlsx(210, 10, "Y")
    #科目配置查询，财务科目名称查询_0210
    def test_zzzm_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[3]").click()  # 财务科目名称
        sleep(1)
        self.Replace(211, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["财务科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        print("Lab1:", Lab1)  #一级枚举A
        if Lab1 == "一级枚举A":
            self.WriteXlsx(211, 11, "PASS")
        else:
            self.WriteXlsx(211, 11, "ERROR")
        self.WriteXlsx(211, 10, "Y")
    #科目配置查询，财务科目名称查询_0211
    def test_zzzn_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[3]").click()  # 财务科目名称
        sleep(1)
        self.Replace(212, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["财务科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        print("Lab1,Lab2:", Lab1,Lab2)  # 三级枚举AAA
        if Lab1 == Lab2=="一级枚举AAA":
            self.WriteXlsx(212, 11, "PASS")
        else:
            self.WriteXlsx(212, 11, "ERROR")
        self.WriteXlsx(212, 10, "Y")
    #科目配置查询，财务科目名称查询_0212
    def test_zzzo_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[3]").click()  # 财务科目名称
        sleep(1)
        self.Replace(213, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["财务科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:", Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(213, 11, "PASS")
        else:
            self.WriteXlsx(213, 11, "ERROR")
        self.WriteXlsx(213, 10, "Y")
    #科目配置查询，财务科目名称查询_0213
    def test_zzzp_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[3]").click()  # 对象名称
        sleep(1)
        self.Replace(214, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["财务科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        print(Lab1, Lab2)  # 三级枚举AAA
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        Lab5 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text
        Lab6 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text
        print("Lab3~Lab4:", Lab3, Lab4)  # 一级枚举A   二级枚举AA
        if Lab1 == Lab2 == Lab5 == Lab6 == "三级枚举AAA":
            if Lab3 == "一级枚举A" and Lab4 == "二级枚举AA":
                self.WriteXlsx(214, 11, "PASS")
            else:
                self.WriteXlsx(214, 11, "ERROR")
        else:
            self.WriteXlsx(214, 11, "ERROR")
        self.WriteXlsx(214, 10, "Y")
    #科目配置查询，财务科目名称查询，特殊符号查询_0214
    def test_zzzq_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[3]").click()  # 财务科目名称
        sleep(1)
        self.Replace(215, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["财务科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:", Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(215, 11, "PASS")
        else:
            self.WriteXlsx(215, 11, "ERROR")
        self.WriteXlsx(215, 10, "Y")
    #科目配置查询，税金科目名称查询_0215
    def test_zzzr_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[4]").click()  # 税金科目名称
        sleep(1)
        self.Replace(216, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["税金科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        print("Lab1,Lab2:", Lab1, Lab2)  # 三级枚举AAA
        if Lab1 == Lab2 == "三级枚举AAA":
            self.WriteXlsx(216, 11, "PASS")
        else:
            self.WriteXlsx(216, 11, "ERROR")
        self.WriteXlsx(216, 10, "Y")
    #科目配置查询，税金科目名称查询_0216
    def test_zzzs_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[4]").click()  # 税金科目名称
        sleep(1)
        self.Replace(217, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["税金科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:", Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(217, 11, "PASS")
        else:
            self.WriteXlsx(217, 11, "ERROR")
        self.WriteXlsx(217, 10, "Y")
    #科目配置查询，税金科目名称查询_0217
    def test_zzzt_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[4]").click()  # 税金科目名称
        sleep(1)
        self.Replace(218, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["税金科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        print(Lab1, Lab2)  # 三级枚举AAA
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        Lab5 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text
        Lab6 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text
        print("Lab3~Lab4:", Lab3, Lab4)  # 一级枚举A   二级枚举AA
        if Lab1 == Lab2 == Lab5 == Lab6 == "三级枚举AAA":
            if Lab3 == "一级枚举A" and Lab4 == "二级枚举AA":
                self.WriteXlsx(218, 11, "PASS")
            else:
                self.WriteXlsx(218, 11, "ERROR")
        else:
            self.WriteXlsx(218, 11, "ERROR")
        self.WriteXlsx(218, 10, "Y")
    #科目配置查询，税金科目名称查询，特殊符号查询_0218
    def test_zzzu_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[4]").click()  # 税金科目名称
        sleep(1)
        self.Replace(219, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["税金科目名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:", Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(219, 11, "PASS")
        else:
            self.WriteXlsx(219, 11, "ERROR")
        self.WriteXlsx(219, 10, "Y")
    #科目配置查询，部门名称查询_0219
    def test_zzzv_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[6]").click()  # 部门名称
        sleep(1)
        self.Replace(220, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["部门名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[12]").text
        print("Lab1:", Lab1)  # 监查室
        if Lab1 == "监查室":
            self.WriteXlsx(220, 11, "PASS")
        else:
            self.WriteXlsx(220, 11, "ERROR")
        self.WriteXlsx(220, 10, "Y")
    #科目配置查询，部门名称查询_0220
    def test_zzzw_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[6]").click()  # 部门名称
        sleep(1)
        self.Replace(221, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["部门名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[12]").text
        print("Lab1:", Lab1)  # 办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Lab1 == "办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司":
            self.WriteXlsx(221, 11, "PASS")
        else:
            self.WriteXlsx(221, 11, "ERROR")
        self.WriteXlsx(221, 10, "Y")
    #科目配置查询，部门名称查询_0221
    def test_zzzx_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[6]").click()  # 部门名称
        sleep(1)
        self.Replace(222, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["部门名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:", Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(222, 11, "PASS")
        else:
            self.WriteXlsx(222, 11, "ERROR")
        self.WriteXlsx(222, 10, "Y")
    #科目配置查询，部门名称查询_0222
    def test_zzzy_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[6]").click()  # 部门名称
        sleep(1)
        self.Replace(223, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["部门名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[12]").text
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[12]").text
        print(Lab1, Lab2)  # 监查室 办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        Lab4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[5]").text
        Lab5 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[5]").text
        Lab6 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text
        print("Lab3~Lab4:", Lab3, Lab4,Lab5,Lab6)  # 一级枚举A   二级枚举AA
        if Lab1 == "办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司" and Lab2=="监查室":
            if Lab3 == "一级枚举A" and Lab4 == "二级枚举AA" and Lab5==Lab6=="三级枚举AAA":
                self.WriteXlsx(223, 11, "PASS")
            else:
                self.WriteXlsx(223, 11, "ERROR")
        else:
            self.WriteXlsx(223, 11, "ERROR")
        self.WriteXlsx(223, 10, "Y")
    #科目配置查询，部门名称查询，特殊符号查询_0223
    def test_zzzz_Subject_config(self):
        self.System_config()
        self.All_DJ()
        self.driver.find_element_by_xpath("//div[@id='searchInfo1']").click()  # --查询条件--
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='searchbox']/div[2]/p[6]").click()  # 部门名称
        sleep(1)
        self.Replace(224, 8)
        self.driver.find_element_by_xpath("//input[@id='searchInput1']").send_keys(C["部门名称"])
        sleep(1)
        self.driver.find_element_by_xpath("//p[@id='clkSearch1']/i").click()  # 查询
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text
        print("Lab1:", Lab1)  # 空
        if Lab1 == "":
            self.WriteXlsx(224, 11, "PASS")
        else:
            self.WriteXlsx(224, 11, "ERROR")
        self.WriteXlsx(224, 10, "Y")
    #新增自由文本科目配置_0224
    def test_zzzza_Subject_config(self):
        self.System_config()
        self.free_Frame1()
        frame  = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        self.Replace(225,8)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys(C["映射文本-自由文本(单位A)"])
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()#财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()#负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[8]/td[1]/div/input").click()#2291
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)#保存成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()#确定
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]/div").text
        print(Lab1)#收
        if Msg=="保存成功!" and Lab1=="收":
            self.WriteXlsx(225, 11, "PASS")
        else:
            self.WriteXlsx(225, 11, "ERROR")
        self.WriteXlsx(225, 10, "Y")
    #新增N9支出项目科目配置_0225
    def test_zzzzb_Subject_config(self):
        self.System_config()
        self.N9_N6()
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()#
        sleep(1)
        frame  = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//span[@id='mytree_3_span']").click()#支出事项A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_93_span']").click()  # 负债
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[8]/td[1]/div/input").click()  # 2291
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[6]/div").text
        print(Lab1)  # 2291
        if Msg == "保存成功!":
            self.WriteXlsx(226, 11, "PASS")
        else:
            self.WriteXlsx(226, 11, "ERROR")
        self.WriteXlsx(226, 10, "Y")
    #新增枚举类型科目配置_0226
    def test_zzzzc_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()#多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 测试帐套001[001]
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  #
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 一级枚举A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_2_span']").click()  # 资产
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='pDiv']/div/a[4]").click()  # 最后一页
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[2]/td[1]/div/input").click()  # 一级枚举科目1191
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[6]/div").text
        print(Lab1)  # 2291
        if Msg == "保存成功!":
            self.WriteXlsx(227, 11, "PASS")
        else:
            self.WriteXlsx(227, 11, "ERROR")
        self.WriteXlsx(227, 10, "Y")
    #新增枚举类型科目配置_0227
    def test_zzzzd_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[3]").click()  # 002
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()  # 新增按钮
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        frame = "myiframe"
        frame1 = "layui-layer-iframe1"
        frame2 = "layui-layer-iframe2"
        frame3 = "layui-layer-iframe3"
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").click()  #
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame2)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()  # 一级枚举A
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame1)
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='FinanceTR']/td/input[2]").click()  # 财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame3)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='mytree_63_span']").click()#净资产
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[1]/div/input").click()#3001
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()  # 确定
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame(frame1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[3]/a[1]").click()  # 确定
        sleep(1)
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 保存成功!
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()  # 确定
        sleep(1)
        self.driver.switch_to.frame(frame)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[6]/div").text
        print(Lab1)  # 3001
        if Msg == "保存成功!":
            self.WriteXlsx(228, 11, "PASS")
        else:
            self.WriteXlsx(228, 11, "ERROR")
        self.WriteXlsx(228, 10, "Y")
    def DJ_2019(self):
        '''年度选择2019，映射类型选择多级枚举-单位枚举(单位A)，账套选择001'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)

    #切换年度_0228
    def test_zzzze_Subject_config(self):
        self.System_config()
        self.DJ_2019()
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]").text#2020
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 一级枚举A
        print(Lab1,Lab2)
        if Lab1=="2020" and Lab2=="一级枚举A":
            self.WriteXlsx(229, 11, "PASS")
        else:
            self.WriteXlsx(229, 11, "ERROR")
        self.WriteXlsx(229, 10, "Y")
    #切换年度_0229
    def test_zzzzf_Subject_config(self):
        self.System_config()
        self.DJ_2019()
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]").text  # 2020
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 一级枚举A
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[4]").text  # 2019
        Lab4 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text  # 三级枚举AAA
        print(Lab1, Lab2,Lab3,Lab4)
        sleep(1)
        if Lab1 == "2020" and Lab2 == "一级枚举A" and Lab3=="2019" and Lab4=="三级枚举AAA":
            self.WriteXlsx(230, 11, "PASS")
        else:
            self.WriteXlsx(230, 11, "ERROR")
        self.WriteXlsx(230, 10, "Y")
    #切换映射类型_0230
    def test_zzzzg_Subject_config(self):
        self.System_config()
        self.DJ_2019()
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()#映射文本-自由文本(单位A)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 收
        print(Lab1)
        if Lab1=="收":
            self.WriteXlsx(231, 11, "PASS")
        else:
            self.WriteXlsx(231, 11, "ERROR")
        self.WriteXlsx(231, 10, "Y")
    #切换映射类型_0231
    def test_zzzzh_Subject_config(self):
        self.System_config()
        self.DJ_2019()
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  # 映射文本-自由文本(单位A)
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  # 支出事项A
        print(Lab1)
        if Lab1 == "支出事项A":
            self.WriteXlsx(232, 11, "PASS")
        else:
            self.WriteXlsx(232, 11, "ERROR")
        self.WriteXlsx(232, 10, "Y")
    #切换映射类型_0232
    def test_zzzzi_Subject_config(self):
        self.System_config()
        self.DJ_2019()
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  #
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text  #
        print("Lab1:",Lab1)
        if Lab1 == "":
            self.WriteXlsx(233, 11, "PASS")
        else:
            self.WriteXlsx(233, 11, "ERROR")
        self.WriteXlsx(233, 10, "Y")
    #切换映射类型_0233
    def test_zzzzj_Subject_config(self):
        self.System_config()
        self.DJ_2019()
        self.driver.find_element_by_xpath("//option[@title='业务枚举多级-业务枚举(单位A)']").click()  #
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//table[@id='myTable1']").text  #
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text  #
        print("Lab1:", Lab1)
        print("Lab2:", Lab2)#三级枚举AAA
        if Lab1 == "" and Lab2=="三级枚举AAA":
            self.WriteXlsx(234, 11, "PASS")
        else:
            self.WriteXlsx(234, 11, "ERROR")
        self.WriteXlsx(234, 10, "Y")
    def DJ_2019_All(self):
        '''年度选择2019，映射类型选择多级枚举-单位枚举(单位A)，账套选择全部'''
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)

    #切换账套_0234
    def test_zzzzk_Subject_config(self):
        self.System_config()
        self.DJ_2019_All()
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text  #
        print("Lab2:", Lab2)  # 三级枚举AAA
        if Lab2 == "三级枚举AAA":
            self.WriteXlsx(235, 11, "PASS")
        else:
            self.WriteXlsx(235, 11, "ERROR")
        self.WriteXlsx(235, 10, "Y")
    #切换账套_0235
    def test_zzzzl_Subject_config(self):
        self.System_config()
        self.DJ_2019_All()
        self.driver.find_element_by_xpath("//select[@id='account1']/option[3]").click()  # 002
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  #
        print("Lab2:", Lab2)  # 一级枚举A
        if Lab2 == "一级枚举A":
            self.WriteXlsx(236, 11, "PASS")
        else:
            self.WriteXlsx(236, 11, "ERROR")
        self.WriteXlsx(236, 10, "Y")
    #切换账套_0236
    def test_zzzzm_Subject_config(self):
        self.System_config()
        self.DJ_2019_All()
        self.driver.find_element_by_xpath("//select[@id='account1']/option[3]").click()  # 002
        sleep(1)
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text  #
        print("Lab2:", Lab2)  # 一级枚举A
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        Lab3 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[5]").text  #
        print("Lab3:", Lab3)  # 三级枚举AAA
        if Lab3 == "一级枚举A" and Lab2 == "一级枚举A":
            self.WriteXlsx(237, 11, "PASS")
        else:
            self.WriteXlsx(237, 11, "ERROR")
        self.WriteXlsx(237, 10, "Y")
    #年结操作提示_0237
    def test_zzzzn_Subject_config(self):
        self.System_config()
        self.DJ_2019_All()
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()#年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)#请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        # self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        if Msg=="请确认是否进行【年结】操作!":
            self.WriteXlsx(238, 11, "PASS")
        else:
            self.WriteXlsx(238, 11, "ERROR")
        self.WriteXlsx(238, 10, "Y")
    #取消年结操作_0238
    def test_zzzzo_Subject_config(self):
        self.System_config()
        self.DJ_2019_All()
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[2]").click()
        sleep(1)
        if Msg == "请确认是否进行【年结】操作!":
            self.WriteXlsx(239, 11, "PASS")
        else:
            self.WriteXlsx(239, 11, "ERROR")
        self.WriteXlsx(239, 10, "Y")
    #账套选择全部进行年结_0239
    def test_zzzzp_Subject_config(self):
        self.System_config()
        self.DJ_2019_All()
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 请先选择账套
        if Msg == "请确认是否进行【年结】操作!" and Msg1=='请先选择账套':
            self.WriteXlsx(240, 11, "PASS")
        else:
            self.WriteXlsx(240, 11, "ERROR")
        self.WriteXlsx(240, 10, "Y")
    #选择没有新年度的账套，进行年结_0240
    def test_zzzzq_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[3]").click()  # 002
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 请先选择账套
        if Msg == "请确认是否进行【年结】操作!" and Msg1 == '已是最大年度，无法向下结转!':
            self.WriteXlsx(241, 11, "PASS")
        else:
            self.WriteXlsx(241, 11, "ERROR")
        self.WriteXlsx(241, 10, "Y")
    #新年度有数据时，进行年结_0241
    def test_zzzzr_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  # 多级枚举-单位枚举(单位A)
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 0001
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 新一年的中存在配置数据，无法进行结转...
        if Msg == "请确认是否进行【年结】操作!" and Msg1 == '新一年的中存在配置数据，无法进行结转...':
            self.WriteXlsx(242, 11, "PASS")
        else:
            self.WriteXlsx(242, 11, "ERROR")
        self.WriteXlsx(242, 10, "Y")
    #旧年度没有数据时，进行年结_0242
    def test_zzzzs_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='请假类型-公共枚举']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 请先选择账套
        if Msg == "请确认是否进行【年结】操作!" and Msg1 == '没有数据可进行结转...':
            self.WriteXlsx(243, 11, "PASS")
        else:
            self.WriteXlsx(243, 11, "ERROR")
        self.WriteXlsx(243, 10, "Y")
    #正常年结N9支出事项科目配置_0243
    def test_zzzzt_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 数据正在结转，结转完毕后会推送消息进行提醒...
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg=="请确认是否进行【年结】操作!":
            if Msg1=="数据正在结转，结转完毕后会推送消息进行提醒...":
                self.WriteXlsx(244, 11, "PASS")
            else:
                self.WriteXlsx(244, 11, "ERROR")
        else:
            self.WriteXlsx(244, 11, "ERROR")
        self.WriteXlsx(244, 10, "Y")
    #正常年结数据验证N9支出事项科目配置_0244
    def test_zzzzu_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]").text#2020
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[7]").text  # 2291[科目A]
        print(Lab1,Lab2)
        if Lab1=="2020":
            self.WriteXlsx(245, 11, "PASS")
        else:
            self.WriteXlsx(245, 11, "ERROR")
        self.WriteXlsx(245, 10, "Y")
    #正常年结自由文本科目配置_0245
    def test_zzzzv_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 数据正在结转，结转完毕后会推送消息进行提醒...
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "请确认是否进行【年结】操作!":
            if Msg1 == "数据正在结转，结转完毕后会推送消息进行提醒...":
                self.WriteXlsx(246, 11, "PASS")
            else:
                self.WriteXlsx(246, 11, "ERROR")
        else:
            self.WriteXlsx(246, 11, "ERROR")
        self.WriteXlsx(246, 10, "Y")
    #正常年结数据验证自由文本科目配置_0246
    def test_zzzzw_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[4]").text  # 2020
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[7]").text  # 2291[科目A]
        print(Lab1, Lab2)
        if Lab1 == "2020":
            self.WriteXlsx(247, 11, "PASS")
        else:
            self.WriteXlsx(247, 11, "ERROR")
        self.WriteXlsx(247, 10, "Y")
    #删除2020年的枚举科目配置_0247
    def test_zzzzx_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[1]").click()  # 全部
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()#s删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 删除成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            if Msg1 == "删除成功!":
                self.WriteXlsx(248, 11, "PASS")
            else:
                self.WriteXlsx(248, 11, "ERROR")
        else:
            self.WriteXlsx(248, 11, "ERROR")
        self.WriteXlsx(248, 10, "Y")
    #正常年结枚举科目配置_0248
    def test_zzzzy_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='YearEnd1']").click()  # 年结
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 请确认是否进行【年结】操作!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 数据正在结转，结转完毕后会推送消息进行提醒...
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "请确认是否进行【年结】操作!":
            if Msg1 == "数据正在结转，结转完毕后会推送消息进行提醒...":
                self.WriteXlsx(249, 11, "PASS")
            else:
                self.WriteXlsx(249, 11, "ERROR")
        else:
            self.WriteXlsx(249, 11, "ERROR")
        self.WriteXlsx(249, 10, "Y")
    #正常年结数据验证枚举科目配置_0249
    def test_zzzzz_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        Lab1 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td[12]").text  # 监查室
        Lab2 = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[4]/td[12]").text  #
        print(Lab1,Lab2)#办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司
        if Lab1=="监查室" and Lab2=="办公室,财务室,综合科,水政水资源科,办公厅,综合司,条法司":
            self.WriteXlsx(250, 11, "PASS")
        else:
            self.WriteXlsx(250, 11, "ERROR")
        self.WriteXlsx(250, 10, "Y")
    #删除单位枚举科目配置数据_0250
    def test_zzzzza_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # s删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 删除成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            if Msg1 == "删除成功!":
                self.WriteXlsx(251, 11, "PASS")
            else:
                self.WriteXlsx(251, 11, "ERROR")
        else:
            self.WriteXlsx(251, 11, "ERROR")
        self.WriteXlsx(251, 10, "Y")
    #删除单位枚举科目配置数据_0251
    def test_zzzzzb_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='多级枚举-单位枚举(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # s删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 删除成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            if Msg1 == "删除成功!":
                self.WriteXlsx(252, 11, "PASS")
            else:
                self.WriteXlsx(252, 11, "ERROR")
        else:
            self.WriteXlsx(252, 11, "ERROR")
        self.WriteXlsx(252, 10, "Y")
    #删除新增的自由文本科目配置_0252
    def test_zzzzzc_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # s删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 删除成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            if Msg1 == "删除成功!":
                self.WriteXlsx(253, 11, "PASS")
            else:
                self.WriteXlsx(253, 11, "ERROR")
        else:
            self.WriteXlsx(253, 11, "ERROR")
        self.WriteXlsx(253, 10, "Y")
    #删除新增的自由文本科目配置_0253
    def test_zzzzzd_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='映射文本-自由文本(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # s删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 删除成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            if Msg1 == "删除成功!":
                self.WriteXlsx(254, 11, "PASS")
            else:
                self.WriteXlsx(254, 11, "ERROR")
        else:
            self.WriteXlsx(254, 11, "ERROR")
        self.WriteXlsx(254, 10, "Y")
    #删除科目配置，单行删除_0254
    def test_zzzzze_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[1]").click()  # 2019
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # s删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 删除成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            if Msg1 == "删除成功!":
                self.WriteXlsx(255, 11, "PASS")
            else:
                self.WriteXlsx(255, 11, "ERROR")
        else:
            self.WriteXlsx(255, 11, "ERROR")
        self.WriteXlsx(255, 10, "Y")
    #删除科目配置，单行删除_0255
    def test_zzzzzf_Subject_config(self):
        self.System_config()
        self.driver.find_element_by_xpath("//li[@id='tab1']").click()  # 点击借方科目配置
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='acctYear1']/option[2]").click()  # 2020
        sleep(1)
        self.driver.find_element_by_xpath("//option[@title='N6/N9支出事项-系统档案(单位A)']").click()  #
        sleep(1)
        self.driver.find_element_by_xpath("//select[@id='account1']/option[2]").click()  # 001
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='hDivBox']/table/thead/tr/th[1]/div/input").click()  # 全选
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # s删除
        sleep(1)
        self.driver.switch_to.default_content()
        Msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg)  # 数据删除后将无法恢复,请确认是否进行删除!
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        Msg1 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
        print(Msg1)  # 删除成功
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        if Msg == "数据删除后将无法恢复,请确认是否进行删除!":
            if Msg1 == "删除成功!":
                self.WriteXlsx(256, 11, "PASS")
            else:
                self.WriteXlsx(256, 11, "ERROR")
        else:
            self.WriteXlsx(256, 11, "ERROR")
        self.WriteXlsx(256, 10, "Y")

if __name__ == '__main__':
    unittest.main(verbosity=2)
