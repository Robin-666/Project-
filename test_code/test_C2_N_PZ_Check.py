import unittest,re,time
from test_page.ZT_information import ZT_ZM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log
from test_page.People_Assist_JM import People_Assist
from test_page.Wait_ZD_JM import Wait_ZD



class TestC2_01(unittest.TestCase):
    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        # cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\凭证制单_凭证查询.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["凭证制单_凭证查询"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\凭证制单_凭证查询.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["凭证制单_凭证查询"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ZT_infor
        ZT_infor = ZT_ZM(self.driver, self.base_url, '')

    def People_FZHS(self):
        global PEOPLE
        PEOPLE = People_Assist(self.driver, self.base_url, '')



    def Log_In_Wait_ZD(self):
        '''从登陆界面进入凭证查询界面'''
        global page, Data ,Wait_ZD_ZJM
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        Wait_ZD_ZJM = Wait_ZD(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        page.CWJC()
        page.Win_PZZD()
        page.Win_Split()
        Wait_ZD_ZJM.PZ_Check_JM()
        Wait_ZD_ZJM.switch_frame(Data.myframe)

    def Clear_All_Input(self):
        js = "$('input[id=formNames]').attr('readonly',false)"  # 3.jQuery，设置为false
        self.driver.execute_script(js)  # 将只能作为点击的（点击选择表单）状态转换为可清除的状态
        Wait_ZD_ZJM.Clear_Form_Select_Button()
        Wait_ZD_ZJM.Sel_Account_All()
        Wait_ZD_ZJM.PZ_Text_XL()
        Wait_ZD_ZJM.Clear_ZD_People()
        Wait_ZD_ZJM.Clear_Synergy_title()
        Wait_ZD_ZJM.Clear_Disest_infor()
        Wait_ZD_ZJM.Clear_PZ_Min_Number()
        Wait_ZD_ZJM.Clear_PZ_Max_Number()

    def test_A_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        Wait_ZD_ZJM.ZT_PZ_Input_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_U8_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS()=="条/共0条记录":
            self.WriteXlsx(2, 11, "PASS")
        else:
            self.WriteXlsx(2, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(2, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.ZT_PZ_Input_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Select_2020_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(3, 11, "PASS")
        else:
            self.WriteXlsx(3, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(3, 10, "Y")


    def test_B_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        #预带了上次的2020账套，此处不再进行账套的选择
        Wait_ZD_ZJM.Form_Select_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Select_C2_Button()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共3条记录":
            self.WriteXlsx(4, 11, "PASS")
        else:
            self.WriteXlsx(4, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(4, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Form_Select_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Select_Robin_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共9条记录":
            self.WriteXlsx(5, 11, "PASS")
        else:
            self.WriteXlsx(5, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(5, 10, "Y")

    def test_C_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        Wait_ZD_ZJM.Form_Select_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.DR_PZ()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(6, 11, "PASS")
        else:
            self.WriteXlsx(6, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(6, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Form_Select_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Unit_MJ()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(7, 11, "PASS")
        else:
            self.WriteXlsx(7, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(7, 10, "Y")

    def test_D_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        Wait_ZD_ZJM.Sel_Account_1()
        self.Glo()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共17条记录":
            self.WriteXlsx(8, 11, "PASS")
        else:
            self.WriteXlsx(8, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(8, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Sel_Account_2()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(9, 11, "PASS")
        else:
            self.WriteXlsx(9, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(9, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Sel_Account_3()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(10, 11, "PASS")
        else:
            self.WriteXlsx(10, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(10, 10, "Y")

    def test_E_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        Wait_ZD_ZJM.PZ_Text_YQ()
        self.Glo()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共18条记录":
            self.WriteXlsx(11, 11, "PASS")
        else:
            self.WriteXlsx(11, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(11, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.PZ_Text_CAI()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(12, 11, "PASS")
        else:
            self.WriteXlsx(12, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(12, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.PZ_Text_JI()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(13, 11, "PASS")
        else:
            self.WriteXlsx(13, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(13, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.PZ_Text_BY()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(14, 11, "PASS")
        else:
            self.WriteXlsx(14, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(14, 10, "Y")

    def test_F_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        Wait_ZD_ZJM.Clear_PZ_Min_Number()
        Wait_ZD_ZJM.Clear_PZ_Max_Number()
        self.Replace(15,8)
        Wait_ZD_ZJM.PZ_Min_Number(C["第一输入框"])
        Wait_ZD_ZJM.PZ_Max_Number(C["第二输入框"])
        self.Glo()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共9条记录":
            self.WriteXlsx(15, 11, "PASS")
        else:
            self.WriteXlsx(15, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(15, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Clear_PZ_Min_Number()
        Wait_ZD_ZJM.Clear_PZ_Max_Number()
        self.Replace(16, 8)
        Wait_ZD_ZJM.PZ_Min_Number(C["第一输入框"])
        Wait_ZD_ZJM.PZ_Max_Number(C["第二输入框"])
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共10条记录":
            self.WriteXlsx(16, 11, "PASS")
        else:
            self.WriteXlsx(16, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(16, 10, "Y")
        Wait_ZD_ZJM.Clear_PZ_Min_Number()
        Wait_ZD_ZJM.Clear_PZ_Max_Number()


    def test_G_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        Wait_ZD_ZJM.Click_ZD_People()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.JCS_KM()
        Wait_ZD_ZJM.Select_CXB_people()
        Wait_ZD_ZJM.Select_Right_Button()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共7条记录":
            self.WriteXlsx(17, 11, "PASS")
        else:
            self.WriteXlsx(17, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(17, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Click_ZD_People()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        Wait_ZD_ZJM.Select_CXB_people()
        Wait_ZD_ZJM.Select_Left_Button()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()

        '''以下代码：选择不同制单人时可以清除界面显示，
        但是在选择人员的界面右侧的人员不能全部移到左边，
        ，再选择人员时，会出现不同的人员到界面上，会造成查询数据异常的情况，暂时不考虑制单人选择的查询情况'''
        # Wait_ZD_ZJM.Condition_Check_Button()
        # Wait_ZD_ZJM.switch_frame_default()
        # Wait_ZD_ZJM.switch_frame(Data.frame3)
        # Wait_ZD_ZJM.Clear_ZD_People()
        # Wait_ZD_ZJM.Click_ZD_People()
        # Wait_ZD_ZJM.switch_frame_default()
        # Wait_ZD_ZJM.switch_frame(Data.frame4)
        # Wait_ZD_ZJM.JCS_KM()
        # Wait_ZD_ZJM.Select_JCS_A_people()
        # Wait_ZD_ZJM.Select_Right_Button()
        # Wait_ZD_ZJM.switch_frame_default()
        # ZT_infor.Save_frame4()
        # Wait_ZD_ZJM.switch_frame_default()
        # ZT_infor.Save_frame3()
        # Wait_ZD_ZJM.switch_frame(Data.myframe)
        # if ZT_infor.CW_XSYS() == "条/共0条记录":
        #     self.WriteXlsx(18, 11, "PASS")
        # else:
        #     self.WriteXlsx(18, 11, "ERROR")
        #     do_log.error("报错：实际与预期不符")
        # self.WriteXlsx(18, 10, "Y")

    def test_H_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        self.Clear_All_Input()
        Wait_ZD_ZJM.Clear_Synergy_title()
        self.Replace(20,8)
        Wait_ZD_ZJM.Synergy_title_Send_Key(C["协同标题"])
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共8条记录":
            self.WriteXlsx(20, 11, "PASS")
        else:
            self.WriteXlsx(20, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(20, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Clear_Synergy_title()
        self.Replace(21, 8)
        Wait_ZD_ZJM.Synergy_title_Send_Key(C["协同标题"])
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共0条记录":
            self.WriteXlsx(21, 11, "PASS")
        else:
            self.WriteXlsx(21, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(21, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Clear_Synergy_title()
        self.Replace(22, 8)
        Wait_ZD_ZJM.Synergy_title_Send_Key(C["协同标题"])
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共6条记录":
            self.WriteXlsx(22, 11, "PASS")
        else:
            self.WriteXlsx(22, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(22, 10, "Y")
        Wait_ZD_ZJM.Clear_Synergy_title()

    def test_I_PZ_Check(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Clear_Disest_infor()
        self.Replace(23,8)
        Wait_ZD_ZJM.Disest_infor_Send_Key(C["摘要标题"])
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共15条记录":
            self.WriteXlsx(23, 11, "PASS")
        else:
            self.WriteXlsx(23, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(23, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Clear_Disest_infor()
        self.Replace(24, 8)
        Wait_ZD_ZJM.Disest_infor_Send_Key(C["摘要标题"])
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共2条记录":
            self.WriteXlsx(24, 11, "PASS")
        else:
            self.WriteXlsx(24, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(24, 10, "Y")
        Wait_ZD_ZJM.Condition_Check_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Clear_Disest_infor()
        self.Replace(25, 8)
        Wait_ZD_ZJM.Disest_infor_Send_Key(C["摘要标题"])
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共16条记录":
            self.WriteXlsx(25, 11, "PASS")
        else:
            self.WriteXlsx(25, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(25, 10, "Y")

















if __name__ == '__main__':
    unittest.main(verbosity=2)




















