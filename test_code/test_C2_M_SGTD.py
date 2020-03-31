import unittest,re,time
from test_page.ZT_information import ZT_ZM
from selenium import webdriver
from openpyxl import load_workbook
from test_page.Subject_Page import Subject_Config_JM
from test_page.ConfigData import AllData
from time import sleep
from report_Auto import do_log
from test_page.People_Assist_JM import People_Assist
from test_page.Wait_ZD_JM import Wait_ZD



class TestC2_01(unittest.TestCase):
    # @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        Data = AllData()
        cls.base_url = Data.URL
        print("Test Start")

    # @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")


    def Replace(self,row, column):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\凭证制单_手工填单.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["凭证制单_手工填单"]
        A = worksheetname.cell(row, column).value
        regexL = r'{.*?}'
        regL = re.compile(regexL, re.S)
        B = re.findall(regL, A)
        global C
        C = eval(B[0])  # str->dict
        Wb.save(Excel_path)

    def WriteXlsx(self,row,column,data):
        Excel_path = r"E:\Lexmis_Auto_Object\Lexmis_V71_SP1\C2_CW_NEW\test_case\凭证制单_手工填单.xlsx"
        # Excel_path = '../test_case\科目配置.xlsx'
        Wb = load_workbook(Excel_path)
        worksheetname = Wb["凭证制单_手工填单"]
        name = worksheetname.title  # 获取表名
        # print('当前sheet-name： ',name)
        worksheetname.cell(row,column,data)
        Wb.save(Excel_path)

    def Glo(self):
        '''定义全局变量，在class外面未定义成功——未知原因'''
        global ZT_infor
        ZT_infor = ZT_ZM(self.driver, self.base_url, '')

    def People_FZHS(self):
        global PEOPLE
        PEOPLE = People_Assist(self.driver, self.base_url, '')



    def Log_In_Wait_ZD(self):
        '''从登陆界面进入手工填单界面'''
        global page, Data ,Wait_ZD_ZJM
        page = Subject_Config_JM(self.driver, self.base_url, '')
        Data = AllData()
        Wait_ZD_ZJM = Wait_ZD(self.driver, self.base_url, '')
        page.open()
        page.input_username(Data.name_b)
        page.input_password(Data.pwd)
        page.click_submit()
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        page.CWJC()
        page.Win_PZZD()
        page.Win_Split()
        Wait_ZD_ZJM.SG_ZD_JM()
        Wait_ZD_ZJM.switch_frame(Data.myframe)


    def test_A_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_2020_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.SG_ZD_JM_New_Add_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        js = "$('input[id=billdate]').attr('readonly',false)"  # 3.jQuery，设置为false
        # js = "$('input[id=billdate]').attr('readonly','')"  # 4.jQuery，设置为空（同3）
        self.driver.execute_script(js)#操作日期控件
        Wait_ZD_ZJM.Clear_Data_time()
        Wait_ZD_ZJM.Data_time(Wait_ZD_ZJM.Call_Data_time())
        self.Replace(2, 8)
        Wait_ZD_ZJM.Double_Click_Input_First()
        Wait_ZD_ZJM.Input_First_1_text_ZY(C["第一行摘要"])
        Wait_ZD_ZJM.Double_Click_Input_First_2()
        Wait_ZD_ZJM.Input_First_2_KM_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        ZT_infor.CW_mytree_162_span()
        Wait_ZD_ZJM.Select_6001()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.JF_RMB()
        Wait_ZD_ZJM.Input_JF_RMB(C["借方金额"])
        Wait_ZD_ZJM.Double_Click_Input_Two()
        Wait_ZD_ZJM.Input_Two_2_text_ZY(C["第二行摘要信息"])
        Wait_ZD_ZJM.Double_Click_Input_Two_2()
        Wait_ZD_ZJM.Input_Two_2_KM_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame4)
        ZT_infor.CW_mytree_162_span()
        Wait_ZD_ZJM.Select_6031()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame4()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.DF_RMB()
        Wait_ZD_ZJM.Input_DF_RMB(C["贷方金额"])
        Wait_ZD_ZJM.Button_Save()
        sleep(5)
        if page.Dialog()=="保存成功":
            self.WriteXlsx(2, 11, "PASS")
        else:
            self.WriteXlsx(2, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(2, 10, "Y")
        page.switch_frame_default()
        page.OK_Msg()


    def test_B_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_2020_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        # Wait_ZD_ZJM.Clear_Sel_PZZ()
        Wait_ZD_ZJM.Sel_PZZ_C()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS()=="条/共0条记录":
            self.WriteXlsx(3, 11, "PASS")
        else:
            self.WriteXlsx(3, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(3, 10, "Y")
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Sel_PZZ_YQ()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共2条记录":
            self.WriteXlsx(4, 11, "PASS")
        else:
            self.WriteXlsx(4, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(4, 10, "Y")

    def test_C_Not_ZD(self):
        self.Log_In_Wait_ZD()
        Wait_ZD_ZJM.Select_ZT_Button()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame1)
        Wait_ZD_ZJM.Select_2020_ZT()
        Wait_ZD_ZJM.switch_frame_default()
        self.Glo()
        ZT_infor.Save_frame1()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame2)
        Wait_ZD_ZJM.Sel_Account_1()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame2()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(5, 11, "PASS")
        else:
            self.WriteXlsx(5, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(5, 10, "Y")
        Wait_ZD_ZJM.Advanced_Query()
        Wait_ZD_ZJM.switch_frame_default()
        Wait_ZD_ZJM.switch_frame(Data.frame3)
        Wait_ZD_ZJM.Sel_Account_3()
        Wait_ZD_ZJM.switch_frame_default()
        ZT_infor.Save_frame3()
        Wait_ZD_ZJM.switch_frame(Data.myframe)
        if ZT_infor.CW_XSYS() == "条/共1条记录":
            self.WriteXlsx(6, 11, "PASS")
        else:
            self.WriteXlsx(6, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(6, 10, "Y")
        Wait_ZD_ZJM.Up_Data_info()
        Wait_ZD_ZJM.switch_frame_default()
        if page.Dialog()=="整理成功":
            self.WriteXlsx(7, 11, "PASS")
        else:
            self.WriteXlsx(7, 11, "ERROR")
            do_log.error("报错：实际与预期不符")
        self.WriteXlsx(7, 10, "Y")




























if __name__ == '__main__':
    unittest.main(verbosity=2)




















