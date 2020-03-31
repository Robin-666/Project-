from openpyxl import load_workbook
import openpyxl,re

def WriteXlsx():
    Wb = load_workbook("C2+财务测试用例_合并.xlsx")
    worksheetname=Wb["系统配置-辅助核算"]
    # print(worksheetname)
    name = worksheetname.title  # 获取表名
    print(name)
    worksheetname.cell(25,16,"Robin")
    Wb.save("C2+财务测试用例_合并.xlsx")



# WriteXlsx()
# Wb = load_workbook(r"E:\Auto_C2+\test_case\C2+Pro_list.xlsx")
# wb1 = Wb.active
# wb1.cell(row,column,data)
# sleep(1)
# Wb.save(r"E:\Auto_C2+\test_case\C2+Pro_list.xlsx")

def Replace(row,column,data):
    Wb = load_workbook(r"E:\Auto_C2+\test_case\C2+Pro_list.xlsx")
    worksheetname=Wb["系统配置-服务器配置"]
    # print(worksheetname.cell(23,7).value)
    A = worksheetname.cell(row,column).value
    # print(A)
    regexL = "{.*?}"
    regL = re.compile(regexL)
    B = re.findall(regL,A)[data]
    print(B.replace('{','').replace('}',''))
    return B.replace('{','').replace('}','')
    Wb.save(r"E:\Auto_C2+\test_case\C2+Pro_list.xlsx")



Replace(4,7,1)
