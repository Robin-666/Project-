import unittest
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.support.select import Select
 
class TestC2_01(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.172:8082"
        # print("Test Start")

    @classmethod
    def tearDown(cls):
        cls.driver.quit()
        print("Test End")



    def WriteXlsx(self,row,column,data):
        Wb = load_workbook(r"E:\Auto_C2+\test_case\test_C2_凭证制单.xlsx")
        wb1 = Wb.active
        wb1.cell(row,column,data)
        sleep(1)
        Wb.save(r"E:\Auto_C2+\test_case\test_C2_凭证制单.xlsx")
    @unittest.skip("直接跳过")
    def testLog_a_页面登录(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("sc01")  #
        sleep(1)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(1)
        self.driver.find_element_by_css_selector("#login_button").click()
        self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        print("当前窗口名称：",self.driver.title)
        if self.driver.title =="A8+ 国际版（集团版）协同管理软件 V7.1SP1, 四川演示01,您好!":
            self.WriteXlsx(2, 7, "Pass")
            self.WriteXlsx(2, 8, "罗彬")
        else:
            self.WriteXlsx(2, 7, "Error")
            self.WriteXlsx(2, 8, "罗彬")

    @unittest.skip("直接跳过")
    def testLog_b_进入新建界面(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("sc01")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        ZJM = self.driver.current_window_handle
        self.driver.find_element_by_xpath("//table[@title='U8-C2_First_Data']").click()  # 在模板点击U8-C2_First_Data
        sleep(0.5)
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                print("进入新建界面窗口名称：",self.driver.title)
                sleep(1)
        if self.driver.title == "新建页面":
            self.WriteXlsx(3, 7, "Pass")
            self.WriteXlsx(3, 8, "罗彬")
        else:
            self.WriteXlsx(3, 7, "Error")
            self.WriteXlsx(3, 8, "罗彬")

    def New_Page(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("sc01")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        ZJM = self.driver.current_window_handle
        sleep(1)
        self.driver.find_element_by_xpath("//table[@title='U8-C2_First_Data']").click()  # 在模板点击U8-C2_First_Data
        sleep(0.5)
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                print("进入新建界面窗口名称：", self.driver.title)
                sleep(1)


    # @unittest.skip("直接跳过")
    def test_Log_c_Data_Dispose(self):
        self.New_Page()
        sleep(2)
        self.driver.switch_to.frame("zwIframe")
        self.driver.find_element_by_xpath("//div[@class='render-eg__rightbox']").click()#去掉警告信息
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='account_field0001']").click()#账套配置
        sleep(2)
        self.driver.switch_to.default_content()
        self.driver.switch_to.frame("layui-layer-iframe1")
        # self.driver.switch_to.frame("layui-layer-iframe2")
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[@align='center']").click()
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#确定按钮
        self.driver.switch_to.frame("zwIframe")
        self.driver.find_element_by_xpath("//input[@id='field0004_format']").send_keys("2019-11-05")
        sleep(1)
        self.driver.find_element_by_xpath("//div[@class='icon CAP cap-icon-xuanbumen']").click()#
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.switch_to.frame("layui-layer-iframe2")
        self.driver.find_element_by_xpath("//div[@title='测试部']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@title='选择']").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        # sleep(1)
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#确定
        sleep(1)
        self.driver.switch_to.frame("zwIframe")
        self.driver.find_element_by_xpath("//div[@class='icon CAP cap-icon-xuanren']").click()#发起人
        self.driver.switch_to.default_content()
        Sel_People = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe']")
        self.driver.switch_to.frame(Sel_People)#先进行定位在切换iframe
        self.driver.find_element_by_xpath("//div[@title='测试部']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//option[@type='Member']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//span[@title='选择']").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 确定
        self.driver.switch_to.frame("zwIframe")
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='account_field0011']").click()#部门
        self.driver.switch_to.default_content()
        Sel_Dep = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe']")
        self.driver.switch_to.frame(Sel_Dep)  # 先进行定位在切换iframe
        self.driver.find_element_by_xpath("//a[@id='mytree_21_a']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@row='0']").click()
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 确定
        self.driver.switch_to.frame("zwIframe")
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='account_field0019']").click()#客户
        self.driver.switch_to.default_content()
        sleep(1)
        Sel_Customer = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe']")
        self.driver.switch_to.frame(Sel_Customer)  # 先进行定位在切换iframe
        self.driver.find_element_by_xpath("//a[@id='mytree_5_a']").click()#客户往来
        sleep(1)
        self.driver.find_element_by_xpath("//input[@row='0']").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 确定
        self.driver.switch_to.frame("zwIframe")
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='account_field0015']").click()#供应商
        self.driver.switch_to.default_content()
        # sleep(1)
        # self.driver.switch_to.default_content()
        sleep(1)
        Sel_Sup = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe']")
        self.driver.switch_to.frame(Sel_Sup)
        self.driver.find_element_by_xpath("//a[@id='mytree_6_a']").click()#05 公共原料供应商
        sleep(1)
        self.driver.find_element_by_xpath("//input[@row='2']").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()  # 确定
        self.driver.switch_to.frame("zwIframe")
        sleep(1)
        self.driver.find_element_by_xpath("//div[@id='account_field0017']").click()#项目
        self.driver.switch_to.default_content()
        sleep(1)
        Sel_Pro = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe']")
        self.driver.switch_to.frame(Sel_Pro)
        self.driver.find_element_by_xpath("//span[@id='mytree_9_span']").click()#现金流量项目
        sleep(1)
        self.driver.find_element_by_xpath("//div[@title='06']").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()
        self.driver.switch_to.frame("zwIframe")
        js = "var q=document.documentElement.scrollTop=300"
        self.driver.execute_script(js)
        print("操作滚动条")
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='cap4-formmain__mTable']/tr[21]/td[5]/div/div/section/div/div/input[@type='text']").send_keys("收付")
        sleep(1)
        self.driver.find_element_by_xpath("//table[@class='cap4-formmain__mTable']/tr[21]/td[7]/div/div/section/div/div/div[2]/input[@type='text']").send_keys("920000")
        sleep(1)
        self.driver.switch_to.default_content()
        sleep(1)
        ZJM = self.driver.current_window_handle
        self.driver.find_element_by_xpath("//a[@id='sendId']").click()#发送
        sleep(3)
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                print("主界面窗口名称：", self.driver.title)
                sleep(1)
        self.driver.find_element_by_xpath("//div[@title='协同工作']").click()
        sleep(1)
        Move1 = self.driver.find_element_by_xpath("//div[@title='待办事项']")
        ActionChains(self.driver).move_to_element(Move1).perform()
        self.driver.find_element_by_xpath("//div[@title='待办事项']").click()#进入待办事项
        sleep(1)
        self.driver.switch_to.frame("mainIframe")
        self.driver.find_element_by_xpath("//div[@id='allPending']").click()#避开悬浮点击全部待办
        sleep(1)
        msg1 = self.driver.find_element_by_xpath("//table[@class='flexme3 ']/tbody[@id='list']/tr/td[2]/div/span[1]").text
        print(msg1)
        if 'U8-C2_First_Data'in msg1:
            self.WriteXlsx(4, 7, "Pass")
            self.WriteXlsx(4, 8, "罗彬")
        else:
            self.WriteXlsx(4, 7, "Error")
            self.WriteXlsx(4, 8, "罗彬")
        print("调用的表单在待办事项中。。")
    # @unittest.skip("直接跳过")
    # def test_Log_d_Data_Dispose单据审批(self):
    #     self.driver.get(self.base_url)
    #     self.driver.maximize_window()
    #     self.driver.find_element_by_css_selector("#login_username").send_keys("sc01")  #
    #     sleep(0.5)
    #     self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
    #     sleep(0.5)
    #     self.driver.find_element_by_css_selector("#login_button").click()
    #     self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
    #     self.driver.find_element_by_xpath("//div[@title='协同工作']").click()
    #     Move1 = self.driver.find_element_by_xpath("//div[@title='待办事项']")
    #     ActionChains(self.driver).move_to_element(Move1).perform()
    #     self.driver.find_element_by_xpath("//div[@title='待办事项']").click()  # 进入待办事项
    #     sleep(1)
    #     self.driver.switch_to.frame("mainIframe")
    #     sleep(1)
    #     self.driver.find_element_by_xpath("//div[@id='allPending']").click()  # 避开悬浮点击全部待办
    #     sleep(1)
        ZJM1 = self.driver.current_window_handle
        # self.driver.find_element_by_xpath("table[@class='flexme3 ']/tbody[@id='list']/tr/td/div/input[@type='checkbox']").click()
        self.driver.find_element_by_xpath("//table[@class='flexme3 ']/tbody[@id='list']/tr/td[2]/div").click()
        sleep(2)
        Win1 = self.driver.window_handles  # 所有窗口句柄
        for windows1 in Win1:
            if windows1 != ZJM1:
                self.driver.switch_to.window(windows1)
                print("审批表单的名称：", self.driver.title)
        self.driver.find_element_by_xpath("//textarea[@id='content_deal_comment']").send_keys("100个赞赞赞,同意")
        sleep(2)
        # self.driver.find_element_by_xpath("//div[@id='_dealDiv']/div[4]/input[@title='同意']").click()
        self.driver.find_element_by_xpath("//input[@title='同意']").click()
        sleep(1)
        print("审批通过")
        self.WriteXlsx(5, 7, "Pass")
        self.WriteXlsx(5, 8, "罗彬")
    # @unittest.skip("直接跳过")
    # def test_Log_e_Data_Dispose凭证制单(self):
    #     self.driver.get(self.base_url)
    #     self.driver.maximize_window()
    #     self.driver.find_element_by_css_selector("#login_username").send_keys("sc01")  #
    #     sleep(0.5)
    #     self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
    #     sleep(0.5)
    #     self.driver.find_element_by_css_selector("#login_button").click()
    #     self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
    #     self.driver.switch_to.frame("zwIframe")
        self.driver.switch_to.window(ZJM1)
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        move = self.driver.find_element_by_xpath("//div[@title='凭证制单']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='凭证制单']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                print("系统配置界面窗口名称：", self.driver.title)
                sleep(1)
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='formTable']/tbody/tr/td/div/input[@type='checkbox']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//button[@id='making']").click()#凭证制单
        sleep(1)
        self.driver.switch_to.default_content()
        # Sel_PZ = self.driver.find_element_by_xpath("//div[@id='makingdialog_main']/iframe")
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='billdate']").send_keys("2019-11-06")
        sleep(1)
        self.driver.find_element_by_xpath("//button[@id='savepz']").click()
        print("保存成功")
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print("msg2: ", msg2)
        if msg2 == "保存成功":
            self.WriteXlsx(6, 7, "Pass")
            self.WriteXlsx(6, 8, "罗彬")
        else:
            self.WriteXlsx(6, 7, "Error")
            self.WriteXlsx(6, 8, "罗彬")
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()


if __name__ == '__main__':
    unittest.main(verbosity=2)
