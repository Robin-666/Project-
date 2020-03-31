import unittest
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from selenium.webdriver.support.select import Select

class TestC2_01(unittest.TestCase):
    @classmethod
    def setUp(cls):
        cls.driver = webdriver.Chrome()
        cls.base_url = "http://192.168.20.172:8082"
        # print("Test Start")

    @classmethod
    def tearDown(cls):
        # cls.driver.quit()
        print("Test End")


    def WriteXlsx(self,row,column,data):
        Wb = load_workbook(r"E:\Auto_C2+\test_case\test_C2_系统配置.xlsx")
        wb1 = Wb.active
        wb1.cell(row,column,data)
        sleep(1)
        Wb.save(r"E:\Auto_C2+\test_case\test_C2_系统配置.xlsx")
    @unittest.skip("直接跳过")
    def testLog_a_页面登录(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxd")  #
        sleep(1)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(1)
        self.driver.find_element_by_css_selector("#login_button").click()
        self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        print("当前窗口名称：",self.driver.title)
        if self.driver.title =="A8+ 国际版（集团版）协同管理软件 V7.1SP1, 四川演示01,您好!":
            self.WriteXlsx(2, 7, "Pass")
            self.WriteXlsx(2, 8, "罗彬")
        else:
            self.WriteXlsx(2, 7, "Error")
            self.WriteXlsx(2, 8, "罗彬")

    @unittest.skip("直接跳过")
    def testLog_b_进入系统配置界面(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("sc01")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        # self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                print("系统配置界面窗口名称：",self.driver.title)
                sleep(1)
        if self.driver.title == "系统配置":
            self.WriteXlsx(3, 7, "Pass")
            self.WriteXlsx(3, 8, "罗彬")
        else:
            self.WriteXlsx(3, 7, "Error")
            self.WriteXlsx(3, 8, "罗彬")

    @unittest.skip("直接跳过")
    def test_Log_c_ERP配置(self):
        pass
    def System_Config(self):
        self.driver.get(self.base_url)
        self.driver.maximize_window()
        self.driver.find_element_by_css_selector("#login_username").send_keys("cxd")  #
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_password").send_keys("123456")  # 密码
        sleep(0.5)
        self.driver.find_element_by_css_selector("#login_button").click()
        self.driver.implicitly_wait(10)  # 等待，直到超出10s还未定位到元素则异常
        # sleep(3)
        self.driver.find_element_by_xpath("//div[@title='财务集成']").click()  # 点击财务集成
        sleep(0.5)
        ZJM = self.driver.current_window_handle
        move = self.driver.find_element_by_xpath("//div[@title='系统配置']")  # 鼠标悬停
        ActionChains(self.driver).move_to_element(move).perform()
        self.driver.find_element_by_xpath("//div[@title='系统配置']").click()
        Win = self.driver.window_handles  # 所有窗口句柄
        for windows in Win:
            if windows != ZJM:
                self.driver.switch_to.window(windows)
                print("系统配置界面窗口名称：", self.driver.title)
                sleep(1)

    # @unittest.skip("直接跳过")
    def test_Log_d_服务器配置_U8(self):
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        # self.driver.find_element_by_class_name("#btn1").click()
        # sleep(1)
        # self.driver.find_element_by_class_name("db_name").send_keys("ufsystem")#服务器名称
        # self.driver.find_element_by_class_name("db_hostip").send_keys("192.168.20.182")#服务器地址
        # self.driver.find_element_by_class_name("db_port").send_keys("1433")#端口
        # self.driver.find_element_by_class_name("db_user").send_keys("sa")
        # self.driver.find_element_by_class_name("db_pwd").send_keys("abc=123")
        # self.driver.find_element_by_class_name("dbtest").click()
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[1]").click()
        self.driver.switch_to.window(self.driver.window_handles[1])
        msg = self.driver.find_element_by_class_name("dialog_main_content_html ").text
        print(msg)
        if msg =="当前数据库信息可用!":
            self.WriteXlsx(5, 7, "Pass")
            self.WriteXlsx(5, 8, "罗彬")
        else:
            self.WriteXlsx(5, 7, "Error")
            self.WriteXlsx(5, 8, "罗彬")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()#弹框的确定按钮
        sleep(1)
        # self.driver.switch_to.default_content()#切出frame
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//div[@class='btnf']/p[2]").click()#确定按钮
        sleep(1)
        # self.driver.switch_to.window(self.driver.window_handles[1])
        # self.driver.switch_to.frame("absolute iframe_shadow")
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg2)
        if msg2 =="服务器数据保存成功!":
            self.WriteXlsx(6, 7, "Pass")
            self.WriteXlsx(6, 8, "罗彬")
        else:
            self.WriteXlsx(6, 7, "Error")
            self.WriteXlsx(6, 8, "罗彬")
        sleep(1)
        self.driver.find_element_by_xpath("//div/span[@class='right padding_t_10 padding_r_10']").click()
        print("test_Log_4_服务器配置_U8_OK")

    # @unittest.skip("直接跳过")
    def test_Log_e_服务器配置_OA(self):
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        sleep(1)
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()#点击新建
        sleep(1)
        self.driver.switch_to.default_content()#切换出来
        self.driver.switch_to.frame("layui-layer-iframe1")
        sleep(1)
        sel = self.driver.find_element_by_xpath("//select[@id='os_acctid']")#搜索结果显示条数
        Select(sel).select_by_value("002")
        sleep(1)
        sel1 = self.driver.find_element_by_xpath("//select[@id='os_acctyear']")  # 搜索结果显示条数
        Select(sel1).select_by_value("2019")
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[3]/input").send_keys("192.168.20.172")#服务器地址
        sleep(1)
        self.driver.find_element_by_xpath("//form[@id='mytable']/div[4]/input").send_keys("8082")#端口
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@id='layui-layer1']/div[@class='layui-layer-btn layui-layer-btn-']/a").click()#保存
        # self.driver.switch_to.frame("absolute iframe_shadow")
        msg3 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg3)
        if msg3 =="保存成功！":
            self.WriteXlsx(7, 7, "Pass")
            self.WriteXlsx(7, 8, "罗彬")
            # self.driver.find_element_by_xpath("//div[@class='dialog_main absolute']/span[@class='right padding_t_10 padding_r_10']").click()#确定
        elif msg3=="该账套对应的年度已被绑定！":
            self.WriteXlsx(8, 7, "Pass")
            self.WriteXlsx(8, 8, "罗彬")
        else:
            self.WriteXlsx(8, 7, "Error")
            self.WriteXlsx(8, 8, "罗彬")
        print("test_Log_5_服务器配置_OA_OK")

    # @unittest.skip("直接跳过")
    def test_Log_f_服务器配置_删除(self):
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title = '服务器配置']").click()
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='btn2']").click()
        print("进入OA服务器配置")
        self.driver.switch_to.frame("templateSysMgr")
        self.driver.find_element_by_xpath("//input[@row='0']").click()#勾选.
        sleep(2)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()#删除按钮
        self.driver.switch_to.default_content()  # 切换出来
        msg4 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg4)
        if msg4=='是否删除该单据绑定数据!':
            self.driver.find_element_by_xpath("//div[@class='dialog_main_footer left align_right w100b']/span/a[1]").click()
            msg5=self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
            print(msg5)
            if msg5=='删除成功！':
                self.WriteXlsx(9, 7, "Pass")
                self.WriteXlsx(9, 8, "罗彬")
            else:
                self.WriteXlsx(9, 7, "Error")
                self.WriteXlsx(9, 8, "罗彬")
        else:
            self.WriteXlsx(9, 7, "提示信息有误")
            self.WriteXlsx(9, 8, "罗彬")
    # @unittest.skip("直接跳过")
    def test_Log_g_Peoole_Set(self):
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title='人员对照']").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//a[@id='autobtn']").click()#按用户名自动绑定
        sleep(1)
        self.driver.switch_to.default_content()
        msg6 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg6)
        if msg6 == '当前选择【单位A】将会按照【单位A】下OA用户名与ERP用户名一致，进行自动绑定；请确认已绑定人员是否更新?':
            self.driver.find_element_by_xpath("//a[@id='btn1_btn']").click()
            msg7 = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']").text
            print(msg7)
            if msg7 == '成功自动绑定1条数据！':
                self.WriteXlsx(10, 7, "Pass")
                self.WriteXlsx(10, 8, "罗彬")
            else:
                self.WriteXlsx(10, 7, "绑定数据有误，请检查！")
                self.WriteXlsx(10, 8, "罗彬")
            self.driver.find_element_by_xpath("//a[@class='common_button common_button_emphasize margin_r_10 hand']").click()
        # sleep(1)
        # self.driver.switch_to.frame("myiframe")
        # self.driver.find_element_by_xpath("//div[@id='north']/a[3]").click()#全部清除
        # self.driver.switch_to.default_content()#切换到最外层
        # msg8=self.driver.find_element_by_xpath("//div[@class='dialog_main_body left']").text
        # print("msg8:  ",msg8)
        # if msg8=="将清空所有绑定的人员信息，是否继续":
        #     self.driver.find_element_by_xpath("//a[@class='common_button common_button_emphasize margin_r_10 hand']").click()
        #     msg9 =self.driver.find_element_by_xpath("//div[@class='dialog_main_body left']").text
        #     print("msg9:  ",msg9)
        #     sleep(1)
        #     if msg9 =="清空成功":
        #         self.WriteXlsx(12, 7, "Pass")
        #         self.WriteXlsx(12, 8, "罗彬")
        #     else:
        #         self.WriteXlsx(12, 7, "Error")
        #         self.WriteXlsx(12, 8, "罗彬")

    # @unittest.skip("直接跳过")
    def test_Log_h_Project_Set_Add(self):#科目配置
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title='科目配置']").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//a[@id='ClkAdd1']").click()#新增
        sleep(1)
        self.driver.switch_to.default_content()
        Project_Set= self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe1']")
        self.driver.switch_to.frame(Project_Set)  # 先进行定位在切换iframe
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='aliasname']").click()#账套信息
        sleep(1)
        self.driver.switch_to.default_content()
        Project_Set_ZT = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe2']")#账套
        self.driver.switch_to.frame(Project_Set_ZT)  # 先进行定位在切换iframe
        sleep(1)
        self.driver.find_element_by_xpath("//tbody[@id='list']/tr[3]/td/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()
        print("点击确定按钮")
        sleep(1)
        # self.driver.switch_to.parent_frame()  # 切换回去
        # Project_Set1 = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe1']")
        self.driver.switch_to.frame(Project_Set)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys("财务集成")
        sleep(1)
        self.driver.find_element_by_xpath("//input[@id='deptnm']").click()#部门名称
        sleep(1)
        self.driver.switch_to.default_content()
        Project_Set_Dev = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe3']")
        self.driver.switch_to.frame(Project_Set_Dev)  # 先进行定位在切换iframe  部门
        sleep(1)
        self.driver.find_element_by_xpath("//ul[@id='myTree_1_ul']/li[3]/span[2]").click()#选择TEST_two
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()
        sleep(1)
        self.driver.switch_to.frame(Project_Set)
        self.driver.find_element_by_xpath("//input[@id='financekmnm']").click()#财务科目
        sleep(1)
        self.driver.switch_to.default_content()
        Project_Set_KM = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe4']")  # 科目信息
        self.driver.switch_to.frame(Project_Set_KM)  # 先进行定位在切换iframe
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id = 'mytree_78_a']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@row='3']").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@id='layui-layer4']/div[3]/a[1]").click()
        self.driver.switch_to.frame(Project_Set)
        self.driver.find_element_by_xpath("//input[@id='taxkmnm']").click()#点税金科目
        sleep(1)
        self.driver.switch_to.default_content()
        Project_Set_SJ = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe5']")
        self.driver.switch_to.frame(Project_Set_SJ)  # 先进行定位在切换iframe  税金科目
        self.driver.find_element_by_xpath("//span[@id='mytree_147_span']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@row='4']").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@id='layui-layer5']/div[3]/a[1]").click()
        self.driver.switch_to.frame(Project_Set)
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()
        msg = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg)
        if msg=="保存成功!":
            self.WriteXlsx(13, 7, "Pass")
            self.WriteXlsx(13, 8, "罗彬")
        else:
            self.WriteXlsx(13, 7, "Error")
            self.WriteXlsx(13, 8, "罗彬")
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()

        #进行修改操作
        sleep(5)
        self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//input[@row='0']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkEdit1']").click()  # 修改
        sleep(1)
        self.driver.switch_to.default_content()
        Project_Set = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe6']")
        self.driver.switch_to.frame(Project_Set)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").clear()
        sleep(1)
        self.driver.find_element_by_xpath("//tr[@id='ConfigTR2']/td/input[2]").send_keys("财务集成-付款100W")
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#保存
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']").click()#弹窗提示消息
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        msg = self.driver.find_element_by_xpath("//tbody[@id='list']/tr[1]/td[5]").text
        print(msg)
        if msg=="财务集成-付款100W":
            self.WriteXlsx(14, 7, "Pass")
            self.WriteXlsx(14, 8, "罗彬")
        else:
            self.WriteXlsx(14, 7, "Error")
            self.WriteXlsx(14, 8, "罗彬")

        # 进行删除操作

        sleep(5)
        # self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//input[@row='0']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='ClkDel1']").click()  # 删除
        sleep(1)
        self.driver.switch_to.default_content()
        msg = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='dialog_main absolute']/div[3]/span/a[1]").click()#确认删除
        sleep(1)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg2)
        self.driver.find_element_by_xpath("//div[@class='dialog_main_footer left align_right w100b']/span/a").click()
        if msg == '数据删除后将无法恢复,请确认是否进行删除!':
            if msg2 == "删除成功!":
                self.WriteXlsx(15, 7, "Pass")
                self.WriteXlsx(15, 8, "罗彬")
            else:
                self.WriteXlsx(15, 7, "Error")
                self.WriteXlsx(15, 8, "罗彬")



    def test_Log_i_Assist(self):#辅助核算
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title='辅助核算']").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        self.driver.switch_to.frame("iframeURL")
        self.driver.find_element_by_xpath("//table[@id='mytable']/tbody/tr[1]/td/div/input").click()
        # sleep(1)
        # self.driver.find_element_by_xpath("//a[@id='autobtn']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//input[@row='7']").click()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@id='addbtn']").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg = self.driver.find_element_by_xpath("//div[@class='dialog_main_content_html ']/table/tbody/tr/td[2]").text
        print(msg)
        if msg =='人员绑定成功！':
            self.WriteXlsx(16, 7, "Pass")
            self.WriteXlsx(16, 8, "罗彬")
        else:
            self.WriteXlsx(16, 7, "Error")
            self.WriteXlsx(16, 8, "罗彬")

    def test_Log_j_SQ(self):
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title='账套授权']").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//a[@class='common_button common_button_icon']").click()#新增
        sleep(1)
        self.driver.switch_to.default_content()
        New_SQ = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe1']")
        self.driver.switch_to.frame(New_SQ)  # 先进行定位在切换iframe
        self.driver.find_element_by_xpath("//input[@id='acctname']").click()#选择账套
        sleep(1)
        self.driver.switch_to.default_content()
        New_SQ1 = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe2']")
        self.driver.switch_to.frame(New_SQ1)
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='MyTable']/tbody/tr[3]/td[1]/div/input").click()
        sleep(1)
        self.driver.switch_to.default_content()  # 切换出来
        self.driver.find_element_by_xpath("//div[@id='layui-layer2']/div[3]/a[1]").click()
        # self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()
        self.driver.switch_to.frame(New_SQ)
        self.driver.find_element_by_xpath("//input[@id='setname']").send_keys("新增一个账套为星空演示001的账套")
        sleep(1)
        Sel = self.driver.find_element_by_xpath("//select[@id='formselect']")#授权表单
        Select(Sel).select_by_value("-1")
        sleep(1)
        Sel_1 = self.driver.find_element_by_xpath("//select[@id='personnelselect']")  # 授权人员
        Select(Sel_1).select_by_value("0")
        sleep(1)
        Sel_2 = self.driver.find_element_by_xpath("//select[@id='departmentselect']")  # 授权部门
        Select(Sel_2).select_by_value("1")
        sleep(1)
        self.driver.find_element_by_xpath("//textarea[@id='department']").click()
        self.driver.switch_to.default_content()
        New_SQ_2 = self.driver.find_element_by_css_selector("iframe[id^='layui-layer-iframe3']")
        self.driver.switch_to.frame(New_SQ_2)
        sleep(1)
        self.driver.find_element_by_xpath("//span[@id='myTree_1_check']").click()
        sleep(1)
        self.driver.switch_to.default_content()
        # self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()
        self.driver.find_element_by_xpath("//div[@id='layui-layer3']/div[3]/a[1]").click()
        self.driver.switch_to.frame(New_SQ)
        sleep(1)
        self.driver.switch_to.default_content()
        self.driver.find_element_by_xpath("//div[@class='layui-layer-btn layui-layer-btn-']/a[1]").click()#保存
        sleep(1)
        msg = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg)
        if msg =='保存成功！':
            self.WriteXlsx(19, 7, "Pass")
            self.WriteXlsx(19, 8, "罗彬")
        else:
            self.WriteXlsx(19, 7, "Error")
            self.WriteXlsx(19, 8, "罗彬")
        # self.driver.switch_to.frame("myiframe")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()

    def test_Log_k_SQ_Del(self):
        self.System_Config()
        self.driver.find_element_by_xpath("//div[@title='账套授权']").click()
        sleep(1)
        self.driver.switch_to.frame("myiframe")
        sleep(1)
        self.driver.find_element_by_xpath("//table[@id='accttable']/tbody[@id='list']/tr[1]/td[1]/div").click()
        sleep(1)
        self.driver.find_element_by_xpath("//a[@href='javascript:del()']").click()#删除按钮
        sleep(1)
        self.driver.switch_to.default_content()
        msg1 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg1)#是否确认删除!
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a[1]").click()
        sleep(1)
        self.driver.switch_to.default_content()
        msg2 = self.driver.find_element_by_xpath("//td[@class='msgbox_content padding_l_10']").text
        print(msg2)#删除成功！
        if msg1 == "是否确认删除!":
            if msg2 == "删除成功！":
                self.WriteXlsx(20, 7, "Pass")
                self.WriteXlsx(20, 8, "罗彬")
            else:
                self.WriteXlsx(20, 7, "Error")
                self.WriteXlsx(20, 8, "罗彬")
        self.driver.find_element_by_xpath("//span[@class='right padding_t_10 padding_r_10']/a").click()


if __name__ == '__main__':
    unittest.main(verbosity=2)
